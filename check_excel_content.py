# -*- coding: utf-8 -*-
"""Проверка содержимого Excel файла"""
import openpyxl

filename = 'test_exports/test_export_kremenchugskaya_2025-11-17.xlsx'

print("="*80)
print("ПРОВЕРКА СОДЕРЖИМОГО EXCEL ФАЙЛА")
print("="*80)
print(f"\nФайл: {filename}\n")

try:
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    print(f"Название листа: {ws.title}")
    print(f"Количество строк: {ws.max_row}")
    print(f"Количество столбцов: {ws.max_column}")
    print("\n" + "-"*80)
    print("СОДЕРЖИМОЕ:")
    print("-"*80)

    # Читаем все строки
    for row_idx in range(1, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data.append(str(cell_value))
            else:
                row_data.append("")

        if any(row_data):  # Печатаем только непустые строки
            print(f"Строка {row_idx}: {' | '.join(row_data)}")

    print("\n" + "="*80)
    print("ПРОВЕРКА ЗАВЕРШЕНА")
    print("="*80)

except Exception as e:
    print(f"[ERROR] Ошибка при чтении файла: {e}")
    import traceback
    traceback.print_exc()
