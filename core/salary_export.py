"""
Экспорт расчёта ЗП (/salary) в .xlsx — рендерер раскладки в openpyxl.

Строки, формулы и оформление берутся из `core/salary_layout.py` (единый
источник для .xlsx и Google Таблицы — см. его докстроку: формат листа,
формулы, две сверки «формула не должна соврать»). Здесь только перенос
раскладки в книгу openpyxl: шрифт, заливки, рамки, ширины, форматы чисел.

ВАЖНО про формулы: openpyxl не пишет кэш вычисленных значений, поэтому
просмотрщики, которые сами не считают (предпросмотр Google Drive, защищённый
режим Excel), показывают формульные ячейки пустыми — жалоба владельца
2026-07-31, из-за неё формулы временно убирали. Смягчение: у книги выставлен
`fullCalcOnLoad` — Excel/LibreOffice/Google Sheets пересчитывают лист при
ОТКРЫТИИ файла, и значения появляются. В предпросмотре без открытия
формульные ячейки по-прежнему могут быть пустыми — это цена живых формул.
У Google-экспорта (`core/salary_gsheet.py`) этой проблемы нет: Таблицы считают
формулы на сервере.

Payload (все суммы в рублях, уже рассчитаны страницей):
    {
      "month": "YYYY-MM",
      "kpi_names": ["Доля кухни (%)", ...],      # показатели месяца, по порядку
      "base_per_kpi": 5000,                       # тариф KPI = фонд / кол-во KPI
      "roles": [{"name": "бармен", "rate": 300}, ...],  # порядок = порядок строк
      "employees": [{
          "name": str,
          "hours_by_role": {role_name: часы},
          "pay_by_role": {role_name: оплата},     # для сверки; в файл идёт формула
          "shifts_count": int,                    # дневные смены графика (база такси)
          "handover_bonus": float,                # премия за передачу смены
          "handover_paid_days": int,              # оплаченных дней передачи кассы —
                                                  # база строки премии (дни смен минус
                                                  # дни без кассы и ручные штрафы)
          "day_plan_bonus": float,                # премия за дневной план
          "kpi_premiums": [float, ...],           # по каждому KPI, уже с koef
          "late_penalty": float,                  # авто-штраф за опоздания
          "adjustments": {категория: сумма}       # vacation/extra_income/deduction_*;
                                                  # с 2026-07-31 фронт шлёт {} —
                                                  # эти строки заполняются в Excel
      }, ...]
    }
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# HANDOVER_RATE/DAY_PLAN_RATE/TAXI_* и month_title/sheet_title импортируются
# для реэкспорта: раскладка переехала в salary_layout, но эти имена уже
# используются как `from core.salary_export import ...` (тесты, routes).
from core.salary_layout import (DAY_PLAN_RATE, FILL_HEADER, FILL_RED_LABEL,  # noqa: F401
                                FILL_TOTAL, FIRST_DATA_ROW, FMT_MONEY, FONT_NAME,
                                FONT_SIZE, HANDOVER_RATE, HEADER_ROW, HEADERS,
                                TAXI_OFFICIAL_SHIFTS, TAXI_RATE_PER_SHIFT,
                                WIDTH_EMP, WIDTH_LABEL, WIDTH_META, WIDTH_TOTAL,
                                Formula, build_sheet, month_title, sheet_title)

_THIN = Side(style='thin', color='BFBFBF')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FILL_CACHE = {}


def _fill(hex_color):
    """PatternFill по hex-цвету раскладки (кешируется — заливок мало)."""
    if not hex_color:
        return None
    if hex_color not in _FILL_CACHE:
        _FILL_CACHE[hex_color] = PatternFill('solid', fgColor=hex_color)
    return _FILL_CACHE[hex_color]


def _font(bold=False, italic=False, color=None):
    """Шрифт эталона (PT Serif 8) с нужными начертаниями."""
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=bold, italic=italic, color=color)


def build_salary_workbook(payload: dict) -> Workbook:
    """Собрать книгу Excel из payload страницы /salary (контракт — в докстроке)."""
    sheet = build_sheet(payload)

    wb = Workbook()
    # Формулы в файле живые, но openpyxl не пишет их вычисленные значения:
    # без этого флага лист открылся бы с пустыми формульными ячейками.
    wb.calculation.fullCalcOnLoad = True
    # Шрифт по умолчанию — для ячеек, которых рендерер не касается (приватный
    # API openpyxl, поэтому под try: без него файл всё равно корректен —
    # каждая записанная ячейка получает шрифт явно).
    try:
        wb._named_styles['Normal'].font = _font()
    except Exception:
        pass

    ws = wb.active
    ws.title = sheet.title
    emp_cols = sheet.emp_cols
    total_col = sheet.total_col

    def put(row, col, value, fill=None, font=None, fmt=None, align=None):
        cell = ws.cell(row=row, column=col, value=value)
        # openpyxl типизирует любую строку с ведущим «=» как формулу: имя
        # сотрудника вида «=Иванов» ушло бы в файл живой формулой. Формулы
        # раскладки помечены Formula, всё остальное с «=» — принудительно текст.
        if isinstance(value, str) and value.startswith('=') and not isinstance(value, Formula):
            cell.data_type = 's'
        cell.border = _BORDER
        cell.font = font or _font()
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        if align:
            cell.alignment = align
        return cell

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- Строка 1 пустая, строка 2 — шапка (как в таблице бухгалтерии) ---
    put(HEADER_ROW, 1, HEADERS[0], font=_font(italic=True), align=center)
    for i, h in enumerate(HEADERS[1:], start=2):
        put(HEADER_ROW, i, h, font=_font(bold=True, italic=True), align=center)
    for i, name in enumerate(sheet.employees):
        put(HEADER_ROW, sheet.first_emp_col + i, name, fill=_fill(FILL_HEADER),
            font=_font(bold=True), align=center)
    put(HEADER_ROW, total_col, 'ИТОГО', fill=_fill(FILL_HEADER),
        font=_font(bold=True), align=center)

    # --- Данные ---
    for r in sheet.rows:
        meta_fill = _fill(r.meta_fill)
        put(r.number, 1, r.label, fill=_fill(r.label_fill),
            font=_font(bold=r.label_bold, italic=r.label_italic,
                       # тёмно-красная плашка вычетов — подпись белым
                       color='FFFFFF' if r.label_fill == FILL_RED_LABEL else None))
        owner, deadline, checker = r.meta
        put(r.number, 2, owner or None, fill=_fill(r.owner_fill) or meta_fill, align=center)
        put(r.number, 3, deadline or None, fill=meta_fill,
            font=_font(italic=True), align=center)
        put(r.number, 4, checker or None, fill=meta_fill, align=center)
        put(r.number, 5, r.tariff, fill=meta_fill, font=_font(italic=True),
            fmt=FMT_MONEY, align=center)
        for i, col in enumerate(emp_cols):
            put(r.number, col, r.cells[i] if i < len(r.cells) else None,
                fill=_fill(r.data_fill), fmt=r.fmt)
        put(r.number, total_col, r.total, fill=_fill(FILL_TOTAL),
            font=_font(bold=True), fmt=r.fmt)

    # --- Оформление листа (ширины — как в таблице бухгалтерии) ---
    ws.column_dimensions['A'].width = WIDTH_LABEL
    for i, w in enumerate(WIDTH_META, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in emp_cols:
        ws.column_dimensions[get_column_letter(col)].width = WIDTH_EMP
    ws.column_dimensions[get_column_letter(total_col)].width = WIDTH_TOTAL
    ws.freeze_panes = ws.cell(row=FIRST_DATA_ROW, column=sheet.first_emp_col)

    return wb
