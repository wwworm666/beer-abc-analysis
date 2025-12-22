"""Проверка вычисленных значений списаний баллов"""
import openpyxl

# Открываем с data_only=True чтобы увидеть вычисленные значения
wb = openpyxl.load_workbook('планы_2025_2026.xlsx', data_only=True)
ws = wb.active

print("Ноябрь 2025 - Выручка (строка 131):")
print(f"  Большой (B131): {ws['B131'].value}")
print(f"  Лиговский (C131): {ws['C131'].value}")
print(f"  Кременчугская (D131): {ws['D131'].value}")
print(f"  Варшавская (E131): {ws['E131'].value}")

print("\nНоябрь 2025 - Списания баллов (строка 132, 5% от выручки):")
b132 = ws['B132'].value
c132 = ws['C132'].value
d132 = ws['D132'].value
e132 = ws['E132'].value

if b132: print(f"  Большой (B132): {b132:.2f}")
if c132: print(f"  Лиговский (C132): {c132:.2f}")
if d132: print(f"  Кременчугская (D132): {d132:.2f}")
if e132: print(f"  Варшавская (E132): {e132:.2f}")

print("\n--- Проверка формул (без data_only) ---")
wb2 = openpyxl.load_workbook('планы_2025_2026.xlsx')
ws2 = wb2.active

print("\nФормулы в строке 132:")
print(f"  B132: {ws2['B132'].value}")
print(f"  C132: {ws2['C132'].value}")
print(f"  D132: {ws2['D132'].value}")
print(f"  E132: {ws2['E132'].value}")
