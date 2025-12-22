"""Проверка добавленных строк списаний баллов"""
import openpyxl

wb = openpyxl.load_workbook('планы_2025_2026.xlsx')
ws = wb.active

print("Ноябрь 2025:")
for r in range(128, 142):
    row_label = ws.cell(r, 1).value or ""
    b_val = ws.cell(r, 2).value
    print(f"  {r}: {str(row_label):20s} | B{r}: {b_val}")

print("\nДекабрь 2025:")
for r in range(143, 157):
    row_label = ws.cell(r, 1).value or ""
    b_val = ws.cell(r, 2).value
    print(f"  {r}: {str(row_label):20s} | B{r}: {b_val}")
