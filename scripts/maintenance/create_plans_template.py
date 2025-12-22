import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar

# Создаем новую книгу
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Планы 2025-2026"

# Список баров
bars = [
    'Большой пр. В.О',
    'Лиговский',
    'Кременчугская',
    'Варшавская',
    'Общая'
]

# Список метрик (в том порядке, в котором они должны идти)
metrics = [
    'Выручка',
    'Прибыль',
    'Чеки',
    'Средний чек',
    'Доля розлив',
    'Доля фасовка',
    'Доля кухня',
    'Наценка',
    'Наценка розлив',
    'Наценка фасовка',
    'Наценка кухня'
]

# Генерируем список всех месяцев 2025 и 2026
months = []
for year in [2025, 2026]:
    for month in range(1, 13):
        month_name = calendar.month_name[month]
        # Русские названия месяцев
        month_names_ru = {
            'January': 'Январь', 'February': 'Февраль', 'March': 'Март',
            'April': 'Апрель', 'May': 'Май', 'June': 'Июнь',
            'July': 'Июль', 'August': 'Август', 'September': 'Сентябрь',
            'October': 'Октябрь', 'November': 'Ноябрь', 'December': 'Декабрь'
        }
        month_ru = month_names_ru[month_name]
        months.append(f"{month_ru} {year}")

# Стили
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
bar_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
bar_font = Font(bold=True, color="FFFFFF", size=11)
metric_font = Font(size=10)
center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Текущая строка
current_row = 1

# Заголовок таблицы
ws.merge_cells(f'A{current_row}:C{current_row}')
header_cell = ws.cell(current_row, 1)
header_cell.value = "ПЛАНЫ НА 2025-2026 ГОД"
header_cell.font = header_fill
header_cell.fill = header_fill
header_cell.alignment = center_alignment
current_row += 1

# Пустая строка
current_row += 1

# Для каждого месяца создаем блок
for month in months:
    # Заголовок месяца
    ws.merge_cells(f'A{current_row}:C{current_row}')
    month_cell = ws.cell(current_row, 1)
    month_cell.value = month
    month_cell.font = Font(bold=True, size=14, color="FFFFFF")
    month_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    month_cell.alignment = center_alignment
    current_row += 1

    # Заголовки столбцов: Метрика | Бар1 | Бар2 | Бар3 | Бар4
    ws.cell(current_row, 1).value = "Метрика"
    ws.cell(current_row, 1).font = header_font
    ws.cell(current_row, 1).fill = header_fill
    ws.cell(current_row, 1).alignment = left_alignment
    ws.cell(current_row, 1).border = thin_border

    for i, bar in enumerate(bars, start=2):
        cell = ws.cell(current_row, i)
        cell.value = bar
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    current_row += 1

    # Строки с метриками
    for metric in metrics:
        ws.cell(current_row, 1).value = metric
        ws.cell(current_row, 1).font = metric_font
        ws.cell(current_row, 1).alignment = left_alignment
        ws.cell(current_row, 1).border = thin_border

        # Пустые ячейки для значений по каждому бару
        for col in range(2, 6):  # Колонки 2-5 для 4 баров
            cell = ws.cell(current_row, col)
            cell.value = ""  # Пустая ячейка для ввода
            cell.alignment = center_alignment
            cell.border = thin_border

        current_row += 1

    # Пустая строка после блока месяца
    current_row += 1

# Настройка ширины столбцов
ws.column_dimensions['A'].width = 30  # Метрика
ws.column_dimensions['B'].width = 20  # Большой пр. В.О
ws.column_dimensions['C'].width = 18  # Лиговский
ws.column_dimensions['D'].width = 18  # Кременчугская
ws.column_dimensions['E'].width = 18  # Варшавская
ws.column_dimensions['F'].width = 18  # Общая

# Закрепляем первую строку
ws.freeze_panes = 'A4'

# Сохраняем файл
wb.save('планы_2025_2026.xlsx')
print("Таблица планов на 2025-2026 создана: планы_2025_2026.xlsx")
print(f"Месяцев: {len(months)}")
print(f"Баров: {len(bars)}")
print(f"Метрик на месяц: {len(metrics)}")
