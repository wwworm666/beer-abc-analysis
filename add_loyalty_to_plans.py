"""
Добавить строку "Списания баллов" в таблицу планов = 5% от выручки
"""
import openpyxl

# Открываем файл
wb = openpyxl.load_workbook('планы_2025_2026.xlsx')
ws = wb.active

# Находим все блоки с данными (где есть "Выручка")
# Добавляем строку "Списания баллов" сразу после "Выручка"

for row_idx in range(1, ws.max_row + 1):
    cell_value = ws.cell(row_idx, 1).value

    # Если нашли строку "Выручка"
    if cell_value and str(cell_value).strip() == 'Выручка':
        # Проверяем есть ли данные (выручка не пустая)
        revenue_cells = [ws.cell(row_idx, col).value for col in range(2, 6)]

        if any(revenue_cells):  # Если хоть одна ячейка не пустая
            # Проверяем, нет ли уже строки "Списания баллов" после этой строки
            next_cell = ws.cell(row_idx + 1, 1).value

            if next_cell and 'списан' in str(next_cell).lower():
                print(f"Строка {row_idx}: Списания баллов уже есть, обновляем формулы...")
                # Обновляем формулы
                for col in range(2, 6):
                    revenue_value = ws.cell(row_idx, col).value
                    if revenue_value:
                        # Ставим формулу = 5% от выручки
                        ws.cell(row_idx + 1, col).value = f"={ws.cell(row_idx, col).coordinate}*0.05"
            else:
                print(f"Строка {row_idx}: Добавляем строку 'Списания баллов'...")
                # Вставляем новую строку после Выручки
                ws.insert_rows(row_idx + 1)

                # Заполняем название
                ws.cell(row_idx + 1, 1).value = "Списания баллов"

                # Добавляем формулы = 5% от выручки для каждого бара
                for col in range(2, 6):
                    revenue_value = ws.cell(row_idx, col).value
                    if revenue_value:
                        # Ставим формулу = 5% от выручки
                        ws.cell(row_idx + 1, col).value = f"={ws.cell(row_idx, col).coordinate}*0.05"

# Сохраняем файл
wb.save('планы_2025_2026.xlsx')
print("\nГотово! Файл сохранён.")
