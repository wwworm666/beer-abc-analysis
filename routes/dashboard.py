from flask import Blueprint, request, jsonify
import time
import json
import os
from datetime import datetime, timedelta
from core.olap_reports import OlapReports
from core.dashboard_analysis import DashboardMetrics
from core.weeks_generator import WeeksGenerator
from extensions import (
    venues_manager, plans_manager, notes_manager, taps_manager,
    comparison_calculator, trends_analyzer, export_manager,
    DASHBOARD_OLAP_CACHE, DASHBOARD_OLAP_CACHE_TTL
)

dashboard_bp = Blueprint('dashboard', __name__)


@dashboard_bp.route('/api/dashboard-analytics', methods=['POST'])
def dashboard_analytics():
    """API endpoint для получения всех метрик дашборда из сырых OLAP данных"""
    try:
        data = request.json
        venue_key = data.get('bar')  # ключ заведения (kremenchugskaya, bolshoy, etc)
        date_from = data.get('date_from')  # YYYY-MM-DD
        date_to = data.get('date_to')      # YYYY-MM-DD

        if not date_from or not date_to:
            return jsonify({'error': 'Требуются параметры date_from и date_to'}), 400

        # Преобразуем venue_key в название для iiko API
        bar_name = venues_manager.get_iiko_name(venue_key) if venue_key and venue_key != 'all' else None

        # iiko API: date_to не включается (exclusive end)
        # Чтобы получить данные за 10.11-16.11 включительно, нужно запросить до 17.11
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
        date_to_inclusive = (date_to_obj + timedelta(days=1)).strftime('%Y-%m-%d')

        print(f"\n[DASHBOARD] Запуск анализа дашборда...")
        print(f"   Venue key: {venue_key}")
        print(f"   Bar name (iiko): {bar_name if bar_name else 'ВСЕ'}")
        print(f"   Period (requested): {date_from} - {date_to}")
        print(f"   Period (iiko API): {date_from} - {date_to_inclusive}")

        # Проверяем кеш OLAP данных
        cache_key = f"{venue_key}_{date_from}_{date_to_inclusive}"
        now = time.time()

        if cache_key in DASHBOARD_OLAP_CACHE:
            cached_entry = DASHBOARD_OLAP_CACHE[cache_key]
            if (now - cached_entry['timestamp']) < DASHBOARD_OLAP_CACHE_TTL:
                all_sales_data = cached_entry['data']
                ttl_remaining = DASHBOARD_OLAP_CACHE_TTL - (now - cached_entry['timestamp'])
                print(f"   [CACHE] Использую кешированные OLAP данные (истекает через {int(ttl_remaining // 60)} мин {int(ttl_remaining % 60)} сек)")
            else:
                # Кеш устарел, удаляем
                del DASHBOARD_OLAP_CACHE[cache_key]
                all_sales_data = None
        else:
            all_sales_data = None

        # Если данных нет в кеше - запрашиваем из iiko API
        if not all_sales_data:
            # Подключаемся к iiko API
            olap = OlapReports()
            if not olap.connect():
                return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

            # 1. Получаем ВСЕ данные ОДНИМ запросом (оптимизация: 1 запрос вместо 3)
            print("   [1/5] Запуск комплексного OLAP запроса...")
            start_time = time.time()

            all_sales_data = olap.get_all_sales_report(date_from, date_to_inclusive, bar_name)

            if not all_sales_data or not all_sales_data.get('data'):
                olap.disconnect()
                return jsonify({'error': 'Не удалось получить данные из OLAP'}), 500

            elapsed = time.time() - start_time
            print(f"   [OK] Комплексный запрос выполнен за {elapsed:.2f}s")

            # Отключаемся от iiko API
            olap.disconnect()

            # Сохраняем в кеш
            DASHBOARD_OLAP_CACHE[cache_key] = {
                'data': all_sales_data,
                'timestamp': now
            }
            print(f"   [CACHE] Данные закешированы на {DASHBOARD_OLAP_CACHE_TTL // 60} минут")

        # 2. Создаем калькулятор метрик и рассчитываем из СЫРЫХ данных
        print("   [2/5] Расчет метрик из сырых OLAP данных...")
        calculator = DashboardMetrics()

        # Считаем метрики из единого отчета
        metrics = calculator.calculate_metrics(all_sales_data)
        table_data = calculator.get_table_data(metrics)

        print(f"   [OK] Метрики рассчитаны успешно!")

        # 3. Добавляем метрику активности кранов за период
        print("   [3/5] Расчет активности кранов...")

        # Маппинг venue_key -> bar_id для TapsManager
        VENUE_TO_BAR_MAPPING = {
            'bolshoy': 'bar1',
            'ligovskiy': 'bar2',
            'kremenchugskaya': 'bar3',
            'varshavskaya': 'bar4',
            'all': None  # Для "all" передаем None чтобы считать все бары
        }

        # Преобразуем venue_key в bar_id для TapsManager
        bar_id_for_taps = VENUE_TO_BAR_MAPPING.get(venue_key)
        print(f"   [DEBUG] venue_key={venue_key} -> bar_id_for_taps={bar_id_for_taps}")

        tap_activity = taps_manager.calculate_tap_activity_for_period(bar_id_for_taps, date_from, date_to)
        metrics['tap_activity'] = tap_activity
        print(f"   [OK] Активность кранов: {tap_activity}%")

        # Преобразуем ключи для совместимости с фронтендом
        # Фронтенд ожидает camelCase, бэкенд возвращает snake_case
        frontend_mapping = {
            'total_revenue': 'revenue',
            'total_checks': 'checks',
            'avg_check': 'averageCheck',
            'draft_share': 'draftShare',
            'bottles_share': 'packagedShare',
            'kitchen_share': 'kitchenShare',
            'draft_revenue': 'revenueDraft',
            'bottles_revenue': 'revenuePackaged',
            'kitchen_revenue': 'revenueKitchen',
            'avg_markup': 'markupPercent',
            'total_margin': 'profit',
            'draft_markup': 'markupDraft',
            'bottles_markup': 'markupPackaged',
            'tap_activity': 'tapActivity',
            'kitchen_markup': 'markupKitchen',
            'loyalty_points_written_off': 'loyaltyWriteoffs'
        }

        # Применяем маппинг
        mapped_metrics = {}
        for old_key, new_key in frontend_mapping.items():
            if old_key in metrics:
                value = metrics[old_key]

                # Наценка в API приходит как дробное число (1.95), а в планах как проценты (195)
                # Умножаем на 100 для единообразия
                if new_key in ['markupPercent', 'markupDraft', 'markupPackaged', 'markupKitchen']:
                    value = value * 100

                mapped_metrics[new_key] = value

        # Формируем ответ с преобразованными ключами
        response = {
            **mapped_metrics,
            'table_data': table_data
        }

        return jsonify(response)

    except Exception as e:
        print(f"[ERROR] Ошибка в /api/dashboard-analytics: {e}")
        import traceback
        traceback.print_exc()
        error_detail = f"{type(e).__name__}: {str(e)}"
        return jsonify({'error': error_detail}), 500


# ============================================================================
# DASHBOARD PLANS API - Управление плановыми показателями
# ============================================================================

@dashboard_bp.route('/api/venues')
def get_venues():
    """Получить список всех заведений для селектора"""
    try:
        print("\n[VENUES API] Запрос списка заведений...")
        venues = venues_manager.get_all_for_dropdown()
        print(f"[VENUES API] Возвращаю {len(venues)} заведений")
        return jsonify({'venues': venues})
    except Exception as e:
        print(f"[VENUES API ERROR] Ошибка при получении списка заведений: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/weeks')
def get_weeks():
    """Получить список всех недель для текущего года"""
    try:
        print("\n[WEEKS API] Генерация недель для текущего года...")
        current_year = datetime.now().year
        weeks = WeeksGenerator.generate_weeks_for_year(current_year)
        current_week = WeeksGenerator.get_current_week()
        print(f"[WEEKS API] Сгенерировано недель: {len(weeks)}")
        print(f"[WEEKS API] Текущая неделя: {current_week['label']}")
        return jsonify({
            'weeks': weeks,
            'current_week': current_week
        })
    except Exception as e:
        print(f"[WEEKS API ERROR] Ошибка при генерации недель: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/plans/<venue_key>/<period_key>')
def get_plan_with_venue(venue_key, period_key):
    """Получить план за конкретное заведение и период"""
    try:
        print(f"\n[PLANS API] Запрос плана для заведения: {venue_key}, период: {period_key}")
        composite_key = f"{venue_key}_{period_key}"
        print(f"[PLANS API] Составной ключ: {composite_key}")
        plan = plans_manager.get_plan(composite_key)
        if plan:
            print(f"[PLANS API] План найден")
            return jsonify(plan)
        else:
            print(f"[PLANS API] План НЕ найден")
            return jsonify({'error': 'Plan not found'}), 404
    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при получении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/plans/calculate/<venue_key>/<start_date>/<end_date>')
def calculate_plan_for_period(venue_key, start_date, end_date):
    """Рассчитать план для произвольного периода на основе месячных планов"""
    try:
        print(f"\n[PLANS API] Расчёт плана для {venue_key}, период: {start_date} - {end_date}")
        venue = '' if venue_key in ('', 'total', 'общая', 'all') else venue_key
        plan = plans_manager.calculate_plan_for_period(venue, start_date, end_date)
        if plan:
            print(f"[PLANS API] План рассчитан: выручка={plan.get('revenue', 0):.0f}")
            return jsonify(plan)
        else:
            print(f"[PLANS API] Не удалось рассчитать план")
            return jsonify({'error': 'No monthly plans found for this period'}), 404
    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при расчёте плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/plans/migrate-to-monthly', methods=['POST'])
def migrate_plans_to_monthly():
    """Миграция недельных планов в месячные (одноразовая)"""
    try:
        print("\n[PLANS API] Запуск миграции планов в месячный формат...")
        monthly_plans = plans_manager.import_monthly_plans_from_weekly()
        return jsonify({
            'success': True,
            'message': f'Migrated to {len(monthly_plans)} monthly plans',
            'plans_count': len(monthly_plans),
            'plan_keys': list(monthly_plans.keys())
        })
    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка миграции: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/plans/import-from-excel', methods=['POST'])
def import_plans_from_excel():
    """Импорт планов из Excel файла планы_2025_2026.xlsx"""
    try:
        import openpyxl

        print("\n[PLANS API] Запуск импорта планов из Excel...")

        # Маппинг названий баров
        VENUE_MAPPING = {
            'Большой пр. В.О': 'bolshoy',
            'Лиговский': 'ligovskiy',
            'Кременчугская': 'kremenchugskaya',
            'Варшавская': 'varshavskaya',
            'Общая': 'all'
        }

        # Маппинг метрик
        METRIC_MAPPING = {
            'Выручка': 'revenue',
            'Прибыль': 'profit',
            'Чеки': 'checks',
            'Средний чек': 'averageCheck',
            'Доля розлив': 'draftShare',
            'Доля фасовка': 'packagedShare',
            'Доля кухня': 'kitchenShare',
            'Наценка': 'markupPercent',
            'Наценка розлив': 'markupDraft',
            'Наценка фасовка': 'markupPackaged',
            'Наценка кухня': 'markupKitchen',
            'Списания баллов': 'loyaltyWriteoffs',
            'Активность кранов': 'tapActivity'
        }

        MONTH_NAMES_RU = {
            'Январь': 1, 'Февраль': 2, 'Март': 3, 'Апрель': 4,
            'Май': 5, 'Июнь': 6, 'Июль': 7, 'Август': 8,
            'Сентябрь': 9, 'Октябрь': 10, 'Ноябрь': 11, 'Декабрь': 12
        }

        def parse_month_header(header_text):
            parts = header_text.strip().split()
            if len(parts) != 2:
                return None, None
            month_name, year_str = parts
            try:
                year = int(year_str)
            except ValueError:
                return None, None
            month = MONTH_NAMES_RU.get(month_name)
            return (year, month) if month else (None, None)

        # Читаем Excel
        file_path = 'data/reports/планы_2025_2026.xlsx'
        if not os.path.exists(file_path):
            return jsonify({'error': f'Файл не найден: {file_path}'}), 404

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        all_plans = {}
        current_month_year = None
        bar_columns = {}

        for row in ws.iter_rows(min_row=1, values_only=False):
            first_cell = row[0]
            if first_cell.value and isinstance(first_cell.value, str):
                year, month = parse_month_header(first_cell.value)
                if year and month:
                    current_month_year = (year, month)
                    print(f"[EXCEL] Найден месяц: {first_cell.value}")
                    continue

            if first_cell.value == 'Метрика':
                bar_columns = {}
                for col_idx, cell in enumerate(row[1:], start=2):
                    if cell.value in VENUE_MAPPING:
                        bar_columns[col_idx] = VENUE_MAPPING[cell.value]
                continue

            if first_cell.value in METRIC_MAPPING and current_month_year and bar_columns:
                metric_key = METRIC_MAPPING[first_cell.value]
                year, month = current_month_year

                for col_idx, venue_key in bar_columns.items():
                    cell_value = row[col_idx - 1].value
                    if cell_value is not None and cell_value != '':
                        try:
                            value = float(cell_value)
                            month_key = f'{venue_key}_{year}-{month:02d}'
                            if month_key not in all_plans:
                                all_plans[month_key] = {}
                            all_plans[month_key][metric_key] = value
                        except (ValueError, TypeError):
                            pass

        # Расчёт выручки по категориям
        for plan_data in all_plans.values():
            if 'revenue' in plan_data:
                revenue = plan_data['revenue']
                if 'draftShare' in plan_data:
                    plan_data['revenueDraft'] = revenue * (plan_data['draftShare'] / 100)
                if 'packagedShare' in plan_data:
                    plan_data['revenuePackaged'] = revenue * (plan_data['packagedShare'] / 100)
                if 'kitchenShare' in plan_data:
                    plan_data['revenueKitchen'] = revenue * (plan_data['kitchenShare'] / 100)
                if 'loyaltyWriteoffs' not in plan_data:
                    plan_data['loyaltyWriteoffs'] = revenue * 0.05

        # Сохраняем в JSON
        replace_ok = plans_manager.replace_all_plans(
            all_plans,
            source='Excel import from РїР»Р°РЅС‹_2025_2026.xlsx'
        )

        if not replace_ok:
            return jsonify({'error': 'Не удалось сохранить планы на Render Disk'}), 500

        print(f"[PLANS API] РРјРїРѕСЂС‚РёСЂРѕРІР°РЅРѕ РїР»Р°РЅРѕРІ: {len(all_plans)}")

        return jsonify({
            'success': True,
            'message': f'РРјРїРѕСЂС‚РёСЂРѕРІР°РЅРѕ {len(all_plans)} РїР»Р°РЅРѕРІ РёР· Excel',
            'plans_count': len(all_plans),
            'plan_keys': sorted(all_plans.keys())
        })

        output_data = {
            'plans': all_plans,
            'metadata': {
                'lastUpdate': datetime.now().isoformat(),
                'version': '1.0',
                'source': 'Excel import from планы_2025_2026.xlsx'
            }
        }

        with open(plans_manager.data_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)

        # Перезагружаем PlansManager
        plans_manager._initialize_file()
        try:
            from core.daily_plans_generator import regenerate_daily_plans
            regenerate_daily_plans()
            print("[PLANS API] daily_plans.json пересчитан после импорта")
        except Exception as regen_error:
            print(f"[PLANS API WARN] Не удалось пересчитать daily_plans.json: {regen_error}")

        print(f"[PLANS API] Импортировано планов: {len(all_plans)}")

        return jsonify({
            'success': True,
            'message': f'Импортировано {len(all_plans)} планов из Excel',
            'plans_count': len(all_plans),
            'plan_keys': sorted(all_plans.keys())
        })

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка импорта из Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/plans/<period_key>')
def get_plan(period_key):
    """Получить план за конкретный период (без указания заведения)"""
    try:
        print(f"\n[PLANS API] Запрос плана для периода: {period_key}")
        plan = plans_manager.get_plan(period_key)
        if plan:
            print(f"[PLANS API] План найден")
            return jsonify(plan)
        else:
            print(f"[PLANS API] План НЕ найден")
            return jsonify({'error': 'Plan not found'}), 404
    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при получении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/plans', methods=['GET'])
def get_all_plans():
    """Получить все планы (для функции копирования)"""
    try:
        print("\n[PLANS API] Запрос всех планов...")
        all_plans = plans_manager.get_all_plans()
        periods = plans_manager.get_periods_with_plans()
        print(f"[PLANS API] Загружено планов: {len(all_plans)}")
        return jsonify({
            'plans': all_plans,
            'periods': periods
        })
    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при получении планов: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/plans', methods=['POST'])
def save_plan():
    """Сохранить или обновить план"""
    try:
        request_data = request.json

        if not request_data:
            return jsonify({'error': 'No data provided'}), 400

        period_key = request_data.get('period')
        plan_data = request_data.get('data')

        if not period_key:
            return jsonify({'error': 'Period key is required'}), 400

        if not plan_data:
            return jsonify({'error': 'Plan data is required'}), 400

        print(f"\n[PLANS API] Сохранение плана для периода: {period_key}")
        print(f"[PLANS API] Данные: {list(plan_data.keys())}")

        success = plans_manager.save_plan(period_key, plan_data)

        if success:
            print(f"[PLANS API] План успешно сохранен")
            return jsonify({
                'success': True,
                'message': 'Plan saved successfully',
                'period': period_key
            })
        else:
            return jsonify({'error': 'Failed to save plan'}), 500

    except ValueError as e:
        print(f"[PLANS API VALIDATION ERROR] {e}")
        return jsonify({'error': f'Validation error: {str(e)}'}), 400

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при сохранении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/plans/<venue_key>/<period_key>', methods=['POST'])
def save_plan_with_venue(venue_key, period_key):
    """Сохранить или обновить план для конкретного заведения и периода"""
    try:
        request_data = request.json

        if not request_data:
            return jsonify({'error': 'No data provided'}), 400

        print(f"\n[PLANS API] Сохранение плана для заведения: {venue_key}, период: {period_key}")
        print(f"[PLANS API] Данные: {list(request_data.keys())}")

        composite_key = f"{venue_key}_{period_key}"
        print(f"[PLANS API] Составной ключ: {composite_key}")

        success = plans_manager.save_plan(composite_key, request_data)

        if success:
            print(f"[PLANS API] План успешно сохранен")
            return jsonify({
                'success': True,
                'message': 'Plan saved successfully',
                'period': period_key,
                'venue': venue_key
            })
        else:
            return jsonify({'error': 'Failed to save plan'}), 500

    except ValueError as e:
        print(f"[PLANS API VALIDATION ERROR] {e}")
        return jsonify({'error': f'Validation error: {str(e)}'}), 400

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при сохранении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/plans/<venue_key>/<period_key>', methods=['DELETE'])
def delete_plan_with_venue(venue_key, period_key):
    """Удалить план за конкретное заведение и период"""
    try:
        print(f"\n[PLANS API] Удаление плана для заведения: {venue_key}, период: {period_key}")
        composite_key = f"{venue_key}_{period_key}"
        print(f"[PLANS API] Составной ключ: {composite_key}")
        success = plans_manager.delete_plan(composite_key)
        if success:
            print(f"[PLANS API] План удален")
            return jsonify({
                'success': True,
                'message': 'Plan deleted successfully'
            })
        else:
            return jsonify({'error': 'Plan not found'}), 404
    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при удалении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/plans/<period_key>', methods=['DELETE'])
def delete_plan(period_key):
    """Удалить план за период (без указания заведения)"""
    try:
        print(f"\n[PLANS API] Удаление плана для периода: {period_key}")
        success = plans_manager.delete_plan(period_key)
        if success:
            print(f"[PLANS API] План удален")
            return jsonify({
                'success': True,
                'message': 'Plan deleted successfully'
            })
        else:
            return jsonify({'error': 'Plan not found'}), 404
    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при удалении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/comments/<venue_key>/<period_key>', methods=['GET'])
def get_comment(venue_key, period_key):
    """Получить комментарий для периода и заведения"""
    try:
        print(f"\n[COMMENTS API] Получение комментария: {venue_key} / {period_key}")
        plan = plans_manager.get_plan(period_key)
        if plan and 'comment' in plan:
            comment = plan['comment']
            print(f"[COMMENTS API] Комментарий найден: {len(comment)} символов")
            return jsonify({'comment': comment})
        else:
            print(f"[COMMENTS API] Комментарий не найден")
            return jsonify({'comment': None})
    except Exception as e:
        print(f"[COMMENTS API ERROR] Ошибка при получении комментария: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/comments/<venue_key>/<period_key>', methods=['POST'])
def save_comment(venue_key, period_key):
    """Сохранить комментарий для периода и заведения"""
    try:
        data = request.json
        comment = data.get('comment', '').strip()

        print(f"\n[COMMENTS API] Сохранение комментария: {venue_key} / {period_key}")
        print(f"[COMMENTS API] Текст: {len(comment)} символов")

        plan = plans_manager.get_plan(period_key)
        if plan is None:
            plan = {}

        plan['comment'] = comment
        success = plans_manager.save_plan(period_key, plan)

        if success:
            print(f"[COMMENTS API] Комментарий сохранен")
            return jsonify({
                'success': True,
                'message': 'Comment saved successfully'
            })
        else:
            return jsonify({'error': 'Failed to save comment'}), 500

    except Exception as e:
        print(f"[COMMENTS API ERROR] Ошибка при сохранении комментария: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ========== COMPARISON, TRENDS & EXPORT APIs ==========

@dashboard_bp.route('/api/comparison/periods', methods=['POST'])
def compare_periods():
    """API для сравнения двух периодов"""
    try:
        data = request.json
        venue_key = data.get('venue_key')
        period1_key = data.get('period1_key')
        period2_key = data.get('period2_key')

        if not all([venue_key, period1_key, period2_key]):
            return jsonify({'error': 'Missing required parameters'}), 400

        return jsonify({
            'success': True,
            'comparison': {},
            'insights': []
        })

    except Exception as e:
        print(f"[ERROR] /api/comparison/periods: {e}")
        return jsonify({'error': str(e)}), 500


def get_dashboard_analytics_data(bar, date_from, date_to):
    """
    Helper функция для получения аналитических данных дашборда
    Используется в экспорте и других местах
    """
    bar_name = venues_manager.get_iiko_name(bar) if bar and bar != 'all' else None
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
    date_to_inclusive = (date_to_obj + timedelta(days=1)).strftime('%Y-%m-%d')

    olap = OlapReports()
    if not olap.connect():
        raise Exception('Не удалось подключиться к iiko API')

    try:
        all_sales_data = olap.get_all_sales_report(date_from, date_to_inclusive, bar_name)
        if not all_sales_data or not all_sales_data.get('data'):
            raise Exception('Не удалось получить данные из OLAP')
        calculator = DashboardMetrics()
        metrics = calculator.calculate_metrics(all_sales_data)
        return metrics
    finally:
        olap.disconnect()


@dashboard_bp.route('/api/revenue-metrics', methods=['POST'])
def revenue_metrics():
    """
    API для вкладки 'Выручка' — 4 ключевые метрики + таблица по барам
    Использует тот же OLAP запрос что и дашборд
    """
    try:
        data = request.json
        venue_key = data.get('bar')
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        if not date_from or not date_to:
            return jsonify({'error': 'Требуются параметры date_from и date_to'}), 400

        print(f"\n[REVENUE-METRICS] Zapros metrik...")
        print(f"   Bar: {venue_key if venue_key else 'OBSSHAYA'}")
        print(f"   Period: {date_from} - {date_to}")

        # Получаем данные так же как dashboard_analytics (вся выручка, не только напитки)
        bar_name = venues_manager.get_iiko_name(venue_key) if venue_key and venue_key != 'all' else None
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
        date_to_inclusive = (date_to_obj + timedelta(days=1)).strftime('%Y-%m-%d')

        cache_key = f"{venue_key}_{date_from}_{date_to_inclusive}"
        now = time.time()

        # Очищаем кеш для отладки
        DASHBOARD_OLAP_CACHE.clear()

        all_sales_data = None
        if cache_key in DASHBOARD_OLAP_CACHE:
            cached_entry = DASHBOARD_OLAP_CACHE[cache_key]
            if (now - cached_entry['timestamp']) < DASHBOARD_OLAP_CACHE_TTL:
                all_sales_data = cached_entry['data']
                print(f"   [CACHE] Using cached OLAP data")
            else:
                print(f"   [CACHE] Cache expired, fetching new")
        else:
            print(f"   [CACHE] Cache miss, fetching new")

        if not all_sales_data:
            olap = OlapReports()
            if not olap.connect():
                return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

            print(f"   [DEBUG] Calling get_all_sales_report...")
            all_sales_data = olap.get_all_sales_report(date_from, date_to_inclusive, bar_name)
            print(f"   [DEBUG] Got {len(all_sales_data.get('data', [])) if all_sales_data else 0} records")
            olap.disconnect()

            DASHBOARD_OLAP_CACHE[cache_key] = {
                'data': all_sales_data,
                'timestamp': now
            }

        # Считаем метрики из сырых OLAP данных
        calculator = DashboardMetrics()
        metrics = calculator.calculate_metrics(all_sales_data)

        # Применяем маппинг как в dashboard_analytics
        frontend_mapping = {
            'total_revenue': 'revenue',
            'total_checks': 'checks',
            'avg_check': 'averageCheck',
            'draft_share': 'draftShare',
            'bottles_share': 'packagedShare',
            'kitchen_share': 'kitchenShare',
            'draft_revenue': 'revenueDraft',
            'bottles_revenue': 'revenuePackaged',
            'kitchen_revenue': 'revenueKitchen',
            'avg_markup': 'markupPercent',
            'total_margin': 'profit',
            'draft_markup': 'markupDraft',
            'bottles_markup': 'markupPackaged',
            'kitchen_markup': 'markupKitchen',
            'tap_activity': 'tapActivity',
            'loyalty_points_written_off': 'loyaltyWriteoffs'
        }

        mapped = {}
        for old_key, new_key in frontend_mapping.items():
            if old_key in metrics:
                value = metrics[old_key]
                # Наценку умножаем на 100
                if new_key in ['markupPercent', 'markupDraft', 'markupPackaged', 'markupKitchen']:
                    value = value * 100
                mapped[new_key] = value

        # 'revenue' = вся выручка после маппинга
        actual_revenue = mapped.get('revenue', 0)

        # Получаем план за ВЕСЬ месяц из plansdashboard.json
        plan_revenue = 0.0

        # Определяем ключ для плана (например, "all_2026-03" или "bolshoy_2026-03")
        month_key = f"{venue_key if venue_key else 'all'}_{datetime.strptime(date_from, '%Y-%m-%d').strftime('%Y-%m')}"

        plans_file = plans_manager.data_file

        # Читаем plansdashboard.json
        try:
            with open(plans_file, 'r', encoding='utf-8') as f:
                plans_data = json.load(f)
            plans = plans_data.get('plans', {})

            # Берём готовый план на месяц
            month_plan = plans.get(month_key, {})
            plan_revenue = month_plan.get('revenue', 0.0)

        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"   [WARN] Не удалось прочитать {plans_file}: {e}")
            plan_revenue = 0.0

        # Расчёт метрик
        start = datetime.strptime(date_from, '%Y-%m-%d')
        end = datetime.strptime(date_to, '%Y-%m-%d')
        total_days = (end - start).days + 1  # Дней с 1-го по сегодня (для средней)

        # Дней в месяце (для прогноза)
        if start.month == 12:
            days_in_month = 31
        else:
            next_month = datetime(start.year, start.month + 1, 1)
            days_in_month = (next_month - datetime(start.year, start.month, 1)).days

        # Средняя в день (факт / прошедшие дни)
        average_daily = actual_revenue / total_days if total_days > 0 else 0

        # Ожидаемая (прогноз на конец месяца)
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        days_elapsed = min((today - start).days + 1, days_in_month)

        if days_elapsed > 0 and days_elapsed < days_in_month:
            # Экстраполируем среднюю на весь месяц
            expected = average_daily * days_in_month
        else:
            # Месяц закончился — ожидаемая = факт
            expected = actual_revenue

        # % выполнения
        completion = (actual_revenue / plan_revenue * 100) if plan_revenue > 0 else 0

        result = {
            'current': round(actual_revenue, 2),
            'plan': round(plan_revenue, 2),
            'expected': round(expected, 2),
            'average': round(average_daily, 2),
            'period_days': total_days,
            'days_in_month': days_in_month,
            'completion_percent': round(completion, 1)
        }

        print(f"   [OK] Metriki rasschitany: current={result['current']}, plan={result['plan']}")
        return jsonify(result)

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/revenue-metrics: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# Кэш для widget API (5 минут)
WIDGET_CACHE = {}
WIDGET_CACHE_TTL = 300  # 5 минут


@dashboard_bp.route('/api/widget/revenue', methods=['GET'])
def widget_revenue():
    """
    API для PWA виджета — 5 баров с процентами выполнения
    ОПТИМИЗАЦИЯ: ОДИН OLAP запрос на все бары вместо пяти
    """
    try:
        # Проверяем кэш
        now = time.time()
        cache_key = 'widget_revenue'

        if cache_key in WIDGET_CACHE:
            cached_entry = WIDGET_CACHE[cache_key]
            if (now - cached_entry['timestamp']) < WIDGET_CACHE_TTL:
                print(f"[WIDGET] Using cached data")
                return jsonify(cached_entry['data'])

        print(f"\n[WIDGET] Generatsiya dannykh dlya vidzheta (ODIN OLAP zapros)...")

        # Период: с 1-го числа по сегодня
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        month_start = today.replace(day=1)
        date_from = month_start.strftime('%Y-%m-%d')
        date_to = today.strftime('%Y-%m-%d')
        date_to_inclusive = (today + timedelta(days=1)).strftime('%Y-%m-%d')

        print(f"   Period: {date_from} - {date_to}")

        # ОДИН OLAP запрос на ВСЕ бары сразу (без фильтра по Store.Name)
        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        print(f"   [1/3] Zapros OLAP na VSE barы...")
        all_sales_data = olap.get_all_sales_report(date_from, date_to_inclusive, None)  # None = все бары
        olap.disconnect()

        if not all_sales_data or not all_sales_data.get('data'):
            print(f"   [WARN] Net dannykh iz OLAP")
            return jsonify([
                {'bar': '', 'name': 'Общая', 'completion': 0.0},
                {'bar': 'bolshoy', 'name': 'Большой пр. В.О', 'completion': 0.0},
                {'bar': 'ligovskiy', 'name': 'Лиговский', 'completion': 0.0},
                {'bar': 'kremenchugskaya', 'name': 'Кременчугская', 'completion': 0.0},
                {'bar': 'varshavskaya', 'name': 'Варшавская', 'completion': 0.0}
            ])

        # Группируем данные по барам (Store.Name)
        print(f"   [2/3] Gruppировка dannykh po barам...")
        from collections import defaultdict
        bar_data = defaultdict(list)

        for record in all_sales_data.get('data', []):
            store_name = record.get('Store.Name', '')
            if store_name:
                bar_data[store_name].append(record)

        # Маппинг названий баров из iiko в ключи
        iiko_to_key = {
            'Большой пр. В.О': 'bolshoy',
            'Лиговский': 'ligovskiy',
            'Кременчугская': 'kremenchugskaya',
            'Варшавская': 'varshavskaya'
        }

        # Читаем планы один раз
        print(f"   [3/3] Rasschёт metrik...")
        plans_file = plans_manager.data_file
        try:
            with open(plans_file, 'r', encoding='utf-8') as f:
                plans_data = json.load(f)
            plans = plans_data.get('plans', {})
        except:
            plans = {}

        results = []
        total_revenue = 0.0
        total_plan = 0.0

        # Порядок баров
        bars_order = [
            ('', 'Общая'),
            ('bolshoy', 'Большой пр. В.О'),
            ('ligovskiy', 'Лиговский'),
            ('kremenchugskaya', 'Кременчугская'),
            ('varshavskaya', 'Варшавская')
        ]

        for bar_key, bar_name_iiko in bars_order:
            if bar_key == '':
                # Общая = сумма по всем барам
                actual_revenue = 0.0
                plan_revenue = 0.0

                for iiko_name, records in bar_data.items():
                    calculator = DashboardMetrics()
                    metrics = calculator.calculate_metrics({'data': records})

                    frontend_mapping = {
                        'total_revenue': 'revenue',
                        'avg_check': 'averageCheck',
                        'total_margin': 'profit',
                    }
                    mapped = {}
                    for old_key, new_key in frontend_mapping.items():
                        if old_key in metrics:
                            value = metrics[old_key]
                            if new_key in ['averageCheck']:
                                value = value * 100
                            mapped[new_key] = value

                    bar_revenue = mapped.get('revenue', 0)
                    actual_revenue += bar_revenue

                    # План по этому бару
                    bar_key_tmp = iiko_to_key.get(iiko_name, '')
                    month_key = f"{bar_key_tmp if bar_key_tmp else 'all'}_{month_start.strftime('%Y-%m')}"
                    month_plan = plans.get(month_key, {})
                    plan_revenue += month_plan.get('revenue', 0.0)

                total_revenue = actual_revenue
                total_plan = plan_revenue
                completion = (actual_revenue / plan_revenue * 100) if plan_revenue > 0 else 0

            else:
                # Конкретный бар
                iiko_name = bar_name_iiko
                records = bar_data.get(iiko_name, [])

                if records:
                    calculator = DashboardMetrics()
                    metrics = calculator.calculate_metrics({'data': records})

                    frontend_mapping = {
                        'total_revenue': 'revenue',
                        'avg_check': 'averageCheck',
                        'total_margin': 'profit',
                    }
                    mapped = {}
                    for old_key, new_key in frontend_mapping.items():
                        if old_key in metrics:
                            value = metrics[old_key]
                            if new_key in ['averageCheck']:
                                value = value * 100
                            mapped[new_key] = value

                    actual_revenue = mapped.get('revenue', 0)
                else:
                    actual_revenue = 0.0

                # План
                month_key = f"{bar_key}_{month_start.strftime('%Y-%m')}"
                month_plan = plans.get(month_key, {})
                plan_revenue = month_plan.get('revenue', 0.0)

                completion = (actual_revenue / plan_revenue * 100) if plan_revenue > 0 else 0

            results.append({
                'bar': bar_key,
                'name': bar_name_iiko,
                'completion': round(completion, 1)
            })
            print(f"   {bar_name_iiko}: {completion:.1f}% (выручка={actual_revenue:.0f}, план={plan_revenue:.0f})")

        # Кэшируем результат
        WIDGET_CACHE[cache_key] = {
            'data': results,
            'timestamp': now
        }

        print(f"   [OK] Widget data generated (1 OLAP zamiesto 5)")
        return jsonify(results)

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/widget/revenue: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/export/text', methods=['POST'])
def export_text():
    """API для экспорта в текстовый формат"""
    try:
        data = request.json
        venue_name = data.get('venue_name', 'Unknown')
        period = data.get('period', {})
        comparison = data.get('comparison', {})
        insights = data.get('insights', [])

        text_report = export_manager.prepare_text_report(
            venue_name,
            period,
            comparison,
            insights
        )

        return jsonify({
            'success': True,
            'text': text_report
        })

    except Exception as e:
        print(f"[ERROR] /api/export/text: {e}")
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/export/excel', methods=['POST'])
def export_excel():
    """API для экспорта в Excel (XLSX) формат"""
    try:
        from io import BytesIO
        import csv

        data = request.json
        bar = data.get('bar')
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        print(f"\n[EXPORT EXCEL] Генерация Excel: {bar} / {date_from} - {date_to}")

        # Получаем данные
        venue = venues_manager.get_venue(bar)
        venue_name = venue['full_name'] if venue else bar

        # Получаем фактические данные
        actual_data = get_dashboard_analytics_data(bar, date_from, date_to)

        # Получаем плановые данные (пропорционально по периоду)
        plan_data = plans_manager.calculate_plan_for_period(bar, date_from, date_to) or {}

        # Пробуем использовать openpyxl если установлен
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Дашборд"

            # Заголовок
            ws.merge_cells('A1:C1')
            ws['A1'] = f"{venue_name} | {date_from} - {date_to}"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Заголовки таблицы - простая таблица факт/план
            headers = ['Метрика', 'План', 'Факт']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

            # Данные метрик (все 16 показателей)
            metrics = [
                ('Выручка (₽)', 'revenue', 'total_revenue'),
                ('Чеки (шт)', 'checks', 'total_checks'),
                ('Средний чек (₽)', 'averageCheck', 'avg_check'),
                ('Доля розлива (%)', 'draftShare', 'draft_share'),
                ('Доля фасовки (%)', 'packagedShare', 'bottles_share'),
                ('Доля кухни (%)', 'kitchenShare', 'kitchen_share'),
                ('Выручка розлив (₽)', 'revenueDraft', 'draft_revenue'),
                ('Выручка фасовка (₽)', 'revenuePackaged', 'bottles_revenue'),
                ('Выручка кухня (₽)', 'revenueKitchen', 'kitchen_revenue'),
                ('Наценка (%)', 'markupPercent', 'avg_markup'),
                ('Прибыль (₽)', 'profit', 'total_margin'),
                ('Наценка розлив (%)', 'markupDraft', 'draft_markup'),
                ('Наценка фасовка (%)', 'markupPackaged', 'bottles_markup'),
                ('Наценка кухня (%)', 'markupKitchen', 'kitchen_markup'),
                ('Списания баллов (₽)', 'loyaltyWriteoffs', 'loyalty_points_written_off'),
                ('Активность кранов (%)', 'tapActivity', 'tap_activity'),
            ]

            row = 5
            for metric_name, plan_key, actual_key in metrics:
                plan_value = plan_data.get(plan_key, 0) or 0
                actual_value = actual_data.get(actual_key, 0) or 0

                ws.cell(row=row, column=1, value=metric_name)
                ws.cell(row=row, column=2, value=round(plan_value, 2))
                ws.cell(row=row, column=3, value=round(actual_value, 2))
                row += 1

            # Сохраняем в BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            print(f"[EXPORT EXCEL] Excel файл сгенерирован (openpyxl)")
            return output.getvalue(), 200, {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename=dashboard_{bar}_{date_from}_{date_to}.xlsx'
            }

        except ImportError:
            # Fallback to CSV if openpyxl not available
            print("[EXPORT EXCEL] openpyxl не установлен, используем CSV")
            output = BytesIO()
            writer = csv.writer(output)

            writer.writerow([f"{venue_name} | {date_from} - {date_to}"])
            writer.writerow([])
            writer.writerow(['Метрика', 'План', 'Факт'])

            metrics = [
                ('Выручка (₽)', 'revenue', 'total_revenue'),
                ('Чеки (шт)', 'checks', 'total_checks'),
                ('Средний чек (₽)', 'averageCheck', 'avg_check'),
                ('Списания баллов (₽)', 'loyaltyWriteoffs', 'loyalty_points_written_off'),
                ('Прибыль (₽)', 'profit', 'total_margin'),
                ('% наценки', 'markupPercent', 'avg_markup'),
                ('Доля розлива (%)', 'draftShare', 'draft_share'),
                ('Доля фасовки (%)', 'packagedShare', 'bottles_share'),
                ('Доля кухни (%)', 'kitchenShare', 'kitchen_share'),
            ]

            for metric_name, plan_key, actual_key in metrics:
                plan_value = plan_data.get(plan_key, 0) or 0
                actual_value = actual_data.get(actual_key, 0) or 0

                # Для наценки конвертируем (API возвращает дробь, план в процентах)
                if 'markup' in actual_key.lower():
                    actual_value = actual_value * 100

                writer.writerow([metric_name, plan_value, actual_value])

            output.seek(0)

            print(f"[EXPORT EXCEL] CSV файл сгенерирован (fallback)")
            return output.getvalue(), 200, {
                'Content-Type': 'text/csv',
                'Content-Disposition': f'attachment; filename=dashboard_{bar}_{date_from}_{date_to}.csv'
            }

    except Exception as e:
        print(f"[ERROR] /api/export/excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    """API для экспорта в PDF формат"""
    try:
        from io import BytesIO

        data = request.json
        bar = data.get('bar')
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        print(f"\n[EXPORT PDF] Генерация PDF: {bar} / {date_from} - {date_to}")

        # Получаем данные
        venue = venues_manager.get_venue(bar)
        venue_name = venue['full_name'] if venue else bar

        # Получаем фактические данные
        actual_data = get_dashboard_analytics_data(bar, date_from, date_to)

        # Получаем плановые данные (пропорционально по периоду)
        plan_data = plans_manager.calculate_plan_for_period(bar, date_from, date_to) or {}

        # Пробуем использовать reportlab если установлен
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)

            elements = []
            styles = getSampleStyleSheet()

            # Заголовок
            title = Paragraph(f"<b>Дашборд - {venue_name}</b>", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))

            subtitle = Paragraph(f"Период: {date_from} - {date_to}", styles['Normal'])
            elements.append(subtitle)
            elements.append(Spacer(1, 20))

            # Таблица метрик (все 16 показателей)
            data_table = [['Метрика', 'План', 'Факт', '% плана', 'Разница']]

            metrics = [
                ('Выручка', 'revenue', 'total_revenue', '₽'),
                ('Чеки', 'checks', 'total_checks', 'шт'),
                ('Средний чек', 'averageCheck', 'avg_check', '₽'),
                ('Доля розлива', 'draftShare', 'draft_share', '%'),
                ('Доля фасовки', 'packagedShare', 'bottles_share', '%'),
                ('Доля кухни', 'kitchenShare', 'kitchen_share', '%'),
                ('Выручка розлив', 'revenueDraft', 'draft_revenue', '₽'),
                ('Выручка фасовка', 'revenuePackaged', 'bottles_revenue', '₽'),
                ('Выручка кухня', 'revenueKitchen', 'kitchen_revenue', '₽'),
                ('Наценка', 'markupPercent', 'avg_markup', '%'),
                ('Прибыль', 'profit', 'total_margin', '₽'),
                ('Наценка розлив', 'markupDraft', 'draft_markup', '%'),
                ('Наценка фасовка', 'markupPackaged', 'bottles_markup', '%'),
                ('Наценка кухня', 'markupKitchen', 'kitchen_markup', '%'),
                ('Списания баллов', 'loyaltyWriteoffs', 'loyalty_points_written_off', '₽'),
                ('Активность кранов', 'tapActivity', 'tap_activity', '%'),
            ]

            for metric_name, plan_key, actual_key, unit in metrics:
                plan_value = plan_data.get(plan_key, 0) or 0
                actual_value = actual_data.get(actual_key, 0) or 0

                percent = (actual_value / plan_value * 100) if plan_value > 0 else 0
                diff = actual_value - plan_value
                data_table.append([
                    f"{metric_name} ({unit})",
                    f"{plan_value:,.2f}",
                    f"{actual_value:,.2f}",
                    f"{percent:.1f}%",
                    f"{diff:,.2f}"
                ])

            table = Table(data_table)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(table)

            doc.build(elements)
            output.seek(0)

            print(f"[EXPORT PDF] PDF файл сгенерирован (reportlab)")
            return output.getvalue(), 200, {
                'Content-Type': 'application/pdf',
                'Content-Disposition': f'attachment; filename=dashboard_{bar}_{date_from}_{date_to}.pdf'
            }

        except ImportError:
            # Fallback to HTML if reportlab not available
            print("[EXPORT PDF] reportlab не установлен, используем HTML")

            metrics = [
                ('Выручка', 'revenue', 'total_revenue', '₽'),
                ('Чеки', 'checks', 'total_checks', 'шт'),
                ('Средний чек', 'averageCheck', 'avg_check', '₽'),
                ('Доля розлива', 'draftShare', 'draft_share', '%'),
                ('Доля фасовки', 'packagedShare', 'bottles_share', '%'),
                ('Доля кухни', 'kitchenShare', 'kitchen_share', '%'),
                ('Выручка розлив', 'revenueDraft', 'draft_revenue', '₽'),
                ('Выручка фасовка', 'revenuePackaged', 'bottles_revenue', '₽'),
                ('Выручка кухня', 'revenueKitchen', 'kitchen_revenue', '₽'),
                ('Наценка', 'markupPercent', 'avg_markup', '%'),
                ('Прибыль', 'profit', 'total_margin', '₽'),
                ('Наценка розлив', 'markupDraft', 'draft_markup', '%'),
                ('Наценка фасовка', 'markupPackaged', 'bottles_markup', '%'),
                ('Наценка кухня', 'markupKitchen', 'kitchen_markup', '%'),
                ('Списания баллов', 'loyaltyWriteoffs', 'loyalty_points_written_off', '₽'),
                ('Активность кранов', 'tapActivity', 'tap_activity', '%'),
            ]

            cards = []
            for name, plan_key, actual_key, unit in metrics:
                plan_val = plan_data.get(plan_key, 0) or 0
                actual_val = actual_data.get(actual_key, 0) or 0

                percent = (actual_val / plan_val * 100) if plan_val > 0 else 0
                diff = actual_val - plan_val

                if percent >= 100:
                    color = '#4CAF50'
                elif percent >= 90:
                    color = '#FFC107'
                else:
                    color = '#F44336'

                cards.append(f"""
                <div class="metric-card">
                    <div class="metric-name">{name}</div>
                    <div class="metric-values">
                        <div class="value-row">
                            <span class="label">План:</span>
                            <span class="value">{plan_val:,.2f} {unit}</span>
                        </div>
                        <div class="value-row">
                            <span class="label">Факт:</span>
                            <span class="value" style="color: {color}; font-weight: bold;">{actual_val:,.2f} {unit}</span>
                        </div>
                        <div class="value-row progress-row">
                            <span class="label">Выполнение:</span>
                            <div class="progress-bar">
                                <div class="progress-fill" style="width: {min(percent, 100)}%; background: {color};"></div>
                                <span class="progress-text">{percent:.1f}%</span>
                            </div>
                        </div>
                    </div>
                </div>
                """)

            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Дашборд - {venue_name}</title>
                <style>
                    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
                    body {{
                        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        padding: 30px;
                        min-height: 100vh;
                    }}
                    .container {{
                        max-width: 1200px;
                        margin: 0 auto;
                        background: white;
                        border-radius: 20px;
                        padding: 40px;
                        box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                    }}
                    .header {{
                        text-align: center;
                        margin-bottom: 40px;
                        padding-bottom: 20px;
                        border-bottom: 3px solid #667eea;
                    }}
                    h1 {{
                        color: #333;
                        font-size: 32px;
                        margin-bottom: 10px;
                    }}
                    .period {{
                        color: #666;
                        font-size: 16px;
                    }}
                    .metrics-grid {{
                        display: grid;
                        grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
                        gap: 20px;
                    }}
                    .metric-card {{
                        background: #f8f9fa;
                        border-radius: 12px;
                        padding: 20px;
                        transition: transform 0.2s, box-shadow 0.2s;
                    }}
                    .metric-card:hover {{
                        transform: translateY(-5px);
                        box-shadow: 0 8px 16px rgba(0,0,0,0.1);
                    }}
                    .metric-name {{
                        font-size: 18px;
                        font-weight: 600;
                        color: #333;
                        margin-bottom: 15px;
                    }}
                    .metric-values {{
                        display: flex;
                        flex-direction: column;
                        gap: 10px;
                    }}
                    .value-row {{
                        display: flex;
                        justify-content: space-between;
                        align-items: center;
                        padding: 8px 0;
                    }}
                    .label {{
                        color: #666;
                        font-size: 14px;
                    }}
                    .value {{
                        font-size: 16px;
                        color: #333;
                    }}
                    .progress-row {{
                        flex-direction: column;
                        align-items: stretch;
                        gap: 5px;
                        margin-top: 5px;
                    }}
                    .progress-bar {{
                        width: 100%;
                        height: 24px;
                        background: #e0e0e0;
                        border-radius: 12px;
                        position: relative;
                        overflow: hidden;
                    }}
                    .progress-fill {{
                        height: 100%;
                        border-radius: 12px;
                        transition: width 0.3s ease;
                    }}
                    .progress-text {{
                        position: absolute;
                        top: 50%;
                        left: 50%;
                        transform: translate(-50%, -50%);
                        font-size: 12px;
                        font-weight: 600;
                        color: #333;
                    }}
                    @media print {{
                        body {{ background: white; padding: 0; }}
                        .container {{ box-shadow: none; }}
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>{venue_name}</h1>
                        <div class="period">{date_from} - {date_to}</div>
                    </div>
                    <div class="metrics-grid">
                        {''.join(cards)}
                    </div>
                </div>
            </body>
            </html>
            """

            print(f"[EXPORT PDF] HTML файл сгенерирован (fallback)")
            return html_content.encode('utf-8'), 200, {
                'Content-Type': 'text/html; charset=utf-8',
                'Content-Disposition': f'attachment; filename=dashboard_{bar}_{date_from}_{date_to}.html'
            }

    except Exception as e:
        print(f"[ERROR] /api/export/pdf: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
