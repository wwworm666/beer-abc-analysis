from flask import Flask, render_template, request, jsonify
import pandas as pd
import json
import os
import time
import asyncio
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from core.olap_reports import OlapReports
from core.data_processor import BeerDataProcessor
from core.abc_analysis import ABCAnalysis
from core.xyz_analysis import XYZAnalysis
from core.category_analysis import CategoryAnalysis
from core.draft_analysis import DraftAnalysis
from core.waiter_analysis import WaiterAnalysis
from core.taps_manager import TapsManager
from core.employee_analysis import EmployeeMetricsCalculator, get_employees_from_waiter_data
from core.employee_plans import get_employee_plan_by_shifts
from core.iiko_api import IikoAPI
from dashboardNovaev.dashboard_analysis import DashboardMetrics
from dashboardNovaev.plans_manager import PlansManager
from dashboardNovaev.weeks_generator import WeeksGenerator
from dashboardNovaev.backend.venues_manager import VenuesManager
from dashboardNovaev.backend.comparison_calculator import ComparisonCalculator
from dashboardNovaev.backend.trends_analyzer import TrendsAnalyzer
from dashboardNovaev.backend.export_manager import ExportManager
import subprocess

def get_git_commit_hash():
    """Получить короткий хеш текущего git commit для версионирования"""
    try:
        return subprocess.check_output(
            ['git', 'rev-parse', '--short', 'HEAD'],
            stderr=subprocess.DEVNULL
        ).decode('ascii').strip()
    except Exception:
        # Fallback на timestamp для случаев когда git недоступен
        from datetime import datetime
        return datetime.now().strftime('%Y%m%d%H%M')

app = Flask(__name__)

# Определяется при старте приложения
APP_VERSION = get_git_commit_hash()

# Кэш для списка сотрудников (не меняется часто)
EMPLOYEES_CACHE = {'data': None, 'timestamp': 0}
EMPLOYEES_CACHE_TTL = 300  # 5 минут

# Кэш для OLAP запросов дашборда (ключ: venue_bar_dateFrom_dateTo)
DASHBOARD_OLAP_CACHE = {}
DASHBOARD_OLAP_CACHE_TTL = 600  # 10 минут

# Инициализируем менеджер кранов
# Если на Render (есть диск /kultura), используем его для постоянного хранения
# Иначе используем локальную папку data/
TAPS_DATA_PATH = os.environ.get('TAPS_DATA_PATH', 'data/taps_data.json')
if os.path.exists('/kultura'):
    TAPS_DATA_PATH = '/kultura/taps_data.json'
    print(f"[INFO] Используется Render Disk: {TAPS_DATA_PATH}")
else:
    print(f"[INFO] Используется локальный путь: {TAPS_DATA_PATH}")

taps_manager = TapsManager(data_file=TAPS_DATA_PATH)

# Инициализируем менеджер планов (использует ту же логику /kultura)
plans_manager = PlansManager()

# Инициализируем менеджер заведений
venues_manager = VenuesManager()

# Инициализируем Telegram бота (webhook режим)
try:
    import telegram_webhook

    # Загружаем маппинг пива для бота
    beer_mapping_file = os.path.join(os.path.dirname(__file__), 'data', 'beer_info_mapping.json')
    beer_mapping_for_bot = {}
    if os.path.exists(beer_mapping_file):
        with open(beer_mapping_file, 'r', encoding='utf-8') as f:
            beer_mapping_for_bot = json.load(f)
        print(f"[TELEGRAM] Загружен маппинг пива: {len(beer_mapping_for_bot)} записей")

    # Передаем источники данных в telegram модуль
    telegram_webhook.set_data_sources(taps_manager, beer_mapping_for_bot)
    TELEGRAM_BOT_ENABLED = True
    print("[TELEGRAM] Бот инициализирован (webhook режим)")
except Exception as e:
    print(f"[TELEGRAM] Ошибка инициализации бота: {e}")
    TELEGRAM_BOT_ENABLED = False

# Кэш для номенклатуры (15 минут TTL)
nomenclature_cache = {
    'data': None,
    'expires_at': None
}

def get_cached_nomenclature(olap):
    """
    Получить номенклатуру из кэша или запросить новую

    Args:
        olap: Экземпляр OlapReports

    Returns:
        Словарь с номенклатурой или None
    """
    now = datetime.now()

    # Проверяем кэш
    if nomenclature_cache['data'] is not None and nomenclature_cache['expires_at'] is not None:
        if now < nomenclature_cache['expires_at']:
            print(f"[CACHE] Использую кэшированную номенклатуру (истекает через {(nomenclature_cache['expires_at'] - now).seconds // 60} мин)")
            return nomenclature_cache['data']

    # Кэш устарел или пуст - запрашиваем новые данные
    print("[CACHE] Запрашиваю свежую номенклатуру из iiko API...")
    nomenclature = olap.get_nomenclature()

    if nomenclature:
        nomenclature_cache['data'] = nomenclature
        nomenclature_cache['expires_at'] = now + timedelta(minutes=15)
        print(f"[CACHE] Номенклатура закэширована на 15 минут (до {nomenclature_cache['expires_at'].strftime('%H:%M:%S')})")

    return nomenclature

# Список баров
BARS = [
    "Большой пр. В.О",
    "Лиговский",
    "Кременчугская",
    "Варшавская"
]

@app.route('/')
def index():
    """Главная страница - Дашборд"""
    return render_template('dashboard.html', bars=BARS)

@app.route('/packaging')
def packaging():
    """Страница фасовки"""
    return render_template('index.html', bars=BARS)

@app.route('/draft')
def draft():
    """Страница разливного пива"""
    return render_template('draft.html', bars=BARS)

@app.route('/waiters')
def waiters():
    """Страница анализа по официантам"""
    return render_template('waiters.html', bars=BARS)

@app.route('/taps')
def taps():
    """Главная страница управления кранами - выбор бара"""
    return render_template('taps_main.html')

@app.route('/stocks')
def stocks():
    """Страница управления стоками и формирования заказов"""
    return render_template('stocks.html', bars=BARS)

@app.route('/employee')
def employee_dashboard():
    """Страница детального дашборда по сотруднику"""
    return render_template('employee.html', bars=BARS)

@app.route('/bonus')
def bonus_page():
    """Страница расчёта бонусов сотрудников"""
    return render_template('bonus.html')

@app.route('/dashboard')
def dashboard():
    """Страница аналитического дашборда"""
    return render_template('dashboard.html', bars=BARS, app_version=APP_VERSION)

@app.route('/taps/<bar_id>')
def taps_bar(bar_id):
    """Страница управления кранами конкретного бара"""
    bars_config = {
        'bar1': {'name': 'Большой пр. В.О', 'taps': 24},
        'bar2': {'name': 'Лиговский', 'taps': 12},
        'bar3': {'name': 'Кременчугская', 'taps': 12},
        'bar4': {'name': 'Варшавская', 'taps': 12}
    }

    if bar_id not in bars_config:
        return "Бар не найден", 404

    bar_info = bars_config[bar_id]
    return render_template('taps_bar.html',
                         bar_id=bar_id,
                         bar_name=bar_info['name'],
                         tap_count=bar_info['taps'])

@app.route('/api/test', methods=['GET', 'POST'])
def test_endpoint():
    """Тестовый endpoint"""
    print("[TEST] Test endpoint called!")
    return jsonify({'status': 'ok', 'message': 'Test successful'})

@app.route('/api/connection-status', methods=['GET'])
def connection_status():
    """API endpoint для проверки подключения к iiko API"""
    try:
        olap = OlapReports()
        is_connected = olap.connect()
        if is_connected:
            olap.disconnect()
            return jsonify({
                'status': 'connected',
                'message': 'Подключение к iiko API успешно'
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'Не удалось подключиться к iiko API'
            }), 500
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Ошибка подключения: {str(e)}'
        }), 500

@app.route('/api/analyze', methods=['POST'])
def analyze():
    """API endpoint для запуска анализа"""
    try:
        data = request.json
        bar_name = data.get('bar')
        days = int(data.get('days', 30))
        
        print(f"\n[ANALIZ] Zapusk analiza...")
        print(f"   Bar: {bar_name if bar_name else 'VSE'}")
        print(f"   Period: {days} dney")
        
            # Подключаемся к iiko API
        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        try:
            # Запрашиваем данные
            date_to_obj = datetime.now()
            date_from = (date_to_obj - timedelta(days=days)).strftime("%Y-%m-%d")
            date_to = (date_to_obj + timedelta(days=1)).strftime("%Y-%m-%d")  # OLAP exclusive

            report_data = olap.get_beer_sales_report(date_from, date_to, bar_name)
        finally:
            olap.disconnect()

        if not report_data or not report_data.get('data'):
            return jsonify({'error': 'Нет данных за выбранный период'}), 404

        # Обрабатываем данные
        processor = BeerDataProcessor(report_data)
        if not processor.prepare_dataframe():
            return jsonify({'error': 'Ошибка обработки данных'}), 500
        
        agg_data = processor.aggregate_by_beer_and_bar()
        
        # ABC анализ
        abc_analyzer = ABCAnalysis(agg_data)
        
        # XYZ анализ
        xyz_analyzer = XYZAnalysis(processor.df)
        
        if bar_name:
            # Анализ для одного бара
            abc_result = abc_analyzer.perform_abc_analysis_by_bar(bar_name)
            xyz_result = xyz_analyzer.perform_xyz_analysis_by_bar(bar_name)
            
            # Объединяем ABC и XYZ
            if not abc_result.empty and not xyz_result.empty:
                combined = abc_result.merge(
                    xyz_result[['Beer', 'XYZ_Category', 'CoefficientOfVariation']],
                    on='Beer',
                    how='left'
                )
                # Заполняем пустые XYZ как Z (нестабильные)
                combined['XYZ_Category'].fillna('Z', inplace=True)
                combined['CoefficientOfVariation'].fillna(100, inplace=True)

                # Полная категория ABC-XYZ
                combined['ABCXYZ_Combined'] = combined['ABC_Combined'] + '-' + combined['XYZ_Category']

                results = {bar_name: combined}
            else:
                results = {bar_name: abc_result}
        else:
            # Анализ для всех баров - объединяем в единую сводку
            abc_results = abc_analyzer.perform_full_analysis()

            # Объединяем все данные из всех баров
            all_bars_data = []
            for bar, abc_df in abc_results.items():
                if not abc_df.empty:
                    # Добавляем колонку с названием бара
                    abc_df_copy = abc_df.copy()
                    abc_df_copy['Bar'] = bar

                    # Добавляем XYZ данные для этого бара
                    xyz_df = xyz_analyzer.perform_xyz_analysis_by_bar(bar)
                    if not xyz_df.empty:
                        abc_df_copy = abc_df_copy.merge(
                            xyz_df[['Beer', 'XYZ_Category', 'CoefficientOfVariation']],
                            on='Beer',
                            how='left'
                        )
                        abc_df_copy['XYZ_Category'].fillna('Z', inplace=True)
                        abc_df_copy['CoefficientOfVariation'].fillna(100, inplace=True)
                        abc_df_copy['ABCXYZ_Combined'] = abc_df_copy['ABC_Combined'] + '-' + abc_df_copy['XYZ_Category']

                    all_bars_data.append(abc_df_copy)

            # Создаём единый DataFrame со всеми данными
            if all_bars_data:
                combined_all = pd.concat(all_bars_data, ignore_index=True)

                # Агрегируем данные по пиву, объединяя все бары
                # Группируем по Beer, Style, Country и агрегируем метрики
                aggregated = combined_all.groupby(['Beer', 'Style', 'Country']).agg({
                    'TotalQty': 'sum',
                    'TotalRevenue': 'sum',
                    'TotalCost': 'sum',
                    'TotalMargin': 'sum',
                    'CostPerUnit': 'max',  # Берём актуальную себестоимость единицы
                    'AvgMarkupPercent': 'max',  # Берём максимальную наценку по всем барам
                }).reset_index()

                # Пересчитываем ABC категории на основе агрегированных данных
                # Создаем временный экземпляр для использования методов
                abc_temp = ABCAnalysis(pd.DataFrame())

                # 1-я буква: ABC по выручке
                aggregated['ABC_Revenue'] = abc_temp.calculate_abc_category(
                    aggregated['TotalRevenue'],
                    ascending=False
                )

                # 2-я буква: ABC по % наценки (фиксированные пороги)
                aggregated['ABC_Markup'] = abc_temp.calculate_markup_category(
                    aggregated['AvgMarkupPercent']
                )

                # 3-я буква: ABC по марже
                aggregated['ABC_Margin'] = abc_temp.calculate_abc_category(
                    aggregated['TotalMargin'],
                    ascending=False
                )

                # Комбинированная категория
                aggregated['ABC_Combined'] = (
                    aggregated['ABC_Revenue'] +
                    aggregated['ABC_Markup'] +
                    aggregated['ABC_Margin']
                )

                # Добавляем XYZ анализ на основе агрегированных данных
                # Для этого нужно пересчитать вариацию по всем барам
                xyz_all = []
                for beer in aggregated['Beer'].unique():
                    beer_data = combined_all[combined_all['Beer'] == beer]
                    if len(beer_data) > 1:
                        cv = beer_data['TotalQty'].std() / beer_data['TotalQty'].mean() * 100 if beer_data['TotalQty'].mean() > 0 else 100
                    else:
                        cv = 100

                    # Категоризация XYZ
                    if cv < 10:
                        xyz_cat = 'X'
                    elif cv < 25:
                        xyz_cat = 'Y'
                    else:
                        xyz_cat = 'Z'

                    xyz_all.append({
                        'Beer': beer,
                        'CoefficientOfVariation': cv,
                        'XYZ_Category': xyz_cat
                    })

                xyz_df_all = pd.DataFrame(xyz_all)
                aggregated = aggregated.merge(xyz_df_all, on='Beer', how='left')
                aggregated['XYZ_Category'].fillna('Z', inplace=True)
                aggregated['CoefficientOfVariation'].fillna(100, inplace=True)
                aggregated['ABCXYZ_Combined'] = aggregated['ABC_Combined'] + '-' + aggregated['XYZ_Category']

                results = {"Общая": aggregated}
            else:
                results = {}
        
        # Формируем ответ
        response_data = {}
        
        for bar, df in results.items():
            if df.empty:
                continue
                
            # Конвертируем в JSON-friendly формат
            records = df.to_dict('records')
            
            # Статистика по ABC категориям
            abc_stats = df['ABC_Combined'].value_counts().to_dict()
            
            # Статистика по XYZ категориям
            xyz_stats = {}
            if 'XYZ_Category' in df.columns:
                xyz_stats = df['XYZ_Category'].value_counts().to_dict()
            
            # Статистика по ABCXYZ комбинациям
            abcxyz_stats = {}
            if 'ABCXYZ_Combined' in df.columns:
                abcxyz_stats = df['ABCXYZ_Combined'].value_counts().to_dict()
            
            # Топ и худшие фасовки
            top_beers = df.nlargest(10, 'TotalRevenue')[
                ['Beer', 'TotalRevenue', 'ABC_Combined'] + 
                (['XYZ_Category'] if 'XYZ_Category' in df.columns else [])
            ].to_dict('records')
            
            worst_beers = df.nsmallest(10, 'TotalRevenue')[
                ['Beer', 'TotalRevenue', 'ABC_Combined'] + 
                (['XYZ_Category'] if 'XYZ_Category' in df.columns else [])
            ].to_dict('records')
            
            response_data[bar] = {
                'records': records,
                'abc_stats': abc_stats,
                'xyz_stats': xyz_stats,
                'abcxyz_stats': abcxyz_stats,
                'top_beers': top_beers,
                'worst_beers': worst_beers,
                'total_beers': len(df),
                'total_revenue': float(df['TotalRevenue'].sum()),
                'total_qty': float(df['TotalQty'].sum())
            }
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/analyze: {e}")
        import traceback
        traceback.print_exc()
        error_detail = f"{type(e).__name__}: {str(e)}"
        return jsonify({'error': error_detail}), 500
@app.route('/api/weekly-chart/<bar_name>/<beer_name>', methods=['GET'])
def get_weekly_chart(bar_name, beer_name):
    """API endpoint для получения данных графика по неделям"""
    try:
        # Загружаем последний отчет из кеша (в реальности нужно запросить заново)
        # Для простоты используем сохраненные данные
        import os
        if not os.path.exists('beer_report.json'):
            return jsonify({'error': 'Нет данных'}), 404
        
        with open('beer_report.json', 'r', encoding='utf-8') as f:
            report_data = json.load(f)
        
        processor = BeerDataProcessor(report_data)
        if not processor.prepare_dataframe():
            return jsonify({'error': 'Ошибка обработки'}), 500
        
        xyz_analyzer = XYZAnalysis(processor.df)
        chart_data = xyz_analyzer.get_weekly_sales_chart_data(bar_name, beer_name)
        
        if not chart_data:
            return jsonify({'error': 'Нет данных для графика'}), 404
        
        return jsonify(chart_data)
        
    except Exception as e:
        print(f"❌ Ошибка графика: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories', methods=['POST'])
def analyze_categories():
    """API endpoint для анализа по категориям пива"""
    try:
        data = request.json
        bar_name = data.get('bar')
        days = int(data.get('days', 30))

        print(f"\n[CATEGORY] Zapusk analiza po kategoriyam...")
        print(f"   Bar: {bar_name if bar_name else 'VSE'}")
        print(f"   Period: {days} dney")

        # Подключаемся к iiko API
        print("   [1/8] Podklyuchenie k iiko API...")
        olap = OlapReports()
        if not olap.connect():
            print("   [ERROR] Ne udalos podklyuchitsya k iiko API")
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500
        print("   [OK] Podklyucheno k iiko API")

        try:
            # Запрашиваем данные
            print("   [2/8] Zapros dannykh iz OLAP...")
            date_to_obj = datetime.now()
            date_from = (date_to_obj - timedelta(days=days)).strftime("%Y-%m-%d")
            date_to = (date_to_obj + timedelta(days=1)).strftime("%Y-%m-%d")  # OLAP exclusive

            report_data = olap.get_beer_sales_report(date_from, date_to, bar_name)

            if not report_data or not report_data.get('data'):
                print(f"   [ERROR] Net dannykh za period {date_from} - {date_to}")
                return jsonify({'error': 'Нет данных за выбранный период'}), 404
        finally:
            olap.disconnect()

        print(f"   [OK] Polucheno {len(report_data.get('data', []))} zapisey")

        # Обрабатываем данные
        print("   [3/8] Obrabotka dannykh...")
        processor = BeerDataProcessor(report_data)
        if not processor.prepare_dataframe():
            print("   [ERROR] Oshibka obrabotki dannykh")
            return jsonify({'error': 'Ошибка обработки данных'}), 500

        print("   [OK] Dannye obrabotany")

        print("   [4/8] Agregaciya dannykh...")
        agg_data = processor.aggregate_by_beer_and_bar()
        print(f"   [OK] Agregirovano {len(agg_data)} zapisey")

        # Создаем анализатор категорий
        print("   [5/8] Sozdanie analizatora kategoriy...")
        cat_analyzer = CategoryAnalysis(agg_data, processor.df)

        # Проверка null значений после создания
        null_count_after = cat_analyzer.data['Style'].isna().sum()
        uncategorized_count = (cat_analyzer.data['Style'] == 'Bez kategorii (F)').sum()
        print(f"   [STATS] Pustykh Style posle sozdaniya CategoryAnalysis: {null_count_after}")
        print(f"   [STATS] Fasovok v kategorii 'Bez kategorii (F)': {uncategorized_count}")
        print("   [OK] Analizator kategoriy sozdan")

        # XYZ анализатор для добавления данных о стабильности
        print("   [6/8] Sozdanie XYZ analizatora...")
        xyz_analyzer = XYZAnalysis(processor.df)
        print("   [OK] XYZ analizator sozdan")

        # Получаем сводку по категориям
        print("   [7/8] Poluchenie svodki po kategoriyam...")
        summary = cat_analyzer.get_category_summary(bar_name)
        print(f"   [OK] Svodka poluchena: {len(summary)} kategoriy")

        # Получаем детальный анализ для всех категорий
        print("   [8/8] Detalniy analiz kategoriy...")
        category_results = {}

        if bar_name:
            # Для одного бара
            categories = cat_analyzer.get_categories(bar_name)
            print(f"   [STATS] Naydeno kategoriy dlya bara '{bar_name}': {len(categories)}")

            for i, category in enumerate(categories, 1):
                print(f"      [{i}/{len(categories)}] Analiz kategorii: {category}")
                result = cat_analyzer.analyze_category(category, bar_name)
                if result:
                    # Добавляем XYZ данные
                    result = cat_analyzer.add_xyz_to_category_analysis(
                        result, xyz_analyzer, bar_name
                    )
                    category_results[category] = result
                    print(f"         [OK] {result['total_beers']} fasovok")
                else:
                    print(f"         [WARN] Net dannykh dlya kategorii")
        else:
            # Для всех баров - группируем по барам
            print(f"   [STATS] Analiz dlya vsekh barov: {', '.join(BARS)}")
            for bar in BARS:
                bar_results = {}
                categories = cat_analyzer.get_categories(bar)
                print(f"      Bar '{bar}': {len(categories)} kategoriy")

                for category in categories:
                    result = cat_analyzer.analyze_category(category, bar)
                    if result:
                        result = cat_analyzer.add_xyz_to_category_analysis(
                            result, xyz_analyzer, bar
                        )
                        bar_results[category] = result

                if bar_results:
                    category_results[bar] = bar_results
                    print(f"         [OK] {len(bar_results)} kategoriy obrabotano")

        print(f"   [OK] Analiz zavershen")
        print(f"   [STATS] Vsego kategoriy v rezultate: {len(category_results)}")

        # Формируем ответ
        response_data = {
            'summary': summary.to_dict('records'),
            'categories': category_results
        }

        print(f"   [OK] Otvet sformirovan i otpravlyaetsya klientu")
        return jsonify(response_data)

    except Exception as e:
        print(f"[ERROR] Oshibka analiza kategoriy: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/draft-analyze', methods=['POST'])
def analyze_draft():
    """API endpoint для анализа разливного пива"""
    try:
        data = request.json
        bar_name = data.get('bar')
        days = int(data.get('days', 30))
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        print(f"\n[DRAFT] Zapusk analiza razlivnogo piva...")
        print(f"   Bar: {bar_name if bar_name else 'VSE'}")

        # Обработка дат: если переданы конкретные даты, используем их, иначе вычисляем
        if not date_from or not date_to:
            date_to_obj = datetime.now()
            date_from = (date_to_obj - timedelta(days=days)).strftime("%Y-%m-%d")
            date_to = date_to_obj.strftime("%Y-%m-%d")
            print(f"   Period: {days} dney (computed: {date_from} - {date_to})")
        else:
            print(f"   Period: {date_from} - {date_to}")

        # OLAP to-дата exclusive → +1 день
        olap_date_to = (datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')

        # Подключаемся к iiko API
        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        try:
            report_data = olap.get_draft_sales_report(date_from, olap_date_to, bar_name)

            if not report_data or not report_data.get('data'):
                return jsonify({'error': 'Нет данных за выбранный период'}), 404
        finally:
            olap.disconnect()

        # Преобразуем в DataFrame
        df = pd.DataFrame(report_data['data'])
        df['DishAmountInt'] = pd.to_numeric(df['DishAmountInt'], errors='coerce')
        df['DishDiscountSumInt'] = pd.to_numeric(df['DishDiscountSumInt'], errors='coerce')
        df['ProductCostBase.ProductCost'] = pd.to_numeric(df['ProductCostBase.ProductCost'], errors='coerce')
        df['ProductCostBase.MarkUp'] = pd.to_numeric(df['ProductCostBase.MarkUp'], errors='coerce')
        df['OpenDate'] = pd.to_datetime(df['OpenDate.Typed'])

        # Добавляем вычисляемые поля
        df['Margin'] = df['DishDiscountSumInt'] - df['ProductCostBase.ProductCost']

        # Создаем анализатор разливного
        draft_analyzer = DraftAnalysis(df)

        if bar_name:
            # Анализ для одного бара с финансами
            summary = draft_analyzer.get_beer_summary(bar_name, include_financials=True)

            # Применяем ABC анализ (3 буквы: выручка + наценка + маржа)
            if 'TotalRevenue' in summary.columns:
                # 1. ABC по выручке (A)
                summary_sorted = summary.sort_values('TotalRevenue', ascending=False).copy()
                summary_sorted['CumulativeRevenue'] = summary_sorted['TotalRevenue'].cumsum()
                total_revenue = summary_sorted['TotalRevenue'].sum()
                summary_sorted['RevenuePercent'] = (summary_sorted['CumulativeRevenue'] / total_revenue * 100)

                def assign_abc_revenue(pct):
                    if pct <= 80:
                        return 'A'
                    elif pct <= 95:
                        return 'B'
                    else:
                        return 'C'

                summary_sorted['ABC_Revenue'] = summary_sorted['RevenuePercent'].apply(assign_abc_revenue)

                # 2. ABC по наценке (B)
                markup_sorted = summary_sorted.sort_values('AvgMarkupPercent', ascending=False).copy()
                markup_sorted = markup_sorted.reset_index(drop=True)
                markup_sorted['MarkupRank'] = (markup_sorted.index + 1) / len(markup_sorted) * 100

                def assign_abc_markup(rank):
                    if rank <= 33.33:
                        return 'A'
                    elif rank <= 66.66:
                        return 'B'
                    else:
                        return 'C'

                markup_sorted['ABC_Markup'] = markup_sorted['MarkupRank'].apply(assign_abc_markup)
                summary_sorted = summary_sorted.merge(markup_sorted[['BeerName', 'ABC_Markup']], on='BeerName', how='left')

                # 3. ABC по марже (C)
                margin_sorted = summary_sorted.sort_values('TotalMargin', ascending=False).copy()
                margin_sorted = margin_sorted.reset_index(drop=True)
                margin_sorted['MarginRank'] = (margin_sorted.index + 1) / len(margin_sorted) * 100

                def assign_abc_margin(rank):
                    if rank <= 33.33:
                        return 'A'
                    elif rank <= 66.66:
                        return 'B'
                    else:
                        return 'C'

                margin_sorted['ABC_Margin'] = margin_sorted['MarginRank'].apply(assign_abc_margin)
                summary_sorted = summary_sorted.merge(margin_sorted[['BeerName', 'ABC_Margin']], on='BeerName', how='left')

                # Объединяем в 3-буквенный код
                summary_sorted['ABC_Combined'] = summary_sorted['ABC_Revenue'] + summary_sorted['ABC_Markup'] + summary_sorted['ABC_Margin']

                summary = summary_sorted

            # Применяем XYZ анализ
            xyz_df = draft_analyzer.calculate_xyz_for_summary(bar_name)
            if not xyz_df.empty:
                summary = summary.merge(
                    xyz_df[['BeerName', 'XYZ_Category', 'CoefficientOfVariation']],
                    on='BeerName',
                    how='left'
                )
                # Заполняем пустые XYZ как Z (нестабильные)
                summary['XYZ_Category'].fillna('Z', inplace=True)
                summary['CoefficientOfVariation'].fillna(100.0, inplace=True)

                # Полная категория ABC-XYZ (если есть ABC)
                if 'ABC_Combined' in summary.columns:
                    summary['ABCXYZ_Combined'] = summary['ABC_Combined'] + '-' + summary['XYZ_Category']

            # Сортируем по объёму (литры) от большего к меньшему
            summary = summary.sort_values('TotalLiters', ascending=False)

            beers = draft_analyzer.format_summary_for_display(summary)

            response_data = {
                bar_name: {
                    'total_liters': float(summary['TotalLiters'].sum()),
                    'total_portions': int(summary['TotalPortions'].sum()),
                    'total_beers': len(summary),
                    'kegs_30l': float(summary['TotalLiters'].sum() / 30),
                    'kegs_50l': float(summary['TotalLiters'].sum() / 50),
                    'total_revenue': float(summary['TotalRevenue'].sum()) if 'TotalRevenue' in summary.columns else 0,
                    'beers': beers
                }
            }
        else:
            # Анализ для всех баров - объединяем с ABC/XYZ для каждого
            all_bars_data = []

            for bar in BARS:
                bar_summary = draft_analyzer.get_beer_summary(bar, include_financials=True)
                if bar_summary.empty:
                    continue

                # Применяем ABC анализ для этого бара
                if 'TotalRevenue' in bar_summary.columns:
                    # 1. ABC по выручке
                    summary_sorted = bar_summary.sort_values('TotalRevenue', ascending=False).copy()
                    summary_sorted['CumulativeRevenue'] = summary_sorted['TotalRevenue'].cumsum()
                    total_revenue = summary_sorted['TotalRevenue'].sum()
                    summary_sorted['RevenuePercent'] = (summary_sorted['CumulativeRevenue'] / total_revenue * 100)

                    def assign_abc_revenue(pct):
                        if pct <= 80:
                            return 'A'
                        elif pct <= 95:
                            return 'B'
                        else:
                            return 'C'

                    summary_sorted['ABC_Revenue'] = summary_sorted['RevenuePercent'].apply(assign_abc_revenue)

                    # 2. ABC по наценке
                    markup_sorted = summary_sorted.sort_values('AvgMarkupPercent', ascending=False).copy()
                    markup_sorted = markup_sorted.reset_index(drop=True)
                    markup_sorted['MarkupRank'] = (markup_sorted.index + 1) / len(markup_sorted) * 100

                    def assign_abc_markup(rank):
                        if rank <= 33.33:
                            return 'A'
                        elif rank <= 66.66:
                            return 'B'
                        else:
                            return 'C'

                    markup_sorted['ABC_Markup'] = markup_sorted['MarkupRank'].apply(assign_abc_markup)
                    summary_sorted = summary_sorted.merge(markup_sorted[['BeerName', 'ABC_Markup']], on='BeerName', how='left')

                    # 3. ABC по марже
                    margin_sorted = summary_sorted.sort_values('TotalMargin', ascending=False).copy()
                    margin_sorted = margin_sorted.reset_index(drop=True)
                    margin_sorted['MarginRank'] = (margin_sorted.index + 1) / len(margin_sorted) * 100

                    def assign_abc_margin(rank):
                        if rank <= 33.33:
                            return 'A'
                        elif rank <= 66.66:
                            return 'B'
                        else:
                            return 'C'

                    margin_sorted['ABC_Margin'] = margin_sorted['MarginRank'].apply(assign_abc_margin)
                    summary_sorted = summary_sorted.merge(margin_sorted[['BeerName', 'ABC_Margin']], on='BeerName', how='left')

                    # Объединяем в 3-буквенный код
                    summary_sorted['ABC_Combined'] = summary_sorted['ABC_Revenue'] + summary_sorted['ABC_Markup'] + summary_sorted['ABC_Margin']

                    bar_summary = summary_sorted

                # Применяем XYZ анализ для этого бара
                xyz_df = draft_analyzer.calculate_xyz_for_summary(bar)
                if not xyz_df.empty:
                    bar_summary = bar_summary.merge(
                        xyz_df[['BeerName', 'XYZ_Category', 'CoefficientOfVariation']],
                        on='BeerName',
                        how='left'
                    )
                    bar_summary['XYZ_Category'].fillna('Z', inplace=True)
                    bar_summary['CoefficientOfVariation'].fillna(100.0, inplace=True)

                    if 'ABC_Combined' in bar_summary.columns:
                        bar_summary['ABCXYZ_Combined'] = bar_summary['ABC_Combined'] + '-' + bar_summary['XYZ_Category']

                # Сортируем по объёму (литры) от большего к меньшему
                bar_summary = bar_summary.sort_values('TotalLiters', ascending=False)

                all_bars_data.append(bar_summary)

            # Объединяем все бары
            if all_bars_data:
                combined_data = pd.concat(all_bars_data, ignore_index=True)

                # Агрегируем по названию пива (объединяем одинаковые сорта из разных баров)
                agg_dict = {
                    'TotalLiters': 'sum',
                    'TotalPortions': 'sum',
                    'WeeksActive': 'max',
                    'AvgPortionSize': 'mean',
                    'Kegs30L': 'sum',
                    'Kegs50L': 'sum'
                }

                if 'TotalRevenue' in combined_data.columns:
                    agg_dict['TotalRevenue'] = 'sum'
                    agg_dict['TotalCost'] = 'sum'
                    agg_dict['AvgMarkupPercent'] = 'mean'
                    agg_dict['TotalMargin'] = 'sum'

                # Группируем по BeerName
                aggregated = combined_data.groupby('BeerName', as_index=False).agg(agg_dict)

                # Пересчитываем AvgLitersPerWeek
                aggregated['AvgLitersPerWeek'] = aggregated['TotalLiters'] / aggregated['WeeksActive']

                # Добавляем Bar = "Общая"
                aggregated['Bar'] = 'Общая'

                # Применяем ABC анализ к агрегированным данным
                if 'TotalRevenue' in aggregated.columns:
                    # 1. ABC по выручке
                    abc_sorted = aggregated.sort_values('TotalRevenue', ascending=False).copy()
                    abc_sorted['CumulativeRevenue'] = abc_sorted['TotalRevenue'].cumsum()
                    total_revenue = abc_sorted['TotalRevenue'].sum()
                    abc_sorted['RevenuePercent'] = (abc_sorted['CumulativeRevenue'] / total_revenue * 100)

                    def assign_abc_revenue(pct):
                        if pct <= 80:
                            return 'A'
                        elif pct <= 95:
                            return 'B'
                        else:
                            return 'C'

                    abc_sorted['ABC_Revenue'] = abc_sorted['RevenuePercent'].apply(assign_abc_revenue)

                    # 2. ABC по наценке
                    markup_sorted = abc_sorted.sort_values('AvgMarkupPercent', ascending=False).copy()
                    markup_sorted = markup_sorted.reset_index(drop=True)
                    markup_sorted['MarkupRank'] = (markup_sorted.index + 1) / len(markup_sorted) * 100

                    def assign_abc_markup(rank):
                        if rank <= 33.33:
                            return 'A'
                        elif rank <= 66.66:
                            return 'B'
                        else:
                            return 'C'

                    markup_sorted['ABC_Markup'] = markup_sorted['MarkupRank'].apply(assign_abc_markup)
                    abc_sorted = abc_sorted.merge(markup_sorted[['BeerName', 'ABC_Markup']], on='BeerName', how='left')

                    # 3. ABC по марже
                    margin_sorted = abc_sorted.sort_values('TotalMargin', ascending=False).copy()
                    margin_sorted = margin_sorted.reset_index(drop=True)
                    margin_sorted['MarginRank'] = (margin_sorted.index + 1) / len(margin_sorted) * 100

                    def assign_abc_margin(rank):
                        if rank <= 33.33:
                            return 'A'
                        elif rank <= 66.66:
                            return 'B'
                        else:
                            return 'C'

                    margin_sorted['ABC_Margin'] = margin_sorted['MarginRank'].apply(assign_abc_margin)
                    abc_sorted = abc_sorted.merge(margin_sorted[['BeerName', 'ABC_Margin']], on='BeerName', how='left')

                    # Объединяем в 3-буквенный код
                    abc_sorted['ABC_Combined'] = abc_sorted['ABC_Revenue'] + abc_sorted['ABC_Markup'] + abc_sorted['ABC_Margin']

                    aggregated = abc_sorted

                # Применяем XYZ анализ к агрегированным данным (без bar_name - по всем барам)
                xyz_df = draft_analyzer.calculate_xyz_for_summary(None)
                if not xyz_df.empty:
                    aggregated = aggregated.merge(
                        xyz_df[['BeerName', 'XYZ_Category', 'CoefficientOfVariation']],
                        on='BeerName',
                        how='left'
                    )
                    aggregated['XYZ_Category'].fillna('Z', inplace=True)
                    aggregated['CoefficientOfVariation'].fillna(100.0, inplace=True)

                    if 'ABC_Combined' in aggregated.columns:
                        aggregated['ABCXYZ_Combined'] = aggregated['ABC_Combined'] + '-' + aggregated['XYZ_Category']

                # Сортируем по объёму (литры) от большего к меньшему
                aggregated = aggregated.sort_values('TotalLiters', ascending=False)

                beers = draft_analyzer.format_summary_for_display(aggregated)

                response_data = {
                    "Общая": {
                        'total_liters': float(aggregated['TotalLiters'].sum()),
                        'total_portions': int(aggregated['TotalPortions'].sum()),
                        'total_beers': len(aggregated),
                        'kegs_30l': float(aggregated['TotalLiters'].sum() / 30),
                        'kegs_50l': float(aggregated['TotalLiters'].sum() / 50),
                        'total_revenue': float(aggregated['TotalRevenue'].sum()) if 'TotalRevenue' in aggregated.columns else 0,
                        'beers': beers
                    }
                }
            else:
                response_data = {}

        return jsonify(response_data)

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/draft-analyze: {e}")
        import traceback
        traceback.print_exc()
        error_detail = f"{type(e).__name__}: {str(e)}"
        return jsonify({'error': error_detail}), 500

@app.route('/api/waiter-analyze', methods=['POST'])
def analyze_waiters():
    """API endpoint для анализа продаж разливного пива по официантам"""
    try:
        data = request.json
        bar_name = data.get('bar')
        days = int(data.get('days', 30))
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        print(f"\n[WAITER] Zapusk analiza po oficiantam...")
        print(f"   Bar: {bar_name if bar_name else 'VSE'}")

        # Обработка дат: если переданы конкретные даты, используем их, иначе вычисляем
        if not date_from or not date_to:
            date_to_obj = datetime.now()
            date_from = (date_to_obj - timedelta(days=days)).strftime("%Y-%m-%d")
            date_to = date_to_obj.strftime("%Y-%m-%d")
            print(f"   Period: {days} dney (computed: {date_from} - {date_to})")
        else:
            print(f"   Period: {date_from} - {date_to}")

        # OLAP to-дата exclusive → +1 день
        olap_date_to = (datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')

        # Подключаемся к iiko API
        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        try:
            # Запрашиваем данные разливного с официантами
            report_data = olap.get_draft_sales_by_waiter_report(date_from, olap_date_to, bar_name)

            if not report_data or not report_data.get('data'):
                return jsonify({'error': 'Нет данных за выбранный период'}), 404
        finally:
            olap.disconnect()

        # Преобразуем в DataFrame
        df = pd.DataFrame(report_data['data'])
        df['DishAmountInt'] = pd.to_numeric(df['DishAmountInt'], errors='coerce')
        df['DishDiscountSumInt'] = pd.to_numeric(df['DishDiscountSumInt'], errors='coerce')
        df['ProductCostBase.ProductCost'] = pd.to_numeric(df['ProductCostBase.ProductCost'], errors='coerce')
        df['ProductCostBase.MarkUp'] = pd.to_numeric(df['ProductCostBase.MarkUp'], errors='coerce')

        print(f"[INFO] Vsego zapisey: {len(df)}")

        # Создаем анализатор по официантам
        waiter_analyzer = WaiterAnalysis(df)

        # Получаем сводку по официантам
        summary = waiter_analyzer.get_waiter_summary(bar_name)

        if summary.empty:
            return jsonify({'error': 'Нет данных об официантах'}), 404

        print(f"[INFO] Vsego oficiantov: {len(summary)}")

        # Форматируем результат
        waiters = waiter_analyzer.format_summary_for_display(summary)

        # Для каждого официанта получаем детали о том, какое пиво он пролил
        for waiter_record in waiters:
            waiter_name = waiter_record['WaiterName']
            # Передаём bar_name из запроса (None = все бары суммируются, иначе конкретный бар)
            beer_details = waiter_analyzer.get_waiter_beer_details(waiter_name, bar_name)
            if not beer_details.empty:
                beer_details_formatted = waiter_analyzer.format_beer_details_for_display(beer_details)
                waiter_record['beers'] = beer_details_formatted  # Все сорта
            else:
                waiter_record['beers'] = []

        response_data = {
            'total_waiters': len(summary),
            'total_liters': float(summary['TotalLiters'].sum()),
            'total_portions': int(summary['TotalPortions'].sum()),
            'total_revenue': float(summary['TotalRevenue'].sum()) if 'TotalRevenue' in summary.columns else 0,
            'total_margin': float(summary['TotalMargin'].sum()) if 'TotalMargin' in summary.columns else 0,
            'waiters': waiters
        }

        return jsonify(response_data)

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/waiter-analyze: {e}")
        import traceback
        traceback.print_exc()
        error_detail = f"{type(e).__name__}: {str(e)}"
        return jsonify({'error': error_detail}), 500

# ============= API для дашборда сотрудника =============

@app.route('/api/employees', methods=['GET'])
def get_employees_list():
    """Получить список всех сотрудников для dropdown"""
    try:
        # Используем данные о продажах за последние 30 дней чтобы получить актуальный список сотрудников
        date_to_obj = datetime.now()
        date_from = (date_to_obj - timedelta(days=30)).strftime("%Y-%m-%d")
        date_to = (date_to_obj + timedelta(days=1)).strftime("%Y-%m-%d")  # OLAP exclusive

        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        try:
            # Получаем данные разливного с официантами (там точно есть WaiterName)
            report_data = olap.get_draft_sales_by_waiter_report(date_from, date_to, None)
        finally:
            olap.disconnect()

        if not report_data:
            return jsonify({'employees': []})

        # Извлекаем уникальные имена сотрудников
        employees = get_employees_from_waiter_data(report_data)

        return jsonify({'employees': employees})

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/employees: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-analytics', methods=['POST'])
def employee_analytics():
    """API endpoint для детальной аналитики по одному сотруднику"""
    try:
        data = request.json
        employee_name = data.get('employee_name')
        bar_name = data.get('bar')  # None для всех баров
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        if not employee_name or not date_from or not date_to:
            return jsonify({'error': 'Требуются параметры: employee_name, date_from, date_to'}), 400

        print(f"\n[EMPLOYEE] Analiz sotrudnika: {employee_name}")
        print(f"   Bar: {bar_name if bar_name else 'VSE'}")
        print(f"   Period: {date_from} - {date_to}")

        # OLAP трактует to-дату как exclusive → добавляем +1 день чтобы включить весь последний день
        # (бар работает до 2-4 ночи, чеки после полуночи попадают на следующий календарный день)
        olap_date_to = (datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')

        # 1. Получаем данные через OLAP (ПАРАЛЛЕЛЬНО для скорости)
        from concurrent.futures import ThreadPoolExecutor, as_completed

        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        try:
            # Запускаем все OLAP запросы параллельно
            with ThreadPoolExecutor(max_workers=6) as executor:
                futures = {
                    executor.submit(olap.get_employee_aggregated_metrics, date_from, olap_date_to, bar_name): 'aggregated',
                    executor.submit(olap.get_draft_sales_by_waiter_report, date_from, olap_date_to, bar_name): 'draft',
                    executor.submit(olap.get_bottles_sales_by_waiter_report, date_from, olap_date_to, bar_name): 'bottles',
                    executor.submit(olap.get_kitchen_sales_by_waiter_report, date_from, olap_date_to, bar_name): 'kitchen',
                    executor.submit(olap.get_cancelled_orders_by_waiter, date_from, olap_date_to, bar_name): 'cancelled',
                    executor.submit(olap.get_new_loyalty_cards_by_waiter, date_from, olap_date_to, bar_name): 'loyalty_cards',
                }

                results = {}
                for future in as_completed(futures):
                    key = futures[future]
                    try:
                        results[key] = future.result()
                    except Exception as e:
                        print(f"   [ERROR] OLAP {key}: {e}")
                        results[key] = None

            aggregated_data = results.get('aggregated')
            draft_data = results.get('draft')
            bottles_data = results.get('bottles')
            kitchen_data = results.get('kitchen')
            cancelled_data = results.get('cancelled')
            loyalty_cards_data = results.get('loyalty_cards', {})
        finally:
            olap.disconnect()

        # 2. Данные из кассовых смен (сотрудник, локация, выручка, часы - всё из одного API)
        shifts_count = 0
        shift_locations = {}
        shifts_revenue = 0.0
        total_hours = 0.0
        late_count = 0
        try:
            # Проверяем кэш сотрудников
            now = time.time()
            if EMPLOYEES_CACHE['data'] and (now - EMPLOYEES_CACHE['timestamp']) < EMPLOYEES_CACHE_TTL:
                employees_list = EMPLOYEES_CACHE['data']
                print(f"   [CACHE] Using cached employees list ({len(employees_list)} employees)")
            else:
                # Загружаем свежий список
                iiko_emp = IikoAPI()
                if iiko_emp.authenticate():
                    try:
                        employees_list = iiko_emp.get_employees()
                        EMPLOYEES_CACHE['data'] = employees_list
                        EMPLOYEES_CACHE['timestamp'] = now
                        print(f"   [CACHE] Loaded fresh employees list ({len(employees_list)} employees)")
                    finally:
                        iiko_emp.logout()
                else:
                    employees_list = []

            # Ищем ID сотрудника (с нормализацией имени)
            def normalize_name(name):
                if not name:
                    return set()
                return set(name.lower().strip().split())

            employee_id = None
            employee_name_normalized = normalize_name(employee_name)
            for emp in employees_list:
                iiko_name = emp.get('name')
                if not iiko_name:
                    continue
                # Сначала точное совпадение
                if iiko_name == employee_name:
                    employee_id = emp.get('id')
                    break
                iiko_normalized = normalize_name(iiko_name)
                # Потом нормализованное (те же слова в любом порядке)
                if iiko_normalized == employee_name_normalized:
                    employee_id = emp.get('id')
                    print(f"   [MATCH] '{employee_name}' -> '{iiko_name}' (exact set)")
                    break
                # OLAP даёт "Имя Отчество", iiko — "Фамилия Имя Отчество"
                if employee_name_normalized and employee_name_normalized.issubset(iiko_normalized) and len(employee_name_normalized) >= 2:
                    employee_id = emp.get('id')
                    print(f"   [MATCH] '{employee_name}' -> '{iiko_name}' (subset)")
                    break

            # Получаем метрики из кассовых смен (unified метод)
            if employee_id:
                print(f"   Found employee_id: {employee_id}")
                iiko = IikoAPI()
                if iiko.authenticate():
                    try:
                        all_metrics = iiko.get_employee_metrics_from_shifts(date_from, date_to)
                        emp_metrics = all_metrics.get(employee_id, {})
                        shifts_count = emp_metrics.get('shifts_count', 0)
                        shift_locations = emp_metrics.get('shift_locations', {})
                        shifts_revenue = emp_metrics.get('total_revenue', 0.0)
                        total_hours = emp_metrics.get('total_hours', 0.0)
                        late_count = emp_metrics.get('late_count', 0)
                        print(f"   Loaded {shifts_count} cash shifts, {total_hours:.1f} hours, revenue: {shifts_revenue:.0f}, late: {late_count}")
                    finally:
                        iiko.logout()
            else:
                print(f"   [WARN] Employee ID not found for: {employee_name}")
        except Exception as shift_error:
            print(f"   [WARN] Error getting cash shifts: {shift_error}")

        # 3. Рассчитываем план на основе кассовых смен
        plan_revenue = get_employee_plan_by_shifts(shift_locations)
        print(f"   Plan calculated from {len(shift_locations)} cash shifts: {plan_revenue:.0f}")

        # 4. Рассчитываем все метрики (выручка из кассовых смен, категории из OLAP)
        calculator = EmployeeMetricsCalculator()
        metrics = calculator.calculate(
            employee_name=employee_name,
            aggregated_data=aggregated_data,
            draft_data=draft_data,
            bottles_data=bottles_data,
            kitchen_data=kitchen_data,
            cancelled_data=cancelled_data,
            plan_revenue=plan_revenue,
            date_from=date_from,
            date_to=date_to,
            shifts_count_override=shifts_count,
            total_hours_override=total_hours,
            late_count_override=late_count,
            loyalty_cards_count=loyalty_cards_data.get(employee_name, 0),
            total_revenue_override=shifts_revenue if employee_id else None
        )

        return jsonify(metrics)

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/employee-analytics: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-compare', methods=['POST'])
def employee_compare():
    """API endpoint для сравнения нескольких сотрудников"""
    try:
        data = request.json
        employee_names = data.get('employee_names', [])
        bar_name = data.get('bar')
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        if not employee_names or not date_from or not date_to:
            return jsonify({'error': 'Требуются параметры: employee_names, date_from, date_to'}), 400

        if len(employee_names) < 2:
            return jsonify({'error': 'Выберите минимум 2 сотрудников для сравнения'}), 400

        print(f"\n[COMPARE] Sravnenie sotrudnikov: {len(employee_names)} chelovek")
        print(f"   Bar: {bar_name if bar_name else 'VSE'}")
        print(f"   Period: {date_from} - {date_to}")

        # OLAP трактует to-дату как exclusive → +1 день
        olap_date_to = (datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')

        # Загружаем данные ОДИН раз для всех (оптимизация)
        from concurrent.futures import ThreadPoolExecutor, as_completed

        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        try:
            with ThreadPoolExecutor(max_workers=6) as executor:
                futures = {
                    executor.submit(olap.get_employee_aggregated_metrics, date_from, olap_date_to, bar_name): 'aggregated',
                    executor.submit(olap.get_draft_sales_by_waiter_report, date_from, olap_date_to, bar_name): 'draft',
                    executor.submit(olap.get_bottles_sales_by_waiter_report, date_from, olap_date_to, bar_name): 'bottles',
                    executor.submit(olap.get_kitchen_sales_by_waiter_report, date_from, olap_date_to, bar_name): 'kitchen',
                    executor.submit(olap.get_cancelled_orders_by_waiter, date_from, olap_date_to, bar_name): 'cancelled',
                    executor.submit(olap.get_new_loyalty_cards_by_waiter, date_from, olap_date_to, bar_name): 'loyalty_cards',
                }

                all_data = {}
                for future in as_completed(futures):
                    key = futures[future]
                    try:
                        all_data[key] = future.result()
                    except Exception as e:
                        print(f"   [ERROR] OLAP {key}: {e}")
                        all_data[key] = None
        finally:
            olap.disconnect()

        # Получаем список сотрудников и метрики из кассовых смен ОДНИМ запросом (оптимизация)
        now = time.time()
        all_employee_metrics = {}

        iiko = IikoAPI()
        if iiko.authenticate():
            try:
                # Список сотрудников (с кэшем)
                if EMPLOYEES_CACHE['data'] and (now - EMPLOYEES_CACHE['timestamp']) < EMPLOYEES_CACHE_TTL:
                    employees_list = EMPLOYEES_CACHE['data']
                    print(f"   [CACHE] Using cached employees ({len(employees_list)})")
                else:
                    employees_list = iiko.get_employees()
                    EMPLOYEES_CACHE['data'] = employees_list
                    EMPLOYEES_CACHE['timestamp'] = now
                    print(f"   [CACHE] Fresh employees loaded ({len(employees_list)})")

                # Получаем метрики из кассовых смен для ВСЕХ сотрудников (unified метод)
                all_employee_metrics = iiko.get_employee_metrics_from_shifts(date_from, date_to)
                print(f"   [SHIFTS] Loaded metrics for {len(all_employee_metrics)} employees from cash shifts")
            finally:
                iiko.logout()
        else:
            employees_list = []
            all_employee_metrics = {}
            print("   [ERROR] Failed to authenticate to iiko")

        # Хелпер для нормализации имен (игнорируем порядок слов)
        def normalize_name(name):
            if not name:
                return set()
            return set(name.lower().strip().split())

        # Создаём маппинг employee_id -> employee для быстрого поиска
        emp_id_to_name = {emp.get('id'): emp.get('name') for emp in employees_list}

        # DEBUG: показываем имена для сопоставления
        print(f"   [DEBUG] OLAP names to compare: {employee_names}")
        iiko_names = [emp.get('name') for emp in employees_list]
        print(f"   [DEBUG] iiko employee names: {iiko_names[:10]}...")  # первые 10

        # Рассчитываем метрики для каждого сотрудника
        calculator = EmployeeMetricsCalculator()
        results = []

        for emp_name in employee_names:
            # Находим ID сотрудника с нормализацией имени
            employee_id = None
            emp_name_normalized = normalize_name(emp_name)

            for emp in employees_list:
                iiko_name = emp.get('name')
                if not iiko_name:
                    continue
                # Сначала пробуем точное совпадение
                if iiko_name == emp_name:
                    employee_id = emp.get('id')
                    break
                iiko_normalized = normalize_name(iiko_name)
                # Потом пробуем нормализованное (те же слова в любом порядке)
                if iiko_normalized == emp_name_normalized:
                    employee_id = emp.get('id')
                    print(f"   [MATCH] '{emp_name}' -> '{iiko_name}' (exact set)")
                    break
                # OLAP даёт "Имя Отчество", iiko — "Фамилия Имя Отчество"
                # Проверяем что OLAP-имя является подмножеством iiko-имени
                if emp_name_normalized and emp_name_normalized.issubset(iiko_normalized) and len(emp_name_normalized) >= 2:
                    employee_id = emp.get('id')
                    print(f"   [MATCH] '{emp_name}' -> '{iiko_name}' (subset)")
                    break

            # Получаем метрики из кассовых смен по employee_id
            emp_metrics = {}
            shifts_count = 0
            shift_locations = {}
            total_hours = 0.0
            late_count = 0
            cash_revenue = 0.0
            if employee_id:
                emp_metrics = all_employee_metrics.get(employee_id, {})
                shifts_count = emp_metrics.get('shifts_count', 0)
                shift_locations = emp_metrics.get('shift_locations', {})
                total_hours = emp_metrics.get('total_hours', 0.0)
                late_count = emp_metrics.get('late_count', 0)
                cash_revenue = emp_metrics.get('total_revenue', 0.0)
                print(f"   [SHIFTS] {emp_name}: {shifts_count} shifts, {total_hours:.1f}h, revenue: {cash_revenue:.0f}, late: {late_count} (employee_id: {employee_id[:8]}...)")
            else:
                print(f"   [WARNING] {emp_name}: employee_id NOT FOUND in iiko!")

            # Рассчитываем план на основе кассовых смен
            plan_revenue = get_employee_plan_by_shifts(shift_locations)

            # Рассчитываем метрики (выручка из кассовых смен, категории из OLAP)
            loyalty_cards_data = all_data.get('loyalty_cards', {})
            metrics = calculator.calculate(
                employee_name=emp_name,
                aggregated_data=all_data.get('aggregated'),
                draft_data=all_data.get('draft'),
                bottles_data=all_data.get('bottles'),
                kitchen_data=all_data.get('kitchen'),
                cancelled_data=all_data.get('cancelled'),
                plan_revenue=plan_revenue,
                date_from=date_from,
                date_to=date_to,
                shifts_count_override=shifts_count,
                total_hours_override=total_hours,
                late_count_override=late_count,
                loyalty_cards_count=loyalty_cards_data.get(emp_name, 0),
                total_revenue_override=cash_revenue if employee_id else None
            )

            results.append({
                'name': emp_name,
                **metrics
            })

        print(f"[OK] Sravnenie zaversheno dlya {len(results)} sotrudnikov")

        return jsonify({
            'employees': results,
            'period': {'from': date_from, 'to': date_to}
        })

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/employee-compare: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/bonus-calculate', methods=['POST'])
def bonus_calculate():
    """
    API endpoint для расчёта бонусов всех сотрудников.

    Формула (считается ПО ДНЯМ):
    - На каждый рабочий день: если выручка сотрудника > плана ТТ → перевыполнение
    - Суммируются только положительные дни
    - Бонус = 1 000 + (сумма дневных перевыполнений × 5%)
    - Если суммарное перевыполнение = 0 → бонус = 0
    """
    try:
        data = request.json
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        if not date_from or not date_to:
            return jsonify({'error': 'Требуются параметры: date_from, date_to'}), 400

        print(f"\n[BONUS] Raschyot bonusov za period: {date_from} - {date_to}")

        # 1. iiko: кассовые смены — выручка, локация, дата (единый источник)
        now = time.time()
        all_employee_metrics = {}

        iiko = IikoAPI()
        if iiko.authenticate():
            try:
                if EMPLOYEES_CACHE['data'] and (now - EMPLOYEES_CACHE['timestamp']) < EMPLOYEES_CACHE_TTL:
                    employees_list = EMPLOYEES_CACHE['data']
                    print(f"   [CACHE] Using cached employees ({len(employees_list)})")
                else:
                    employees_list = iiko.get_employees()
                    EMPLOYEES_CACHE['data'] = employees_list
                    EMPLOYEES_CACHE['timestamp'] = now
                    print(f"   [CACHE] Fresh employees loaded ({len(employees_list)})")

                all_employee_metrics = iiko.get_employee_metrics_from_shifts(date_from, date_to)
                print(f"   [SHIFTS] Loaded metrics for {len(all_employee_metrics)} employees")
            finally:
                iiko.logout()
        else:
            employees_list = []
            print("   [ERROR] Failed to authenticate to iiko")

        if not all_employee_metrics:
            return jsonify({'error': 'Нет данных за выбранный период'}), 404

        # 2. Подготовка: планы ТТ (всегда свежие данные с диска)
        from core.employee_plans import BarPlansReader, normalize_bar_name, BAR_NAME_MAPPING, clear_plans_cache
        clear_plans_cache()
        plans_reader = BarPlansReader()

        # Маппинг employee_id -> имя
        emp_id_to_name = {emp.get('id'): emp.get('name') for emp in employees_list}

        system_users = EmployeeMetricsCalculator.SYSTEM_USERS + ['']

        # 3. Расчёт по дням для каждого сотрудника (данные из кассовых смен)
        results = []
        total_bonus = 0.0

        for emp_id, emp_metrics in all_employee_metrics.items():
            emp_name = emp_id_to_name.get(emp_id, '')
            if not emp_name or emp_name in system_users:
                continue

            shift_locations = emp_metrics.get('shift_locations', {})
            shift_revenues = emp_metrics.get('shift_revenues', {})
            shifts_count = emp_metrics.get('shifts_count', 0)

            # Считаем по дням
            total_revenue = 0.0
            total_plan = 0.0
            total_overperformance = 0.0
            days_detail = []

            for date_str in sorted(set(list(shift_locations.keys()) + list(shift_revenues.keys()))):
                day_revenue = shift_revenues.get(date_str, 0.0)
                total_revenue += day_revenue

                # План на этот день
                location = shift_locations.get(date_str, '')
                day_plan = 0.0
                if location:
                    normalized = normalize_bar_name(location)
                    mapped_name = BAR_NAME_MAPPING.get(normalized, location)
                    day_plan = plans_reader.get_bar_plan(date_str, mapped_name)

                total_plan += day_plan

                # Перевыполнение: только положительная разница
                day_over = max(0, day_revenue - day_plan) if day_plan > 0 else 0
                total_overperformance += day_over

                # Бонус за смену: 1000 + перевыполнение × 5% (только если план выполнен)
                day_bonus = (1000 + day_over * 0.05) if day_over > 0 else 0

                days_detail.append({
                    'date': date_str,
                    'revenue': round(day_revenue, 2),
                    'plan': round(day_plan, 2),
                    'overperformance': round(day_over, 2),
                    'day_bonus': round(day_bonus, 2)
                })

            # Формула бонуса: 1000 за каждую успешную смену + 5% от перевыполнения
            plan_percent = (total_revenue / total_plan * 100) if total_plan > 0 else 0
            bonus = sum(d['day_bonus'] for d in days_detail)

            total_bonus += bonus

            # Сортируем дни по дате
            days_detail.sort(key=lambda x: x['date'])

            results.append({
                'name': emp_name,
                'plan_revenue': round(total_plan, 2),
                'total_revenue': round(total_revenue, 2),
                'plan_percent': round(plan_percent, 1),
                'overperformance': round(total_overperformance, 2),
                'bonus': round(bonus, 2),
                'shifts_count': shifts_count,
                'days': days_detail
            })

        # Сортируем: сначала с бонусом (по убыванию), потом без
        results.sort(key=lambda x: x['bonus'], reverse=True)

        print(f"[OK] Bonus calculated for {len(results)} employees, total: {total_bonus:.0f}")

        return jsonify({
            'employees': results,
            'total_bonus': round(total_bonus, 2),
            'period': {'from': date_from, 'to': date_to}
        })

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/bonus-calculate: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-metrics-breakdown', methods=['POST'])
def employee_metrics_breakdown():
    """
    API endpoint для получения разбивки метрик по сотрудникам.
    Используется для раскрытия карточек на дашборде.
    """
    try:
        data = request.json
        venue_key = data.get('venue_key', 'all')
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        if not date_from or not date_to:
            return jsonify({'error': 'Требуются параметры: date_from, date_to'}), 400

        print(f"\n[BREAKDOWN] Razbiyka metrik po sotrudnikam")
        print(f"   Venue: {venue_key}, Period: {date_from} - {date_to}")

        # Маппинг venue_key -> bar_name для iiko
        venue_to_bar = {
            'bolshoy': 'Большой пр. В.О',
            'ligovskiy': 'Лиговский',
            'kremenchugskaya': 'Кременчугская',
            'varshavskaya': 'Варшавская',
            'all': None
        }
        bar_name = venue_to_bar.get(venue_key)

        # Загружаем данные OLAP для всех сотрудников
        from concurrent.futures import ThreadPoolExecutor, as_completed

        # OLAP to-дата exclusive → +1 день
        olap_date_to = (datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')

        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        try:
            with ThreadPoolExecutor(max_workers=4) as executor:
                futures = {
                    executor.submit(olap.get_employee_aggregated_metrics, date_from, olap_date_to, bar_name): 'aggregated',
                    executor.submit(olap.get_draft_sales_by_waiter_report, date_from, olap_date_to, bar_name): 'draft',
                    executor.submit(olap.get_bottles_sales_by_waiter_report, date_from, olap_date_to, bar_name): 'bottles',
                    executor.submit(olap.get_kitchen_sales_by_waiter_report, date_from, olap_date_to, bar_name): 'kitchen',
                }

                all_data = {}
                for future in as_completed(futures):
                    key = futures[future]
                    try:
                        all_data[key] = future.result()
                    except Exception as e:
                        print(f"   [ERROR] OLAP {key}: {e}")
                        all_data[key] = None
        finally:
            olap.disconnect()

        # Собираем метрики по сотрудникам
        # aggregated - это dict {name: {metrics}}, draft/bottles/kitchen - OLAP ответы {'data': [...]}
        aggregated = all_data.get('aggregated') or {}
        draft_raw = all_data.get('draft') or {}
        bottles_raw = all_data.get('bottles') or {}
        kitchen_raw = all_data.get('kitchen') or {}
        draft = draft_raw.get('data', []) if isinstance(draft_raw, dict) else []
        bottles = bottles_raw.get('data', []) if isinstance(bottles_raw, dict) else []
        kitchen = kitchen_raw.get('data', []) if isinstance(kitchen_raw, dict) else []

        # Строим breakdown для каждой метрики
        employees_data = {}

        # Из aggregated получаем выручку, чеки (это dict с ключами = имена)
        if isinstance(aggregated, dict):
            for name, metrics in aggregated.items():
                if not name or name == 'Итого':
                    continue
                employees_data[name] = {
                    'name': name,
                    'revenue': float(metrics.get('DishDiscountSumInt', 0) or 0),
                    'checks': int(metrics.get('OrderNum', 0) or 0),
                    'draft_revenue': 0,
                    'bottles_revenue': 0,
                    'kitchen_revenue': 0
                }

        # Добавляем выручку по категориям из OLAP данных
        # Ключ имени официанта в OLAP: WaiterName
        for row in draft:
            if isinstance(row, dict):
                name = row.get('WaiterName') or row.get('Waiter') or row.get('waiter')
                if name and name != 'Итого' and name in employees_data:
                    employees_data[name]['draft_revenue'] += float(row.get('DishDiscountSumInt', 0) or 0)

        for row in bottles:
            if isinstance(row, dict):
                name = row.get('WaiterName') or row.get('Waiter') or row.get('waiter')
                if name and name != 'Итого' and name in employees_data:
                    employees_data[name]['bottles_revenue'] += float(row.get('DishDiscountSumInt', 0) or 0)

        for row in kitchen:
            if isinstance(row, dict):
                name = row.get('WaiterName') or row.get('Waiter') or row.get('waiter')
                if name and name != 'Итого' and name in employees_data:
                    employees_data[name]['kitchen_revenue'] += float(row.get('DishDiscountSumInt', 0) or 0)

        # Рассчитываем производные метрики
        result = []
        for name, data in employees_data.items():
            total_revenue = data['revenue']
            checks = data['checks']

            # Средний чек
            avg_check = total_revenue / checks if checks > 0 else 0

            # Доли
            draft_share = (data['draft_revenue'] / total_revenue * 100) if total_revenue > 0 else 0
            bottles_share = (data['bottles_revenue'] / total_revenue * 100) if total_revenue > 0 else 0
            kitchen_share = (data['kitchen_revenue'] / total_revenue * 100) if total_revenue > 0 else 0

            result.append({
                'name': name,
                'revenue': round(total_revenue, 0),
                'checks': checks,
                'averageCheck': round(avg_check, 0),
                'draftShare': round(draft_share, 1),
                'packagedShare': round(bottles_share, 1),
                'kitchenShare': round(kitchen_share, 1),
                'revenueDraft': round(data['draft_revenue'], 0),
                'revenuePackaged': round(data['bottles_revenue'], 0),
                'revenueKitchen': round(data['kitchen_revenue'], 0)
            })

        # Сортируем по выручке
        result.sort(key=lambda x: x['revenue'], reverse=True)

        print(f"[OK] Razbiyka: {len(result)} sotrudnikov")

        return jsonify({
            'employees': result,
            'period': {'from': date_from, 'to': date_to},
            'venue': venue_key
        })

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/employee-metrics-breakdown: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============= API для управления пивными кранами =============

@app.route('/api/taps/bars', methods=['GET'])
def get_bars_list():
    """Получить список всех баров"""
    try:
        bars = taps_manager.get_bars()
        return jsonify(bars)
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/taps/bars: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/taps/<bar_id>', methods=['GET'])
def get_bar_taps(bar_id):
    """Получить состояние кранов конкретного бара"""
    try:
        result = taps_manager.get_bar_taps(bar_id)
        if 'error' in result:
            return jsonify(result), 404
        return jsonify(result)
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/taps/{bar_id}: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/taps/<bar_id>/start', methods=['POST'])
def start_tap(bar_id):
    """Подключить кегу (начать работу крана)"""
    try:
        data = request.json
        tap_number = data.get('tap_number')
        beer_name = data.get('beer_name')
        keg_id = data.get('keg_id')

        # Проверяем только обязательные поля (keg_id может быть пустым, тогда создастся AUTO)
        if not tap_number or not beer_name:
            return jsonify({'error': 'Требуются: tap_number, beer_name'}), 400

        # Если keg_id пустой, генерируем автоматический
        if not keg_id:
            keg_id = f'AUTO-{int(time.time() * 1000)}'

        result = taps_manager.start_tap(bar_id, int(tap_number), beer_name, keg_id)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/taps/{bar_id}/start: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/taps/<bar_id>/stop', methods=['POST'])
def stop_tap(bar_id):
    """Остановить кран (кега закончилась)"""
    try:
        data = request.json
        tap_number = data.get('tap_number')

        if not tap_number:
            return jsonify({'error': 'Требуется: tap_number'}), 400

        result = taps_manager.stop_tap(bar_id, int(tap_number))

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/taps/{bar_id}/stop: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/taps/<bar_id>/replace', methods=['POST'])
def replace_tap(bar_id):
    """Заменить кегу (смена сорта пива)"""
    try:
        print(f"[DEBUG] /api/taps/{bar_id}/replace called")
        data = request.json
        print(f"[DEBUG] Request data: {data}")

        tap_number = data.get('tap_number')
        beer_name = data.get('beer_name')
        keg_id = data.get('keg_id')

        print(f"[DEBUG] tap_number={tap_number}, beer_name={beer_name}, keg_id={keg_id}")

        # Проверяем только обязательные поля (keg_id может быть пустым, тогда создастся AUTO)
        if not tap_number or not beer_name:
            print(f"[ERROR] Missing required fields")
            return jsonify({'error': 'Требуются: tap_number, beer_name'}), 400

        # Если keg_id пустой, генерируем автоматический
        if not keg_id:
            keg_id = f'AUTO-{int(time.time() * 1000)}'
            print(f"[DEBUG] Generated auto keg_id: {keg_id}")

        print(f"[DEBUG] Calling taps_manager.replace_tap...")
        result = taps_manager.replace_tap(bar_id, int(tap_number), beer_name, keg_id)
        print(f"[DEBUG] Result: {result}")

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/taps/{bar_id}/replace: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/taps/<bar_id>/<int:tap_number>/history', methods=['GET'])
def get_tap_history(bar_id, tap_number):
    """Получить историю действий крана"""
    try:
        limit = request.args.get('limit', 50, type=int)
        result = taps_manager.get_tap_history(bar_id, tap_number, limit)

        if 'error' in result:
            return jsonify(result), 404
        return jsonify(result)
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/taps/{bar_id}/{tap_number}/history: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/taps/events/all', methods=['GET'])
def get_all_events():
    """Получить все события"""
    try:
        bar_id = request.args.get('bar_id', None)
        limit = request.args.get('limit', 100, type=int)

        events = taps_manager.get_all_events(bar_id, limit)
        return jsonify({'events': events})
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/taps/events/all: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/taps/statistics', methods=['GET'])
def get_statistics():
    """Получить статистику по кранам"""
    try:
        bar_id = request.args.get('bar_id', None)
        result = taps_manager.get_statistics(bar_id)

        if 'error' in result:
            return jsonify(result), 404
        return jsonify(result)
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/taps/statistics: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/taps/export-taplist', methods=['GET'])
def export_taplist():
    """Экспортировать текущий таплист в CSV формате"""
    try:
        from io import StringIO
        import csv

        # Опциональный параметр bar_id для фильтрации по конкретному бару
        bar_id_filter = request.args.get('bar_id', None)

        # Получаем список всех баров
        bars = taps_manager.get_bars()

        # Фильтруем по конкретному бару, если указан
        if bar_id_filter:
            bars = [bar for bar in bars if bar['bar_id'] == bar_id_filter]

        # Создаём CSV в памяти
        output = StringIO()
        writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_MINIMAL)

        # Заголовок
        writer.writerow(['Бар', 'Номер крана', 'Название пива'])

        # Собираем данные по всем барам (или по одному, если bar_id_filter указан)
        for bar in bars:
            bar_id = bar['bar_id']
            bar_name = bar['name']

            # Получаем краны бара
            bar_data = taps_manager.get_bar_taps(bar_id)
            if 'error' in bar_data:
                continue

            # Добавляем все краны (активные и пустые)
            for tap in bar_data.get('taps', []):
                beer_name = tap['current_beer'] if tap['current_beer'] else '(пусто)'
                writer.writerow([
                    bar_name,
                    tap['tap_number'],
                    beer_name
                ])

        # Готовим ответ
        output.seek(0)
        csv_content = output.getvalue()
        output.close()

        # Формируем имя файла
        if bar_id_filter and bars:
            # Используем bar_id вместо имени для безопасности
            filename = f"taplist_{bar_id_filter}.csv"
        else:
            filename = "taplist.csv"

        # Создаём response с правильными заголовками
        from flask import make_response
        from urllib.parse import quote
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        # Используем RFC 5987 для корректной работы с кириллицей
        response.headers['Content-Disposition'] = f"attachment; filename={filename}; filename*=UTF-8''{quote(filename)}"

        return response

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/taps/export-taplist: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def load_beer_info_mapping():
    """Загружает маппинг информации о пиве из JSON файла"""
    mapping_file = os.path.join(os.path.dirname(__file__), 'data', 'beer_info_mapping.json')
    if os.path.exists(mapping_file):
        try:
            with open(mapping_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"[ERROR] Ошибка загрузки beer_info_mapping.json: {e}")
    return {}


def find_beer_info(beer_name, mapping):
    """
    Ищет информацию о пиве в маппинге с fuzzy matching.
    Использует difflib для нечёткого сравнения строк.
    """
    import re
    from difflib import SequenceMatcher

    if not beer_name or not mapping:
        return None

    def normalize(name):
        """Нормализует название для сравнения"""
        name = name.lower()
        # Убираем "кег "
        name = name.replace('кег ', '')
        # Заменяем тире на пробел
        name = name.replace(' — ', ' ').replace('—', ' ').replace('-', ' ')
        # Убираем запятые и точки
        name = name.replace(',', '').replace('.', '')
        # Убираем объёмы и единицы измерения
        name = re.sub(r'\d+\s*(л|l|кг|kg|ml|мл)', '', name)
        # Убираем типичные суффиксы
        for suffix in ['светлое', 'темное', 'тёмное', 'нефильтрованное', 'фильтрованное', 'пшеничное', 'полусухой', 'полусладкий']:
            name = name.replace(suffix, '')
        # Убираем лишние пробелы
        name = ' '.join(name.split())
        return name.strip()

    def similarity(a, b):
        """Возвращает степень схожести двух строк (0-1)"""
        return SequenceMatcher(None, a, b).ratio()

    # Прямое совпадение
    if beer_name in mapping:
        return mapping[beer_name]

    # Нормализуем искомое название
    normalized_search = normalize(beer_name)

    # Ищем лучшее совпадение
    best_match = None
    best_score = 0
    threshold = 0.75  # Минимальная схожесть 75%

    for key in mapping:
        normalized_key = normalize(key)

        # Точное совпадение после нормализации
        if normalized_search == normalized_key:
            return mapping[key]

        # Fuzzy matching
        score = similarity(normalized_search, normalized_key)
        if score > best_score:
            best_score = score
            best_match = key

    # Возвращаем лучшее совпадение если оно выше порога
    if best_match and best_score >= threshold:
        return mapping[best_match]

    return None


@app.route('/api/taps/taplist-full', methods=['GET'])
def get_taplist_full():
    """
    Получить полный таплист с расширенной информацией о пиве.
    Возвращает JSON с данными: пивоварня, название, untappd URL, стиль, ABV, IBU, описание.

    Параметры:
    - bar_id: фильтр по конкретному бару (опционально)
    - active_only: true - только активные краны (по умолчанию true)
    """
    try:
        bar_id_filter = request.args.get('bar_id', None)
        active_only = request.args.get('active_only', 'true').lower() == 'true'

        # Загружаем маппинг
        beer_mapping = load_beer_info_mapping()

        # Получаем список баров
        bars = taps_manager.get_bars()

        if bar_id_filter:
            bars = [bar for bar in bars if bar['bar_id'] == bar_id_filter]

        result = []

        for bar in bars:
            bar_id = bar['bar_id']
            bar_name = bar['name']

            bar_data = taps_manager.get_bar_taps(bar_id)
            if 'error' in bar_data:
                continue

            for tap in bar_data.get('taps', []):
                if active_only and tap['status'] != 'active':
                    continue

                beer_name = tap.get('current_beer')
                if not beer_name:
                    continue

                # Ищем расширенную информацию
                beer_info = find_beer_info(beer_name, beer_mapping)

                tap_data = {
                    'bar': bar_name,
                    'bar_id': bar_id,
                    'tap_number': tap['tap_number'],
                    'iiko_name': beer_name,
                    'started_at': tap.get('started_at'),
                }

                if beer_info:
                    tap_data.update({
                        'brewery': beer_info.get('brewery'),
                        'beer_name': beer_info.get('beer_name'),
                        'untappd_url': beer_info.get('untappd_url'),
                        'style': beer_info.get('style'),
                        'abv': beer_info.get('abv'),
                        'ibu': beer_info.get('ibu'),
                        'description': beer_info.get('description'),
                        'mapped': True
                    })
                else:
                    tap_data['mapped'] = False

                result.append(tap_data)

        return jsonify({
            'success': True,
            'count': len(result),
            'taplist': result
        })

    except Exception as e:
        print(f"[ERROR] Ошибка в /api/taps/taplist-full: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/taps/export-taplist-full', methods=['GET'])
def export_taplist_full():
    """
    Экспортировать расширенный таплист в CSV формате.
    Включает: Бар, Кран, Пивоварня, Название пива, Untappd URL, Стиль, ABV, IBU, Описание
    """
    try:
        from io import StringIO
        import csv

        bar_id_filter = request.args.get('bar_id', None)
        active_only = request.args.get('active_only', 'true').lower() == 'true'

        # Загружаем маппинг
        beer_mapping = load_beer_info_mapping()

        bars = taps_manager.get_bars()

        if bar_id_filter:
            bars = [bar for bar in bars if bar['bar_id'] == bar_id_filter]

        output = StringIO()
        writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_ALL)

        # Заголовок
        writer.writerow(['Бар', 'Кран', 'Пивоварня', 'Название пива', 'Untappd URL', 'Стиль', 'ABV', 'IBU', 'Описание'])

        for bar in bars:
            bar_id = bar['bar_id']
            bar_name = bar['name']

            bar_data = taps_manager.get_bar_taps(bar_id)
            if 'error' in bar_data:
                continue

            for tap in bar_data.get('taps', []):
                if active_only and tap['status'] != 'active':
                    continue

                beer_name = tap.get('current_beer')
                if not beer_name:
                    continue

                beer_info = find_beer_info(beer_name, beer_mapping)

                if beer_info:
                    writer.writerow([
                        bar_name,
                        tap['tap_number'],
                        beer_info.get('brewery', ''),
                        beer_info.get('beer_name', ''),
                        beer_info.get('untappd_url', ''),
                        beer_info.get('style', ''),
                        beer_info.get('abv', ''),
                        beer_info.get('ibu', ''),
                        beer_info.get('description', '')
                    ])
                else:
                    # Если нет маппинга - используем название из iiko
                    writer.writerow([
                        bar_name,
                        tap['tap_number'],
                        '',
                        beer_name,
                        '',
                        '',
                        '',
                        '',
                        ''
                    ])

        output.seek(0)
        csv_content = output.getvalue()
        output.close()

        if bar_id_filter and bars:
            filename = f"taplist_full_{bar_id_filter}.csv"
        else:
            filename = "taplist_full.csv"

        from flask import make_response
        from urllib.parse import quote
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f"attachment; filename={filename}; filename*=UTF-8''{quote(filename)}"

        return response

    except Exception as e:
        print(f"[ERROR] Ошибка в /api/taps/export-taplist-full: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/taps/<bar_id>/stats', methods=['GET'])
def get_bar_stats(bar_id):
    """Получить краткую статистику для карточки бара"""
    try:
        result = taps_manager.get_bar_taps(bar_id)
        if 'error' in result:
            return jsonify({'active': 0, 'empty': 12, 'activity_7d': 0}), 200

        taps = result.get('taps', [])
        active = len([t for t in taps if t.get('status') == 'active'])
        total = result.get('total_taps', 12)
        empty = total - len([t for t in taps if t.get('status') in ['active', 'replacing']])

        # Рассчитываем активность за последние 7 дней
        from datetime import datetime, timedelta
        today = datetime.now()
        week_ago = today - timedelta(days=7)
        date_from = week_ago.strftime('%Y-%m-%d')
        date_to = today.strftime('%Y-%m-%d')

        activity_7d = taps_manager.calculate_tap_activity_for_period(bar_id, date_from, date_to)

        return jsonify({
            'active': active,
            'empty': empty,
            'total': total,
            'activity_7d': round(activity_7d, 1)  # Процент активности за последние 7 дней
        })
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/taps/{bar_id}/stats: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'active': 0, 'empty': 12, 'activity_7d': 0}), 200

@app.route('/api/beers/draft', methods=['GET'])
def get_draft_beers():
    """Получить список разливного пива из номенклатуры"""
    try:
        # Читаем файл с номенклатурой
        import os
        products_file = os.path.join('data', 'all_products.json')

        if not os.path.exists(products_file):
            # Если файла нет, возвращаем пустой список
            return jsonify({'beers': []})

        with open(products_file, 'r', encoding='utf-8') as f:
            products = json.load(f)

        # Фильтруем только разливное пиво
        # Разливное обычно содержит "кег", "KEG", "30L", "50L" в названии
        draft_beers = []
        seen_names = set()

        for product in products:
            name = product.get('name', '')
            # Ищем признаки разливного пива
            if any(keyword in name.upper() for keyword in ['КЕГ', 'KEG', '30L', '50L', 'DRAFT']):
                # Убираем дубликаты по названию
                clean_name = name.strip()
                if clean_name and clean_name not in seen_names:
                    draft_beers.append({
                        'id': product.get('id'),
                        'name': clean_name,
                        'num': product.get('num')
                    })
                    seen_names.add(clean_name)

        # Сортируем по названию
        draft_beers.sort(key=lambda x: x['name'])

        return jsonify({'beers': draft_beers})
    except Exception as e:
        print(f"[ERROR] Oshibka v /api/beers/draft: {e}")
        return jsonify({'beers': []}), 200

@app.route('/api/update-nomenclature', methods=['POST'])
def update_nomenclature():
    """Обновить список продуктов из iiko"""
    try:
        import requests
        import xml.etree.ElementTree as ET
        from core.iiko_api import IikoAPI

        print("[INFO] Начинаем обновление номенклатуры...")

        # Подключаемся к iiko
        api = IikoAPI()
        if not api.authenticate():
            print("[ERROR] Не удалось авторизоваться в iiko")
            return jsonify({'success': False, 'error': 'Authentication failed'}), 500

        # Получаем список продуктов
        url = f"{api.base_url}/products"
        params = {"key": api.token}

        response = requests.get(url, params=params, timeout=30)

        if response.status_code != 200:
            api.logout()
            print(f"[ERROR] Ошибка получения продуктов: {response.status_code}")
            return jsonify({'success': False, 'error': f'API error: {response.status_code}'}), 500

        # Парсим XML
        root = ET.fromstring(response.content)
        products = []

        for product in root.findall('.//productDto'):
            product_id = product.find('id').text if product.find('id') is not None else None
            name = product.find('name').text if product.find('name') is not None else None
            parent_id = product.find('parentId').text if product.find('parentId') is not None else None
            num = product.find('num').text if product.find('num') is not None else None

            products.append({
                'id': product_id,
                'name': name,
                'parent_id': parent_id,
                'num': num
            })

        # Сохраняем в файл
        products_file = os.path.join('data', 'all_products.json')
        with open(products_file, 'w', encoding='utf-8') as f:
            json.dump(products, f, indent=2, ensure_ascii=False)

        api.logout()

        print(f"[OK] Обновлено продуктов: {len(products)}")
        return jsonify({'success': True, 'count': len(products)})

    except Exception as e:
        print(f"[ERROR] Ошибка при обновлении номенклатуры: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/stocks/taplist', methods=['GET'])
def get_taplist_stocks():
    """API endpoint для получения остатков КЕГ из iiko API (только активные на кранах)"""
    try:
        bar = request.args.get('bar', '')

        if not bar:
            return jsonify({'error': 'Требуется параметр bar'}), 400

        # Получаем список активных кег из taps_manager
        bar_id_map = {
            'Большой пр. В.О': 'bar1',
            'Лиговский': 'bar2',
            'Кременчугская': 'bar3',
            'Варшавская': 'bar4',
            'Общая': None  # Для "Общая" покажем все бары
        }

        # Маппинг баров на склады iiko (store_id)
        store_id_map = {
            'bar1': 'a4c88d1c-be9a-4366-9aca-68ddaf8be40d',  # Большой пр. В.О
            'bar2': '91d7d070-875b-4d98-a81c-ae628eca45fd',  # Лиговский
            'bar3': '1239d270-1bbe-f64f-b7ea-5f00518ef508',  # Кременчугская
            'bar4': '1ebd631f-2e6d-4f74-8b32-0e54d9efd97d',  # Варшавская
        }

        # Собираем список активных кег со всех кранов с номерами кранов
        active_beers = set()
        beer_to_taps = {}  # { beer_name: [tap_numbers] }

        import re

        if bar == 'Общая':
            # Для "Общая" собираем со всех баров
            for bar_id in ['bar1', 'bar2', 'bar3', 'bar4']:
                result = taps_manager.get_bar_taps(bar_id)
                if 'taps' in result:
                    for tap in result['taps']:
                        if tap.get('status') == 'active' and tap.get('current_beer'):
                            beer_name = tap['current_beer']
                            tap_number = tap.get('tap_number', '?')
                            # Обрабатываем название так же, как из остатков
                            beer_name = re.sub(r'^КЕГ\s+', '', beer_name, flags=re.IGNORECASE)
                            beer_name = re.sub(r'^Кег\s+', '', beer_name, flags=re.IGNORECASE)
                            beer_name = re.sub(r',?\s*\d+\s*л.*', '', beer_name)
                            beer_name = re.sub(r'\s+л\s*$', '', beer_name)
                            beer_name = re.sub(r',?\s*кег.*', '', beer_name, flags=re.IGNORECASE)
                            beer_name = re.sub(r',\s*$', '', beer_name)
                            beer_name = beer_name.strip()
                            active_beers.add(beer_name)
                            if beer_name not in beer_to_taps:
                                beer_to_taps[beer_name] = []
                            beer_to_taps[beer_name].append(tap_number)
        else:
            bar_id = bar_id_map.get(bar)
            if bar_id:
                result = taps_manager.get_bar_taps(bar_id)
                if 'taps' in result:
                    for tap in result['taps']:
                        if tap.get('status') == 'active' and tap.get('current_beer'):
                            beer_name = tap['current_beer']
                            tap_number = tap.get('tap_number', '?')
                            # Обрабатываем название так же, как из остатков
                            beer_name = re.sub(r'^КЕГ\s+', '', beer_name, flags=re.IGNORECASE)
                            beer_name = re.sub(r'^Кег\s+', '', beer_name, flags=re.IGNORECASE)
                            beer_name = re.sub(r',?\s*\d+\s*л.*', '', beer_name)
                            beer_name = re.sub(r'\s+л\s*$', '', beer_name)
                            beer_name = re.sub(r',?\s*кег.*', '', beer_name, flags=re.IGNORECASE)
                            beer_name = re.sub(r',\s*$', '', beer_name)
                            beer_name = beer_name.strip()
                            active_beers.add(beer_name)
                            if beer_name not in beer_to_taps:
                                beer_to_taps[beer_name] = []
                            beer_to_taps[beer_name].append(tap_number)

        # Подключаемся к iiko API
        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        try:
            # Получаем номенклатуру для маппинга GUID -> информация о товаре
            nomenclature = get_cached_nomenclature(olap)
            if not nomenclature:
                return jsonify({'error': 'Не удалось получить номенклатуру'}), 500

            # Получаем РЕАЛЬНЫЕ остатки на складах (текущее время)
            from datetime import datetime
            current_time = datetime.now()
            print(f"[DEBUG] Текущее время сервера: {current_time.strftime('%Y-%m-%d %H:%M:%S')}", flush=True)

            balances = olap.get_store_balances()
            if not balances:
                return jsonify({'error': 'Не удалось получить остатки'}), 500

            print(f"[DEBUG] Получено {len(balances)} записей остатков", flush=True)
        finally:
            olap.disconnect()

        # Определяем склад для фильтрации
        target_store_id = None
        if bar != 'Общая':
            bar_id = bar_id_map.get(bar)
            if bar_id:
                target_store_id = store_id_map.get(bar_id)

        # Обрабатываем остатки кег (GOODS в литрах)
        beer_stocks = {}

        for balance in balances:
            product_id = balance.get('product')
            amount = balance.get('amount', 0)
            store_id = balance.get('store')

            # Фильтруем по складу конкретного бара (если не "Общая")
            if target_store_id and store_id != target_store_id:
                continue

            # НЕ пропускаем отрицательные остатки - они важны для контроля!
            # Пропускаем только нулевые, если кега не активна
            if amount == 0:
                # Пропустим потом при проверке на активность
                pass

            # Получаем информацию о товаре из номенклатуры
            product_info = nomenclature.get(product_id)
            if not product_info:
                continue

            # Берем только GOODS (кеги - это товары, а не блюда!)
            if product_info.get('type') != 'GOODS':
                continue

            # Берем только литры (кеги измеряются в литрах)
            if product_info.get('mainUnit') != 'л':
                continue

            product_name = product_info.get('name', product_id)

            # Убираем "Кег" и объёмы из названия для агрегации
            import re
            base_name = product_name
            base_name = re.sub(r'^Кег\s+', '', base_name, flags=re.IGNORECASE)
            # Убираем объемы: ", 20 л", "20 л", или просто " л" в конце
            base_name = re.sub(r',?\s*\d+\s*л.*', '', base_name)
            base_name = re.sub(r'\s+л\s*$', '', base_name)  # Убираем " л" в конце если осталось
            base_name = re.sub(r',?\s*кег.*', '', base_name, flags=re.IGNORECASE)
            base_name = re.sub(r',\s*$', '', base_name)  # Убираем запятую в конце
            base_name = base_name.strip()

            # ФИЛЬТР: Показываем только кеги, которые стоят на кранах
            # Названия идентичны в iiko и на кранах, поэтому только точное сравнение
            if active_beers:
                is_active = base_name in active_beers

                # Если кега не активна, пропускаем
                if not is_active:
                    continue

            category = product_info.get('category', 'Разливное')

            if base_name not in beer_stocks:
                beer_stocks[base_name] = {
                    'remaining_liters': 0,
                    'category': category,
                    'on_tap': base_name in active_beers if active_beers else False
                }

            # Суммируем остатки по базовому названию
            beer_stocks[base_name]['remaining_liters'] += amount

        # Добавляем активные краны, которых нет в остатках (с нулевым остатком)
        for active_beer in active_beers:
            if active_beer not in beer_stocks:
                beer_stocks[active_beer] = {
                    'remaining_liters': 0,
                    'category': 'Разливное',
                    'on_tap': True
                }

        # Формируем итоговый список
        taps_data = []
        total_liters = 0
        low_stock_count = 0
        negative_stock_count = 0
        active_taps_count = len(active_beers)

        for beer_name, beer_data in beer_stocks.items():
            remaining_liters = beer_data['remaining_liters']

            # НЕ пропускаем активные краны даже с нулевым или отрицательным остатком
            # (чтобы видеть какие кеги нужно пополнить)
            if remaining_liters == 0 and not beer_data.get('on_tap', False):
                continue

            total_liters += remaining_liters if remaining_liters > 0 else 0

            # Определяем уровень остатка
            if remaining_liters < 0:
                stock_level = 'negative'
                negative_stock_count += 1
                low_stock_count += 1
            elif remaining_liters < 10:
                stock_level = 'low'
                low_stock_count += 1
            elif remaining_liters < 25:
                stock_level = 'medium'
            else:
                stock_level = 'high'

            # Получаем номера кранов для этого пива
            tap_numbers = beer_to_taps.get(beer_name, [])
            tap_numbers_str = ', '.join(map(str, sorted(tap_numbers))) if tap_numbers else '—'

            taps_data.append({
                'beer_name': beer_name,
                'category': beer_data['category'],
                'remaining_liters': round(remaining_liters, 1),
                'stock_level': stock_level,
                'on_tap': beer_data.get('on_tap', False),
                'tap_numbers': tap_numbers_str,
                'taps_count': len(tap_numbers)
            })

        # Сортируем по остатку (от меньшего к большему - что заканчивается сверху)
        taps_data.sort(key=lambda x: x['remaining_liters'])

        return jsonify({
            'total_items': len(taps_data),
            'total_liters': round(total_liters, 1),
            'low_stock_count': low_stock_count,
            'negative_stock_count': negative_stock_count,
            'active_taps_count': active_taps_count,
            'taps': taps_data
        })

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/stocks/taplist: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/stocks/kitchen', methods=['GET'])
def get_kitchen_stocks():
    """API endpoint для получения товаров с остатками"""
    try:
        bar = request.args.get('bar', '')

        if not bar:
            return jsonify({'error': 'Требуется параметр bar'}), 400

        # Подключаемся к iiko API
        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        # Получаем складские операции за последние 30 дней
        # API требует формат DD.MM.YYYY
        date_to_obj = datetime.now()
        date_from_obj = datetime.now() - timedelta(days=30)
        date_to = date_to_obj.strftime("%d.%m.%Y")
        date_from = date_from_obj.strftime("%d.%m.%Y")

        try:
            # Получаем номенклатуру товаров (GUID -> название)
            nomenclature = get_cached_nomenclature(olap)

            if not nomenclature:
                return jsonify({'error': 'Не удалось получить номенклатуру товаров'}), 500

            # Получаем отчет по складским операциям с товарами
            bar_name = bar if bar != 'Общая' else None
            store_data = olap.get_store_operations_report(date_from, date_to, bar_name)

            if not store_data:
                return jsonify({'error': 'Не удалось получить данные о товарах'}), 500
        finally:
            olap.disconnect()

        # Обрабатываем данные о товарах
        products_dict = {}

        for record in store_data:
            product_id = record.get('product')
            if not product_id:
                continue

            # Получаем информацию о товаре из номенклатуры
            product_info = nomenclature.get(product_id)
            if not product_info:
                continue  # Товар не найден в номенклатуре

            # Фильтруем товары: исключаем пиво и напитки
            # Пропускаем DISH (блюда - это разливное пиво)
            if product_info.get('type') == 'DISH':
                continue

            # Используем белый список категорий для еды/продуктов
            # Остальное (пиво, напитки) исключаем
            category = product_info.get('category', '') or ''

            # Белый список - только эти категории попадают в стоки
            food_categories = [
                'Метро', 'ООО "Май"', 'ООО "КВГ"', 'ГС Маркет',
                'ИП Тихомиров', 'ООО "Кулинарпродторг"',
                'ИП Новиков', 'ООО МП Арсенал', 'Лента',
                'ООО "МП-Арсенал АО"', 'Криспи',
                'ООО "ВУРСТХАУСМАНУФАКТУР"', 'ООО "Арбореал"'
            ]

            # Если категория не указана или не в белом списке - пропускаем
            if not category or not any(food_cat in category for food_cat in food_categories):
                continue

            # Дополнительно фильтруем по названию (на случай если категория не указана)
            product_name = product_info.get('name', product_id)
            if any(keyword in product_name.lower() for keyword in ['пиво', 'beer', 'ipa', 'лагер', 'эль', 'стаут']):
                continue

            if product_id not in products_dict:
                products_dict[product_id] = {
                    'id': product_id,
                    'name': product_name,
                    'category': category or 'Товары',
                    'incoming': 0,
                    'outgoing': 0,
                    'operations_count': 0
                }

            # Учитываем приход и расход
            amount = float(record.get('amount', 0) or 0)
            is_incoming = record.get('incoming', 'false') == 'true'

            products_dict[product_id]['operations_count'] += 1

            if is_incoming:
                products_dict[product_id]['incoming'] += amount
            else:
                products_dict[product_id]['outgoing'] += abs(amount)

        # Формируем список товаров с остатками
        items = []
        for product_id, product_data in products_dict.items():
            # Рассчитываем текущий остаток
            current_stock = product_data['incoming'] - product_data['outgoing']

            # Рассчитываем средний расход в день
            days_in_period = 30
            avg_consumption = product_data['outgoing'] / days_in_period if days_in_period > 0 else 0

            # Определяем уровень остатков (сколько дней хватит)
            if avg_consumption > 0:
                days_left = current_stock / avg_consumption
                if days_left < 3:
                    stock_level = 'low'
                elif days_left < 7:
                    stock_level = 'medium'
                else:
                    stock_level = 'high'
            else:
                stock_level = 'high'

            items.append({
                'category': product_data['category'],
                'name': product_data['name'],
                'stock': round(current_stock, 1),
                'avg_sales': round(avg_consumption, 2),
                'stock_level': stock_level
            })

        # Сортируем по категориям и названиям
        items.sort(key=lambda x: (x['category'], x['name']))

        low_stock_count = len([item for item in items if item['stock_level'] == 'low'])

        return jsonify({
            'total_items': len(items),
            'low_stock_count': low_stock_count,
            'items': items
        })

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/stocks/kitchen: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/stocks/bottles', methods=['GET'])
def get_bottles_stocks():
    """API endpoint для получения фасованного пива с остатками"""
    try:
        bar = request.args.get('bar', '')

        if not bar:
            return jsonify({'error': 'Требуется параметр bar'}), 400

        # Подключаемся к iiko API
        olap = OlapReports()
        if not olap.connect():
            return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

        # Получаем складские операции за последние 30 дней
        date_to_obj = datetime.now()
        date_from_obj = datetime.now() - timedelta(days=30)
        date_to = date_to_obj.strftime("%d.%m.%Y")
        date_from = date_from_obj.strftime("%d.%m.%Y")

        try:
            # Получаем номенклатуру товаров
            nomenclature = get_cached_nomenclature(olap)

            if not nomenclature:
                return jsonify({'error': 'Не удалось получить номенклатуру товаров'}), 500

            # ID группы "Напитки Фасовка"
            FASOVKA_GROUP_ID = '6103ecbf-e6f8-49fe-8cd2-6102d49e14a6'

            # Получаем все товары из группы "Напитки Фасовка" (включая подгруппы)
            fasovka_product_ids = olap.get_products_in_group(FASOVKA_GROUP_ID, nomenclature)
            print(f"[BOTTLES DEBUG] Товаров в группе 'Напитки Фасовка': {len(fasovka_product_ids)}")

            # Получаем отчет по складским операциям
            bar_name = bar if bar != 'Общая' else None
            store_data = olap.get_store_operations_report(date_from, date_to, bar_name)

            if not store_data:
                return jsonify({'error': 'Не удалось получить данные о товарах'}), 500
        finally:
            olap.disconnect()

        # Обрабатываем данные о фасовке
        products_dict = {}
        checked_products = 0
        matched_products = 0

        for record in store_data:
            product_id = record.get('product')
            if not product_id:
                continue

            # Получаем информацию о товаре из номенклатуры
            product_info = nomenclature.get(product_id)
            if not product_info:
                continue

            checked_products += 1

            # Фильтруем по группе "Напитки Фасовка"
            if product_id not in fasovka_product_ids:
                continue

            product_name = product_info.get('name', product_id)

            matched_products += 1
            supplier = product_info.get('category', 'Без поставщика')

            # Отладочный вывод
            if matched_products <= 5:
                print(f"[DEBUG] Найдена фасовка: {product_name}, категория: {supplier}")

            if product_id not in products_dict:
                products_dict[product_id] = {
                    'id': product_id,
                    'name': product_name,
                    'category': supplier,
                    'incoming': 0,
                    'outgoing': 0,
                    'operations_count': 0
                }

            # Учитываем приход и расход
            amount = float(record.get('amount', 0) or 0)
            is_incoming = record.get('incoming', 'false') == 'true'

            products_dict[product_id]['operations_count'] += 1

            if is_incoming:
                products_dict[product_id]['incoming'] += amount
            else:
                products_dict[product_id]['outgoing'] += abs(amount)

        # Формируем список товаров с остатками
        items = []
        for product_id, product_data in products_dict.items():
            # Рассчитываем текущий остаток
            current_stock = product_data['incoming'] - product_data['outgoing']

            # Рассчитываем средний расход в день
            days_in_period = 30
            avg_consumption = product_data['outgoing'] / days_in_period if days_in_period > 0 else 0

            # Определяем уровень остатков (сколько дней хватит)
            if avg_consumption > 0:
                days_left = current_stock / avg_consumption
                if days_left < 3:
                    stock_level = 'low'
                elif days_left < 7:
                    stock_level = 'medium'
                else:
                    stock_level = 'high'
            else:
                stock_level = 'high'

            items.append({
                'category': product_data['category'],
                'name': product_data['name'],
                'stock': round(current_stock, 1),
                'avg_sales': round(avg_consumption, 2),
                'stock_level': stock_level
            })

        # Сортируем по категориям и названиям
        items.sort(key=lambda x: (x['category'], x['name']))

        low_stock_count = len([item for item in items if item['stock_level'] == 'low'])

        # Отладочная информация
        print(f"\n[BOTTLES DEBUG]")
        print(f"   Проверено товаров: {checked_products}")
        print(f"   Найдено фасовки: {matched_products}")
        print(f"   Итого позиций: {len(items)}")
        print(f"   Требуют пополнения: {low_stock_count}")

        return jsonify({
            'total_items': len(items),
            'low_stock_count': low_stock_count,
            'items': items
        })

    except Exception as e:
        print(f"[ERROR] Oshibka v /api/stocks/bottles: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard-analytics', methods=['POST'])
def dashboard_analytics():
    """API endpoint для получения всех метрик дашборда из сырых OLAP данных"""
    try:
        data = request.json
        venue_key = data.get('bar')  # ключ заведения (kremenchugskaya, bolshoy, etc)
        date_from = data.get('date_from')  # YYYY-MM-DD
        date_to = data.get('date_to')      # YYYY-MM-DD

        if not date_from or not date_to:
            return jsonify({'error': 'Требуются параметры date_from и date_to'}), 400

        # Преобразуем venue_key в название для iiko API
        bar_name = venues_manager.get_iiko_name(venue_key) if venue_key and venue_key != 'all' else None

        # iiko API: date_to не включается (exclusive end)
        # Чтобы получить данные за 10.11-16.11 включительно, нужно запросить до 17.11
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
        date_to_inclusive = (date_to_obj + timedelta(days=1)).strftime('%Y-%m-%d')

        print(f"\n[DASHBOARD] Запуск анализа дашборда...")
        print(f"   Venue key: {venue_key}")
        print(f"   Bar name (iiko): {bar_name if bar_name else 'ВСЕ'}")
        print(f"   Period (requested): {date_from} - {date_to}")
        print(f"   Period (iiko API): {date_from} - {date_to_inclusive}")

        # Проверяем кеш OLAP данных
        cache_key = f"{venue_key}_{date_from}_{date_to_inclusive}"
        now = time.time()

        if cache_key in DASHBOARD_OLAP_CACHE:
            cached_entry = DASHBOARD_OLAP_CACHE[cache_key]
            if (now - cached_entry['timestamp']) < DASHBOARD_OLAP_CACHE_TTL:
                all_sales_data = cached_entry['data']
                ttl_remaining = DASHBOARD_OLAP_CACHE_TTL - (now - cached_entry['timestamp'])
                print(f"   [CACHE] Использую кешированные OLAP данные (истекает через {int(ttl_remaining // 60)} мин {int(ttl_remaining % 60)} сек)")
            else:
                # Кеш устарел, удаляем
                del DASHBOARD_OLAP_CACHE[cache_key]
                all_sales_data = None
        else:
            all_sales_data = None

        # Если данных нет в кеше - запрашиваем из iiko API
        if not all_sales_data:
            # Подключаемся к iiko API
            olap = OlapReports()
            if not olap.connect():
                return jsonify({'error': 'Не удалось подключиться к iiko API'}), 500

            # 1. Получаем ВСЕ данные ОДНИМ запросом (оптимизация: 1 запрос вместо 3)
            print("   [1/5] Запуск комплексного OLAP запроса...")
            start_time = time.time()

            all_sales_data = olap.get_all_sales_report(date_from, date_to_inclusive, bar_name)

            if not all_sales_data or not all_sales_data.get('data'):
                olap.disconnect()
                return jsonify({'error': 'Не удалось получить данные из OLAP'}), 500

            elapsed = time.time() - start_time
            print(f"   [OK] Комплексный запрос выполнен за {elapsed:.2f}s")

            # Отключаемся от iiko API
            olap.disconnect()

            # Сохраняем в кеш
            DASHBOARD_OLAP_CACHE[cache_key] = {
                'data': all_sales_data,
                'timestamp': now
            }
            print(f"   [CACHE] Данные закешированы на {DASHBOARD_OLAP_CACHE_TTL // 60} минут")

        # 2. Создаем калькулятор метрик и рассчитываем из СЫРЫХ данных
        print("   [2/5] Расчет метрик из сырых OLAP данных...")
        calculator = DashboardMetrics()

        # Считаем метрики из единого отчета
        metrics = calculator.calculate_metrics(all_sales_data)
        table_data = calculator.get_table_data(metrics)

        print(f"   [OK] Метрики рассчитаны успешно!")

        # 3. Добавляем метрику активности кранов за период
        print("   [3/5] Расчет активности кранов...")

        # Маппинг venue_key -> bar_id для TapsManager
        VENUE_TO_BAR_MAPPING = {
            'bolshoy': 'bar1',
            'ligovskiy': 'bar2',
            'kremenchugskaya': 'bar3',
            'varshavskaya': 'bar4',
            'all': None  # Для "all" передаем None чтобы считать все бары
        }

        # Преобразуем venue_key в bar_id для TapsManager
        bar_id_for_taps = VENUE_TO_BAR_MAPPING.get(venue_key)
        print(f"   [DEBUG] venue_key={venue_key} -> bar_id_for_taps={bar_id_for_taps}")

        tap_activity = taps_manager.calculate_tap_activity_for_period(bar_id_for_taps, date_from, date_to)
        metrics['tap_activity'] = tap_activity
        print(f"   [OK] Активность кранов: {tap_activity}%")

        # Преобразуем ключи для совместимости с фронтендом
        # Фронтенд ожидает camelCase, бэкенд возвращает snake_case
        frontend_mapping = {
            'total_revenue': 'revenue',
            'total_checks': 'checks',
            'avg_check': 'averageCheck',
            'draft_share': 'draftShare',
            'bottles_share': 'packagedShare',
            'kitchen_share': 'kitchenShare',
            'draft_revenue': 'revenueDraft',
            'bottles_revenue': 'revenuePackaged',
            'kitchen_revenue': 'revenueKitchen',
            'avg_markup': 'markupPercent',
            'total_margin': 'profit',
            'draft_markup': 'markupDraft',
            'bottles_markup': 'markupPackaged',
            'tap_activity': 'tapActivity',
            'kitchen_markup': 'markupKitchen',
            'loyalty_points_written_off': 'loyaltyWriteoffs'
        }

        # Применяем маппинг
        mapped_metrics = {}
        for old_key, new_key in frontend_mapping.items():
            if old_key in metrics:
                value = metrics[old_key]

                # Наценка в API приходит как дробное число (1.95), а в планах как проценты (195)
                # Умножаем на 100 для единообразия
                if new_key in ['markupPercent', 'markupDraft', 'markupPackaged', 'markupKitchen']:
                    value = value * 100

                mapped_metrics[new_key] = value

        # Формируем ответ с преобразованными ключами
        response = {
            **mapped_metrics,
            'table_data': table_data
        }

        return jsonify(response)

    except Exception as e:
        print(f"[ERROR] Ошибка в /api/dashboard-analytics: {e}")
        import traceback
        traceback.print_exc()
        error_detail = f"{type(e).__name__}: {str(e)}"
        return jsonify({'error': error_detail}), 500


# ============================================================================
# DASHBOARD PLANS API - Управление плановыми показателями
# ============================================================================

@app.route('/api/venues')
def get_venues():
    """
    Получить список всех заведений для селектора

    Returns:
        JSON: {
            'venues': [
                {'key': 'all', 'label': '📊 Все заведения', ...},
                {'key': 'bolshoy', 'label': '🍺 Культура - Большой пр. В.О', ...},
                ...
            ]
        }
    """
    try:
        print("\n[VENUES API] Запрос списка заведений...")

        venues = venues_manager.get_all_for_dropdown()

        print(f"[VENUES API] Возвращаю {len(venues)} заведений")

        return jsonify({'venues': venues})

    except Exception as e:
        print(f"[VENUES API ERROR] Ошибка при получении списка заведений: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/weeks')
def get_weeks():
    """
    Получить список всех недель для текущего года

    Returns:
        JSON: {
            'weeks': [...],
            'current_week': {...}
        }
    """
    try:
        print("\n[WEEKS API] Генерация недель для текущего года...")

        current_year = datetime.now().year
        weeks = WeeksGenerator.generate_weeks_for_year(current_year)
        current_week = WeeksGenerator.get_current_week()

        print(f"[WEEKS API] Сгенерировано недель: {len(weeks)}")
        print(f"[WEEKS API] Текущая неделя: {current_week['label']}")

        return jsonify({
            'weeks': weeks,
            'current_week': current_week
        })

    except Exception as e:
        print(f"[WEEKS API ERROR] Ошибка при генерации недель: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans/<venue_key>/<period_key>')
def get_plan_with_venue(venue_key, period_key):
    """
    Получить план за конкретное заведение и период

    URL: /api/plans/{venueKey}/{periodKey}

    Args:
        venue_key: Ключ заведения
        period_key: Ключ периода (YYYY-MM-DD_YYYY-MM-DD)

    Returns:
        JSON: Данные плана или 404
    """
    try:
        print(f"\n[PLANS API] Запрос плана для заведения: {venue_key}, период: {period_key}")

        # Используем составной ключ: venue_period
        composite_key = f"{venue_key}_{period_key}"
        print(f"[PLANS API] Составной ключ: {composite_key}")

        plan = plans_manager.get_plan(composite_key)

        if plan:
            print(f"[PLANS API] План найден")
            return jsonify(plan)
        else:
            print(f"[PLANS API] План НЕ найден")
            return jsonify({'error': 'Plan not found'}), 404

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при получении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans/calculate/<venue_key>/<start_date>/<end_date>')
def calculate_plan_for_period(venue_key, start_date, end_date):
    """
    Рассчитать план для произвольного периода на основе месячных планов

    Пример: /api/plans/calculate/kremenchugskaya/2025-10-01/2025-10-31
    Пример: /api/plans/calculate/kremenchugskaya/2025-10-15/2025-11-15

    Логика:
    - Находит все месяцы в периоде
    - Берёт пропорциональную долю от каждого месячного плана
    - Суммирует абсолютные метрики, усредняет относительные

    Args:
        venue_key: Ключ заведения (bolshoy, ligovskiy, kremenchugskaya, varshavskaya) или пустой для общей
        start_date: Дата начала (YYYY-MM-DD)
        end_date: Дата конца (YYYY-MM-DD)

    Returns:
        JSON: Рассчитанный план или 404 если нет данных
    """
    try:
        print(f"\n[PLANS API] Расчёт плана для {venue_key}, период: {start_date} - {end_date}")

        # Для "общей" передаём пустую строку
        venue = '' if venue_key in ('', 'total', 'общая', 'all') else venue_key

        plan = plans_manager.calculate_plan_for_period(venue, start_date, end_date)

        if plan:
            print(f"[PLANS API] План рассчитан: выручка={plan.get('revenue', 0):.0f}")
            return jsonify(plan)
        else:
            print(f"[PLANS API] Не удалось рассчитать план")
            return jsonify({'error': 'No monthly plans found for this period'}), 404

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при расчёте плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans/migrate-to-monthly', methods=['POST'])
def migrate_plans_to_monthly():
    """
    Миграция недельных планов в месячные (одноразовая)

    POST /api/plans/migrate-to-monthly

    Returns:
        JSON: Результат миграции
    """
    try:
        print("\n[PLANS API] Запуск миграции планов в месячный формат...")

        monthly_plans = plans_manager.import_monthly_plans_from_weekly()

        return jsonify({
            'success': True,
            'message': f'Migrated to {len(monthly_plans)} monthly plans',
            'plans_count': len(monthly_plans),
            'plan_keys': list(monthly_plans.keys())
        })

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка миграции: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans/import-from-excel', methods=['POST'])
def import_plans_from_excel():
    """
    Импорт планов из Excel файла планы_2025_2026.xlsx

    POST /api/plans/import-from-excel

    Returns:
        JSON: Результат импорта с количеством загруженных планов
    """
    try:
        import openpyxl
        from datetime import datetime

        print("\n[PLANS API] Запуск импорта планов из Excel...")

        # Маппинг названий баров
        VENUE_MAPPING = {
            'Большой пр. В.О': 'bolshoy',
            'Лиговский': 'ligovskiy',
            'Кременчугская': 'kremenchugskaya',
            'Варшавская': 'varshavskaya',
            'Общая': 'all'
        }

        # Маппинг метрик
        METRIC_MAPPING = {
            'Выручка': 'revenue',
            'Прибыль': 'profit',
            'Чеки': 'checks',
            'Средний чек': 'averageCheck',
            'Доля розлив': 'draftShare',
            'Доля фасовка': 'packagedShare',
            'Доля кухня': 'kitchenShare',
            'Наценка': 'markupPercent',
            'Наценка розлив': 'markupDraft',
            'Наценка фасовка': 'markupPackaged',
            'Наценка кухня': 'markupKitchen',
            'Списания баллов': 'loyaltyWriteoffs',
            'Активность кранов': 'tapActivity'
        }

        MONTH_NAMES_RU = {
            'Январь': 1, 'Февраль': 2, 'Март': 3, 'Апрель': 4,
            'Май': 5, 'Июнь': 6, 'Июль': 7, 'Август': 8,
            'Сентябрь': 9, 'Октябрь': 10, 'Ноябрь': 11, 'Декабрь': 12
        }

        def parse_month_header(header_text):
            parts = header_text.strip().split()
            if len(parts) != 2:
                return None, None
            month_name, year_str = parts
            try:
                year = int(year_str)
            except ValueError:
                return None, None
            month = MONTH_NAMES_RU.get(month_name)
            return (year, month) if month else (None, None)

        # Читаем Excel
        file_path = 'data/reports/планы_2025_2026.xlsx'
        if not os.path.exists(file_path):
            return jsonify({'error': f'Файл не найден: {file_path}'}), 404

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        all_plans = {}
        current_month_year = None
        bar_columns = {}

        for row in ws.iter_rows(min_row=1, values_only=False):
            first_cell = row[0]
            if first_cell.value and isinstance(first_cell.value, str):
                year, month = parse_month_header(first_cell.value)
                if year and month:
                    current_month_year = (year, month)
                    print(f"[EXCEL] Найден месяц: {first_cell.value}")
                    continue

            if first_cell.value == 'Метрика':
                bar_columns = {}
                for col_idx, cell in enumerate(row[1:], start=2):
                    if cell.value in VENUE_MAPPING:
                        bar_columns[col_idx] = VENUE_MAPPING[cell.value]
                continue

            if first_cell.value in METRIC_MAPPING and current_month_year and bar_columns:
                metric_key = METRIC_MAPPING[first_cell.value]
                year, month = current_month_year

                for col_idx, venue_key in bar_columns.items():
                    cell_value = row[col_idx - 1].value
                    if cell_value is not None and cell_value != '':
                        try:
                            value = float(cell_value)
                            month_key = f'{venue_key}_{year}-{month:02d}'
                            if month_key not in all_plans:
                                all_plans[month_key] = {}
                            all_plans[month_key][metric_key] = value
                        except (ValueError, TypeError):
                            pass

        # Расчёт выручки по категориям
        for plan_data in all_plans.values():
            if 'revenue' in plan_data:
                revenue = plan_data['revenue']
                if 'draftShare' in plan_data:
                    plan_data['revenueDraft'] = revenue * (plan_data['draftShare'] / 100)
                if 'packagedShare' in plan_data:
                    plan_data['revenuePackaged'] = revenue * (plan_data['packagedShare'] / 100)
                if 'kitchenShare' in plan_data:
                    plan_data['revenueKitchen'] = revenue * (plan_data['kitchenShare'] / 100)
                if 'loyaltyWriteoffs' not in plan_data:
                    plan_data['loyaltyWriteoffs'] = revenue * 0.05

        # Сохраняем в JSON
        output_data = {
            'plans': all_plans,
            'metadata': {
                'lastUpdate': datetime.now().isoformat(),
                'version': '1.0',
                'source': 'Excel import from планы_2025_2026.xlsx'
            }
        }

        with open('data/plansdashboard.json', 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)

        # Перезагружаем PlansManager
        plans_manager._initialize_file()

        print(f"[PLANS API] Импортировано планов: {len(all_plans)}")

        return jsonify({
            'success': True,
            'message': f'Импортировано {len(all_plans)} планов из Excel',
            'plans_count': len(all_plans),
            'plan_keys': sorted(all_plans.keys())
        })

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка импорта из Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans/<period_key>')
def get_plan(period_key):
    """
    Получить план за конкретный период (без указания заведения)

    Args:
        period_key: Ключ периода (YYYY-MM-DD_YYYY-MM-DD)

    Returns:
        JSON: Данные плана или 404
    """
    try:
        print(f"\n[PLANS API] Запрос плана для периода: {period_key}")

        plan = plans_manager.get_plan(period_key)

        if plan:
            print(f"[PLANS API] План найден")
            return jsonify(plan)
        else:
            print(f"[PLANS API] План НЕ найден")
            return jsonify({'error': 'Plan not found'}), 404

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при получении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans', methods=['GET'])
def get_all_plans():
    """
    Получить все планы (для функции копирования)

    Returns:
        JSON: {
            'plans': {...},
            'periods': [...]
        }
    """
    try:
        print("\n[PLANS API] Запрос всех планов...")

        all_plans = plans_manager.get_all_plans()
        periods = plans_manager.get_periods_with_plans()

        print(f"[PLANS API] Загружено планов: {len(all_plans)}")

        return jsonify({
            'plans': all_plans,
            'periods': periods
        })

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при получении планов: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans', methods=['POST'])
def save_plan():
    """
    Сохранить или обновить план

    Body:
        {
            'period': 'YYYY-MM-DD_YYYY-MM-DD',
            'data': {
                'revenue': 1000000,
                'checks': 500,
                ...
            }
        }

    Returns:
        JSON: {'success': True} или {'error': '...'}
    """
    try:
        request_data = request.json

        if not request_data:
            return jsonify({'error': 'No data provided'}), 400

        period_key = request_data.get('period')
        plan_data = request_data.get('data')

        if not period_key:
            return jsonify({'error': 'Period key is required'}), 400

        if not plan_data:
            return jsonify({'error': 'Plan data is required'}), 400

        print(f"\n[PLANS API] Сохранение плана для периода: {period_key}")
        print(f"[PLANS API] Данные: {list(plan_data.keys())}")

        # Сохраняем план (валидация внутри PlansManager)
        success = plans_manager.save_plan(period_key, plan_data)

        if success:
            print(f"[PLANS API] План успешно сохранен")
            return jsonify({
                'success': True,
                'message': 'Plan saved successfully',
                'period': period_key
            })
        else:
            return jsonify({'error': 'Failed to save plan'}), 500

    except ValueError as e:
        # Ошибки валидации
        print(f"[PLANS API VALIDATION ERROR] {e}")
        return jsonify({'error': f'Validation error: {str(e)}'}), 400

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при сохранении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans/<venue_key>/<period_key>', methods=['POST'])
def save_plan_with_venue(venue_key, period_key):
    """
    Сохранить или обновить план для конкретного заведения и периода

    URL: /api/plans/{venueKey}/{periodKey}

    Body: объект с данными плана (например: {revenue: 1000, checks: 500, ...})

    Returns:
        JSON: {'success': True} или {'error': '...'}
    """
    try:
        request_data = request.json

        if not request_data:
            return jsonify({'error': 'No data provided'}), 400

        print(f"\n[PLANS API] Сохранение плана для заведения: {venue_key}, период: {period_key}")
        print(f"[PLANS API] Данные: {list(request_data.keys())}")

        # Используем составной ключ: venue_period
        composite_key = f"{venue_key}_{period_key}"
        print(f"[PLANS API] Составной ключ: {composite_key}")

        # Сохраняем план (валидация внутри PlansManager)
        success = plans_manager.save_plan(composite_key, request_data)

        if success:
            print(f"[PLANS API] План успешно сохранен")
            return jsonify({
                'success': True,
                'message': 'Plan saved successfully',
                'period': period_key,
                'venue': venue_key
            })
        else:
            return jsonify({'error': 'Failed to save plan'}), 500

    except ValueError as e:
        # Ошибки валидации
        print(f"[PLANS API VALIDATION ERROR] {e}")
        return jsonify({'error': f'Validation error: {str(e)}'}), 400

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при сохранении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans/<venue_key>/<period_key>', methods=['DELETE'])
def delete_plan_with_venue(venue_key, period_key):
    """
    Удалить план за конкретное заведение и период

    URL: /api/plans/{venueKey}/{periodKey}

    Args:
        venue_key: Ключ заведения
        period_key: Ключ периода

    Returns:
        JSON: {'success': True} или {'error': '...'}
    """
    try:
        print(f"\n[PLANS API] Удаление плана для заведения: {venue_key}, период: {period_key}")

        # Используем составной ключ: venue_period
        composite_key = f"{venue_key}_{period_key}"
        print(f"[PLANS API] Составной ключ: {composite_key}")

        success = plans_manager.delete_plan(composite_key)

        if success:
            print(f"[PLANS API] План удален")
            return jsonify({
                'success': True,
                'message': 'Plan deleted successfully'
            })
        else:
            return jsonify({'error': 'Plan not found'}), 404

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при удалении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans/<period_key>', methods=['DELETE'])
def delete_plan(period_key):
    """
    Удалить план за период (без указания заведения)

    Args:
        period_key: Ключ периода

    Returns:
        JSON: {'success': True} или {'error': '...'}
    """
    try:
        print(f"\n[PLANS API] Удаление плана для периода: {period_key}")

        success = plans_manager.delete_plan(period_key)

        if success:
            print(f"[PLANS API] План удален")
            return jsonify({
                'success': True,
                'message': 'Plan deleted successfully'
            })
        else:
            return jsonify({'error': 'Plan not found'}), 404

    except Exception as e:
        print(f"[PLANS API ERROR] Ошибка при удалении плана: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/comments/<venue_key>/<period_key>', methods=['GET'])
def get_comment(venue_key, period_key):
    """
    Получить комментарий для периода и заведения

    Args:
        venue_key: Ключ заведения
        period_key: Ключ периода

    Returns:
        JSON: {'comment': '...'} или {'comment': null}
    """
    try:
        print(f"\n[COMMENTS API] Получение комментария: {venue_key} / {period_key}")

        plan = plans_manager.get_plan(period_key)

        if plan and 'comment' in plan:
            comment = plan['comment']
            print(f"[COMMENTS API] Комментарий найден: {len(comment)} символов")
            return jsonify({'comment': comment})
        else:
            print(f"[COMMENTS API] Комментарий не найден")
            return jsonify({'comment': None})

    except Exception as e:
        print(f"[COMMENTS API ERROR] Ошибка при получении комментария: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/comments/<venue_key>/<period_key>', methods=['POST'])
def save_comment(venue_key, period_key):
    """
    Сохранить комментарий для периода и заведения

    Args:
        venue_key: Ключ заведения
        period_key: Ключ периода

    Body:
        {'comment': 'текст комментария'}

    Returns:
        JSON: {'success': True} или {'error': '...'}
    """
    try:
        data = request.json
        comment = data.get('comment', '').strip()

        print(f"\n[COMMENTS API] Сохранение комментария: {venue_key} / {period_key}")
        print(f"[COMMENTS API] Текст: {len(comment)} символов")

        # Получаем существующий план или создаем пустой
        plan = plans_manager.get_plan(period_key)
        if plan is None:
            plan = {}

        # Обновляем комментарий
        plan['comment'] = comment

        # Сохраняем план обратно
        success = plans_manager.save_plan(period_key, plan)

        if success:
            print(f"[COMMENTS API] Комментарий сохранен")
            return jsonify({
                'success': True,
                'message': 'Comment saved successfully'
            })
        else:
            return jsonify({'error': 'Failed to save comment'}), 500

    except Exception as e:
        print(f"[COMMENTS API ERROR] Ошибка при сохранении комментария: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# END OF DASHBOARD PLANS API
# ============================================================================


# ========== COMPARISON, TRENDS & EXPORT APIs ==========

# Инициализируем калькуляторы
comparison_calculator = ComparisonCalculator()
trends_analyzer = TrendsAnalyzer()
export_manager = ExportManager()


@app.route('/api/comparison/periods', methods=['POST'])
def compare_periods():
    """API для сравнения двух периодов"""
    try:
        data = request.json
        venue_key = data.get('venue_key')
        period1_key = data.get('period1_key')
        period2_key = data.get('period2_key')

        if not all([venue_key, period1_key, period2_key]):
            return jsonify({'error': 'Missing required parameters'}), 400

        # Получаем данные для обоих периодов
        # TODO: Implement period data fetching
        # Пока возвращаем mock данные
        return jsonify({
            'success': True,
            'comparison': {},
            'insights': []
        })

    except Exception as e:
        print(f"[ERROR] /api/comparison/periods: {e}")
        return jsonify({'error': str(e)}), 500


def get_dashboard_analytics_data(bar, date_from, date_to):
    """
    Helper функция для получения аналитических данных дашборда
    Используется в экспорте и других местах

    Args:
        bar: ключ заведения (kremenchugskaya, bolshoy, etc) или 'all'
        date_from: дата начала периода (YYYY-MM-DD)
        date_to: дата окончания периода (YYYY-MM-DD)

    Returns:
        dict: словарь с метриками в snake_case формате
    """
    # Преобразуем venue_key в название для iiko API
    bar_name = venues_manager.get_iiko_name(bar) if bar and bar != 'all' else None

    # iiko API: date_to не включается (exclusive end)
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
    date_to_inclusive = (date_to_obj + timedelta(days=1)).strftime('%Y-%m-%d')

    # Подключаемся к iiko API
    olap = OlapReports()
    if not olap.connect():
        raise Exception('Не удалось подключиться к iiko API')

    try:
        # Получаем комплексный отчет со всеми данными
        all_sales_data = olap.get_all_sales_report(date_from, date_to_inclusive, bar_name)

        if not all_sales_data or not all_sales_data.get('data'):
            raise Exception('Не удалось получить данные из OLAP')

        # Создаем калькулятор и рассчитываем метрики
        calculator = DashboardMetrics()
        metrics = calculator.calculate_metrics(all_sales_data)

        return metrics

    finally:
        olap.disconnect()


@app.route('/api/export/text', methods=['POST'])
def export_text():
    """API для экспорта в текстовый формат"""
    try:
        data = request.json
        venue_name = data.get('venue_name', 'Unknown')
        period = data.get('period', {})
        comparison = data.get('comparison', {})
        insights = data.get('insights', [])

        text_report = export_manager.prepare_text_report(
            venue_name,
            period,
            comparison,
            insights
        )

        return jsonify({
            'success': True,
            'text': text_report
        })

    except Exception as e:
        print(f"[ERROR] /api/export/text: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """API для экспорта в Excel (XLSX) формат"""
    try:
        from io import BytesIO
        import csv

        data = request.json
        bar = data.get('bar')
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        print(f"\n[EXPORT EXCEL] Генерация Excel: {bar} / {date_from} - {date_to}")

        # Получаем данные
        venue = venues_manager.get_venue(bar)
        venue_name = venue['full_name'] if venue else bar

        # Получаем фактические данные
        actual_data = get_dashboard_analytics_data(bar, date_from, date_to)

        # Получаем плановые данные
        plan_key = f"{bar}_{date_from}"
        plan_data = plans_manager.get_plan(plan_key) or {}

        # Пробуем использовать openpyxl если установлен
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Дашборд"

            # Заголовок
            ws.merge_cells('A1:C1')
            ws['A1'] = f"{venue_name} | {date_from} - {date_to}"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Заголовки таблицы - простая таблица факт/план
            headers = ['Метрика', 'План', 'Факт']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

            # Данные метрик с маппингом между plan и actual
            metrics = [
                ('Выручка (₽)', 'revenue', 'total_revenue'),
                ('Чеки (шт)', 'checks', 'total_checks'),
                ('Средний чек (₽)', 'averageCheck', 'avg_check'),
                ('Списания баллов (₽)', 'loyaltyWriteoffs', 'loyalty_points_written_off'),
                ('Прибыль (₽)', 'profit', 'total_margin'),
                ('% наценки', 'markupPercent', 'avg_markup'),
                ('Доля розлива (%)', 'draftShare', 'draft_share'),
                ('Доля фасовки (%)', 'packagedShare', 'bottles_share'),
                ('Доля кухни (%)', 'kitchenShare', 'kitchen_share'),
            ]

            row = 5
            for metric_name, plan_key, actual_key in metrics:
                plan_value = plan_data.get(plan_key, 0) or 0
                actual_value = actual_data.get(actual_key, 0) or 0

                # Для наценки конвертируем (API возвращает дробь, план в процентах)
                if 'markup' in actual_key.lower():
                    actual_value = actual_value * 100

                ws.cell(row=row, column=1, value=metric_name)
                ws.cell(row=row, column=2, value=round(plan_value, 2))
                ws.cell(row=row, column=3, value=round(actual_value, 2))
                row += 1

            # Сохраняем в BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            print(f"[EXPORT EXCEL] Excel файл сгенерирован (openpyxl)")
            return output.getvalue(), 200, {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename=dashboard_{bar}_{date_from}_{date_to}.xlsx'
            }

        except ImportError:
            # Fallback to CSV if openpyxl not available
            print("[EXPORT EXCEL] openpyxl не установлен, используем CSV")
            output = BytesIO()
            writer = csv.writer(output)

            writer.writerow([f"{venue_name} | {date_from} - {date_to}"])
            writer.writerow([])
            writer.writerow(['Метрика', 'План', 'Факт'])

            metrics = [
                ('Выручка (₽)', 'revenue', 'total_revenue'),
                ('Чеки (шт)', 'checks', 'total_checks'),
                ('Средний чек (₽)', 'averageCheck', 'avg_check'),
                ('Списания баллов (₽)', 'loyaltyWriteoffs', 'loyalty_points_written_off'),
                ('Прибыль (₽)', 'profit', 'total_margin'),
                ('% наценки', 'markupPercent', 'avg_markup'),
                ('Доля розлива (%)', 'draftShare', 'draft_share'),
                ('Доля фасовки (%)', 'packagedShare', 'bottles_share'),
                ('Доля кухни (%)', 'kitchenShare', 'kitchen_share'),
            ]

            for metric_name, plan_key, actual_key in metrics:
                plan_value = plan_data.get(plan_key, 0) or 0
                actual_value = actual_data.get(actual_key, 0) or 0

                # Для наценки конвертируем (API возвращает дробь, план в процентах)
                if 'markup' in actual_key.lower():
                    actual_value = actual_value * 100

                writer.writerow([metric_name, plan_value, actual_value])

            output.seek(0)

            print(f"[EXPORT EXCEL] CSV файл сгенерирован (fallback)")
            return output.getvalue(), 200, {
                'Content-Type': 'text/csv',
                'Content-Disposition': f'attachment; filename=dashboard_{bar}_{date_from}_{date_to}.csv'
            }

    except Exception as e:
        print(f"[ERROR] /api/export/excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    """API для экспорта в PDF формат"""
    try:
        from io import BytesIO

        data = request.json
        bar = data.get('bar')
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        print(f"\n[EXPORT PDF] Генерация PDF: {bar} / {date_from} - {date_to}")

        # Получаем данные
        venue = venues_manager.get_venue(bar)
        venue_name = venue['full_name'] if venue else bar

        # Получаем фактические данные
        actual_data = get_dashboard_analytics_data(bar, date_from, date_to)

        # Получаем плановые данные
        plan_key = f"{bar}_{date_from}"
        plan_data = plans_manager.get_plan(plan_key) or {}

        # Пробуем использовать reportlab если установлен
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)

            elements = []
            styles = getSampleStyleSheet()

            # Заголовок
            title = Paragraph(f"<b>Дашборд - {venue_name}</b>", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))

            subtitle = Paragraph(f"Период: {date_from} - {date_to}", styles['Normal'])
            elements.append(subtitle)
            elements.append(Spacer(1, 20))

            # Таблица метрик
            data_table = [['Метрика', 'План', 'Факт', '% плана', 'Разница']]

            metrics = [
                ('Выручка (₽)', 'revenue', 'total_revenue'),
                ('Чеки (шт)', 'checks', 'total_checks'),
                ('Средний чек (₽)', 'averageCheck', 'avg_check'),
                ('Списания баллов (₽)', 'loyaltyWriteoffs', 'loyalty_points_written_off'),
                ('Прибыль (₽)', 'profit', 'total_margin'),
                ('% наценки', 'markupPercent', 'avg_markup'),
                ('Доля розлива (%)', 'draftShare', 'draft_share'),
                ('Доля фасовки (%)', 'packagedShare', 'bottles_share'),
                ('Доля кухни (%)', 'kitchenShare', 'kitchen_share'),
            ]

            for metric_name, plan_key, actual_key in metrics:
                plan_value = plan_data.get(plan_key, 0) or 0
                actual_value = actual_data.get(actual_key, 0) or 0

                # Для наценки конвертируем (API возвращает дробь, план в процентах)
                if 'markup' in actual_key.lower():
                    actual_value = actual_value * 100

                percent = (actual_value / plan_value * 100) if plan_value > 0 else 0
                diff = actual_value - plan_value
                data_table.append([
                    metric_name,
                    f"{plan_value:,.2f}",
                    f"{actual_value:,.2f}",
                    f"{percent:.1f}%",
                    f"{diff:,.2f}"
                ])

            table = Table(data_table)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(table)

            doc.build(elements)
            output.seek(0)

            print(f"[EXPORT PDF] PDF файл сгенерирован (reportlab)")
            return output.getvalue(), 200, {
                'Content-Type': 'application/pdf',
                'Content-Disposition': f'attachment; filename=dashboard_{bar}_{date_from}_{date_to}.pdf'
            }

        except ImportError:
            # Fallback to HTML if reportlab not available
            print("[EXPORT PDF] reportlab не установлен, используем HTML")

            # Создаем карточки метрик для HTML
            metrics = [
                ('Выручка', 'revenue', 'total_revenue', '₽'),
                ('Чеки', 'checks', 'total_checks', 'шт'),
                ('Средний чек', 'averageCheck', 'avg_check', '₽'),
                ('Списания баллов', 'loyaltyWriteoffs', 'loyalty_points_written_off', '₽'),
                ('Прибыль', 'profit', 'total_margin', '₽'),
                ('Наценка', 'markupPercent', 'avg_markup', '%'),
                ('Доля розлива', 'draftShare', 'draft_share', '%'),
                ('Доля фасовки', 'packagedShare', 'bottles_share', '%'),
                ('Доля кухни', 'kitchenShare', 'kitchen_share', '%'),
            ]

            cards = []
            for name, plan_key, actual_key, unit in metrics:
                plan_val = plan_data.get(plan_key, 0) or 0
                actual_val = actual_data.get(actual_key, 0) or 0

                # Для наценки конвертируем
                if 'markup' in actual_key.lower():
                    actual_val = actual_val * 100

                percent = (actual_val / plan_val * 100) if plan_val > 0 else 0
                diff = actual_val - plan_val

                # Определяем цвет по выполнению плана
                if percent >= 100:
                    color = '#4CAF50'  # зеленый
                elif percent >= 90:
                    color = '#FFC107'  # желтый
                else:
                    color = '#F44336'  # красный

                cards.append(f"""
                <div class="metric-card">
                    <div class="metric-name">{name}</div>
                    <div class="metric-values">
                        <div class="value-row">
                            <span class="label">План:</span>
                            <span class="value">{plan_val:,.2f} {unit}</span>
                        </div>
                        <div class="value-row">
                            <span class="label">Факт:</span>
                            <span class="value" style="color: {color}; font-weight: bold;">{actual_val:,.2f} {unit}</span>
                        </div>
                        <div class="value-row progress-row">
                            <span class="label">Выполнение:</span>
                            <div class="progress-bar">
                                <div class="progress-fill" style="width: {min(percent, 100)}%; background: {color};"></div>
                                <span class="progress-text">{percent:.1f}%</span>
                            </div>
                        </div>
                    </div>
                </div>
                """)

            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Дашборд - {venue_name}</title>
                <style>
                    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
                    body {{
                        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        padding: 30px;
                        min-height: 100vh;
                    }}
                    .container {{
                        max-width: 1200px;
                        margin: 0 auto;
                        background: white;
                        border-radius: 20px;
                        padding: 40px;
                        box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                    }}
                    .header {{
                        text-align: center;
                        margin-bottom: 40px;
                        padding-bottom: 20px;
                        border-bottom: 3px solid #667eea;
                    }}
                    h1 {{
                        color: #333;
                        font-size: 32px;
                        margin-bottom: 10px;
                    }}
                    .period {{
                        color: #666;
                        font-size: 16px;
                    }}
                    .metrics-grid {{
                        display: grid;
                        grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
                        gap: 20px;
                    }}
                    .metric-card {{
                        background: #f8f9fa;
                        border-radius: 12px;
                        padding: 20px;
                        transition: transform 0.2s, box-shadow 0.2s;
                    }}
                    .metric-card:hover {{
                        transform: translateY(-5px);
                        box-shadow: 0 8px 16px rgba(0,0,0,0.1);
                    }}
                    .metric-name {{
                        font-size: 18px;
                        font-weight: 600;
                        color: #333;
                        margin-bottom: 15px;
                    }}
                    .metric-values {{
                        display: flex;
                        flex-direction: column;
                        gap: 10px;
                    }}
                    .value-row {{
                        display: flex;
                        justify-content: space-between;
                        align-items: center;
                        padding: 8px 0;
                    }}
                    .label {{
                        color: #666;
                        font-size: 14px;
                    }}
                    .value {{
                        font-size: 16px;
                        color: #333;
                    }}
                    .progress-row {{
                        flex-direction: column;
                        align-items: stretch;
                        gap: 5px;
                        margin-top: 5px;
                    }}
                    .progress-bar {{
                        width: 100%;
                        height: 24px;
                        background: #e0e0e0;
                        border-radius: 12px;
                        position: relative;
                        overflow: hidden;
                    }}
                    .progress-fill {{
                        height: 100%;
                        border-radius: 12px;
                        transition: width 0.3s ease;
                    }}
                    .progress-text {{
                        position: absolute;
                        top: 50%;
                        left: 50%;
                        transform: translate(-50%, -50%);
                        font-size: 12px;
                        font-weight: 600;
                        color: #333;
                    }}
                    @media print {{
                        body {{ background: white; padding: 0; }}
                        .container {{ box-shadow: none; }}
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>{venue_name}</h1>
                        <div class="period">{date_from} - {date_to}</div>
                    </div>
                    <div class="metrics-grid">
                        {''.join(cards)}
                    </div>
                </div>
            </body>
            </html>
            """

            print(f"[EXPORT PDF] HTML файл сгенерирован (fallback)")
            return html_content.encode('utf-8'), 200, {
                'Content-Type': 'text/html; charset=utf-8',
                'Content-Disposition': f'attachment; filename=dashboard_{bar}_{date_from}_{date_to}.html'
            }

    except Exception as e:
        print(f"[ERROR] /api/export/pdf: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/test_modules')
def test_modules():
    """Route to serve test_modules.html"""
    return render_template('test_modules.html')


@app.route('/api/debug/taps-data')
def debug_taps_data():
    """DEBUG: Показать содержимое файла taps_data.json"""
    try:
        import os
        import json

        info = {
            'file_path': TAPS_DATA_PATH,
            'file_exists': os.path.exists(TAPS_DATA_PATH),
            'file_size': os.path.getsize(TAPS_DATA_PATH) if os.path.exists(TAPS_DATA_PATH) else 0
        }

        if os.path.exists(TAPS_DATA_PATH):
            with open(TAPS_DATA_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # Показываем структуру первого уровня
            info['top_level_keys'] = list(data.keys()) if isinstance(data, dict) else f'Not a dict: {type(data)}'
            info['data_type'] = str(type(data))

            # Показываем полную структуру одного крана
            if isinstance(data, dict) and data:
                first_bar_key = list(data.keys())[0]
                first_bar = data[first_bar_key]

                info['structure'] = {
                    'bar_keys': list(first_bar.keys()) if isinstance(first_bar, dict) else 'not a dict'
                }

                # Если есть поле taps
                if isinstance(first_bar, dict) and 'taps' in first_bar:
                    taps = first_bar['taps']
                    if isinstance(taps, dict) and taps:
                        first_tap_key = list(taps.keys())[0]
                        first_tap = taps[first_tap_key]

                        info['tap_sample'] = {
                            'bar': first_bar_key,
                            'tap_number': first_tap_key,
                            'tap_data': first_tap
                        }

        return jsonify(info)

    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================
# TELEGRAM BOT WEBHOOK ENDPOINTS
# ============================================================

def run_async(coro):
    """Безопасный запуск асинхронной функции из синхронного контекста"""
    try:
        loop = asyncio.get_event_loop()
        if loop.is_closed():
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

    return loop.run_until_complete(coro)


@app.route('/telegram/webhook', methods=['POST'])
def telegram_webhook_handler():
    """
    Webhook endpoint для Telegram бота.
    Telegram отправляет сюда все обновления.
    """
    if not TELEGRAM_BOT_ENABLED:
        return jsonify({'error': 'Telegram bot not enabled'}), 503

    try:
        update_data = request.get_json()
        if not update_data:
            return jsonify({'error': 'No data'}), 400

        # Обрабатываем update асинхронно
        result = run_async(telegram_webhook.process_telegram_update(update_data))

        return jsonify({'ok': result})
    except Exception as e:
        print(f"[TELEGRAM WEBHOOK ERROR] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/telegram/setup-webhook', methods=['POST', 'GET'])
def setup_telegram_webhook():
    """
    Установить webhook URL для Telegram бота.
    Вызывается один раз после деплоя на Render.
    """
    if not TELEGRAM_BOT_ENABLED:
        return jsonify({'error': 'Telegram bot not enabled'}), 503

    try:
        # Получаем базовый URL из запроса или используем Render URL
        if request.method == 'POST' and request.is_json:
            data = request.get_json() or {}
            base_url = data.get('base_url')
        else:
            base_url = request.args.get('base_url')

        if not base_url:
            # Пытаемся определить URL автоматически
            base_url = os.environ.get('RENDER_EXTERNAL_URL')
            if not base_url:
                # Если Render URL не найден, используем URL из запроса
                base_url = request.url_root.rstrip('/')

        webhook_url = f"{base_url}/telegram/webhook"

        result = run_async(telegram_webhook.set_webhook(webhook_url))

        if result:
            return jsonify({
                'success': True,
                'webhook_url': webhook_url,
                'message': 'Webhook установлен успешно'
            })
        else:
            return jsonify({'error': 'Failed to set webhook'}), 500

    except Exception as e:
        print(f"[TELEGRAM SETUP ERROR] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/telegram/webhook-info', methods=['GET'])
def get_telegram_webhook_info():
    """Получить информацию о текущем webhook"""
    if not TELEGRAM_BOT_ENABLED:
        return jsonify({'error': 'Telegram bot not enabled'}), 503

    try:
        info = run_async(telegram_webhook.get_webhook_info())

        if info:
            return jsonify({
                'success': True,
                'webhook_info': info
            })
        else:
            return jsonify({'error': 'Failed to get webhook info'}), 500

    except Exception as e:
        print(f"[TELEGRAM INFO ERROR] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/telegram/delete-webhook', methods=['POST'])
def delete_telegram_webhook():
    """
    Удалить webhook (для переключения обратно на polling режим).
    Используется для локальной разработки.
    """
    if not TELEGRAM_BOT_ENABLED:
        return jsonify({'error': 'Telegram bot not enabled'}), 503

    try:
        result = run_async(telegram_webhook.delete_webhook())

        if result:
            return jsonify({
                'success': True,
                'message': 'Webhook удален. Теперь можно использовать polling режим.'
            })
        else:
            return jsonify({'error': 'Failed to delete webhook'}), 500

    except Exception as e:
        print(f"[TELEGRAM DELETE ERROR] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("\n" + "="*60)
    print("BEER ABC/XYZ ANALYSIS")
    print("="*60)
    print("\nZapusk veb-servera...")
    print("Otkroyte v brauzere: http://localhost:5000")
    print("Test modules: http://localhost:5000/test_modules")
    if TELEGRAM_BOT_ENABLED:
        print("Telegram bot: ENABLED (webhook mode)")
    print("\nDlya ostanovki nazhmite Ctrl+C\n")

    app.run(debug=True, host='0.0.0.0', port=5000)